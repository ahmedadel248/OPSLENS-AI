from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

from fastapi import Body, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles

from src.api import history_store
from src.api import services as api_services
from src.api.schemas import FeedbackRequest, InvestigationRequest


PROJECT_ROOT = Path(__file__).resolve().parents[2]
WEB_DIR = PROJECT_ROOT / "web"
EXPORTS_DIR = PROJECT_ROOT / "data" / "final_incident_reports" / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="OpsLens AI API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    history_store.init_db()


def _safe_name(value: Any, fallback: str = "opslens_report") -> str:
    value = str(value or fallback).strip()
    allowed = [ch if (ch.isalnum() or ch in "-_.") else "_" for ch in value]
    name = "".join(allowed).strip("_")
    return name or fallback



def _dict_or_empty(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _list_or_empty(value: Any) -> list:
    return value if isinstance(value, list) else []


def _payload_to_report(payload: Any) -> Dict[str, Any]:
    """Accept old/new frontend payload shapes and always return a report dict."""
    if isinstance(payload, str):
        try:
            return _payload_to_report(json.loads(payload))
        except Exception:
            raise ValueError("Report payload is a string, not a report JSON object.")

    if isinstance(payload, list):
        for item in payload:
            try:
                parsed = _payload_to_report(item)
                if parsed:
                    return parsed
            except Exception:
                continue
        raise ValueError("Report payload list does not contain a report object.")

    if not isinstance(payload, dict):
        raise ValueError("Report payload must be a JSON object.")

    if isinstance(payload.get("report"), (dict, str, list)):
        return _payload_to_report(payload.get("report"))

    return dict(payload)


def _normalize_report(report: Any) -> Dict[str, Any]:
    report = _payload_to_report(report)

    affected = report.get("affected_resources")
    if not isinstance(affected, dict):
        report["affected_resources"] = {}

    fix = report.get("recommended_fix")
    if isinstance(fix, list):
        report["recommended_fix"] = {
            "strategy": "\n".join(_stringify(item) for item in fix),
            "actions": [],
            "commands": [],
        }
    elif isinstance(fix, str):
        report["recommended_fix"] = {"strategy": fix, "actions": [], "commands": []}
    elif not isinstance(fix, dict):
        report["recommended_fix"] = {"strategy": "", "actions": [], "commands": []}
    else:
        fix = dict(fix)
        fix.setdefault("strategy", "")
        fix.setdefault("actions", [])
        fix.setdefault("commands", fix.get("safe_commands", []))
        if not isinstance(fix.get("actions"), list):
            fix["actions"] = []
        if not isinstance(fix.get("commands"), list):
            fix["commands"] = []
        report["recommended_fix"] = fix

    verification = report.get("verification")
    if isinstance(verification, list):
        report["verification"] = {
            "intent": "\n".join(_stringify(item) for item in verification),
            "commands": [],
        }
    elif isinstance(verification, str):
        report["verification"] = {"intent": verification, "commands": []}
    elif not isinstance(verification, dict):
        report["verification"] = {"intent": "", "commands": []}
    else:
        verification = dict(verification)
        verification.setdefault("intent", "")
        if not isinstance(verification.get("commands"), list):
            verification["commands"] = []
        report["verification"] = verification

    for key in ("agent_reasoning", "additional_findings"):
        if not isinstance(report.get(key), list):
            report[key] = []

    return report

def _report_service(report: Dict[str, Any]) -> str:
    affected = _dict_or_empty(report.get("affected_resources"))
    return (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or "opslens_report"
    )


def _record_id_for_report(report: Dict[str, Any]) -> str:
    """
    Stable ID for history records.

    Do NOT hash the full report because fields like created_at, modified_at,
    json paths, and export paths can change between renders and create duplicates.
    Prefer job_id. Otherwise hash only stable incident identity fields.
    """
    for key in ("job_id", "__opslens_job_id", "investigation_id", "run_id", "record_id", "id"):
        value = report.get(key)
        if value:
            return history_store.safe(str(value))

    if report.get("json_report_path"):
        return history_store.safe(Path(str(report["json_report_path"])).stem)

    affected = report.get("affected_resources") or {}

    stable_source = {
        "title": report.get("title", ""),
        "namespace": affected.get("namespace", ""),
        "service": affected.get("service") or affected.get("service_name") or "",
        "deployment": affected.get("deployment") or affected.get("deployment_name") or "",
        "node": affected.get("node") or affected.get("node_name") or "",
        "summary": report.get("incident_summary", ""),
        "root": report.get("root_cause_story", ""),
        "scenario": report.get("source_scenario", ""),
    }

    raw = json.dumps(stable_source, sort_keys=True, ensure_ascii=False, default=str)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return history_store.safe(digest)


def _export_base_name(report: Dict[str, Any], fallback: str = "opslens_report") -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _safe_name(f"{_report_service(report) or fallback}_{stamp}")[:140]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _report_to_markdown(report: Dict[str, Any]) -> str:
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    lines = [
        f"# {report.get('title') or 'OpsLens Incident Report'}",
        "",
        "## Overview",
        f"- Severity: {report.get('severity', 'unknown')}",
        f"- Confidence: {report.get('confidence', 'unknown')}",
        f"- Namespace: {affected.get('namespace', '')}",
        f"- Node: {affected.get('node', '')}",
        f"- Service: {affected.get('service') or affected.get('service_name') or ''}",
        "",
        "## Incident Summary",
        _stringify(report.get("incident_summary", "")),
        "",
        "## Root Cause",
        _stringify(report.get("root_cause_story", "")),
        "",
        "## Recommended Fix",
        _stringify(fix.get("strategy", fix)),
        "",
        "## Verification",
        _stringify(verification.get("intent", verification)),
    ]
    return "\n".join(lines).strip() + "\n"


def _write_csv_report(report: Dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    rows = [
        ["section", "key", "value"],
        ["overview", "title", report.get("title", "")],
        ["overview", "severity", report.get("severity", "")],
        ["overview", "confidence", report.get("confidence", "")],
        ["affected_resources", "namespace", affected.get("namespace", "")],
        ["affected_resources", "node", affected.get("node", "")],
        ["affected_resources", "service", affected.get("service") or affected.get("service_name") or ""],
        ["affected_resources", "deployment", affected.get("deployment", "")],
        ["incident", "summary", report.get("incident_summary", "")],
        ["incident", "root_cause", report.get("root_cause_story", "")],
        ["recommended_fix", "strategy", fix.get("strategy", "") if isinstance(fix, dict) else _stringify(fix)],
        ["verification", "intent", verification.get("intent", "") if isinstance(verification, dict) else _stringify(verification)],
    ]

    for index, item in enumerate(report.get("agent_reasoning") or [], start=1):
        rows.append(["agent_reasoning", f"row_{index}", _stringify(item)])

    if isinstance(fix, dict):
        for index, item in enumerate(fix.get("actions") or [], start=1):
            rows.append(["recommended_fix.actions", f"action_{index}", _stringify(item)])
        for index, command in enumerate(fix.get("commands") or fix.get("safe_commands") or [], start=1):
            rows.append(["recommended_fix.commands", f"command_{index}", _stringify(command)])

    if isinstance(verification, dict):
        for index, command in enumerate(verification.get("commands") or [], start=1):
            rows.append(["verification.commands", f"command_{index}", _stringify(command)])

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return output_path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    report = _normalize_report(report)
    if not report:
        raise ValueError("No report payload was provided for export.")
    fmt = export_format.lower().strip()
    base = _export_base_name(report, fallback=fallback_name)

    if fmt == "csv":
        return _write_csv_report(report, EXPORTS_DIR / f"{base}.csv")

    if fmt in {"json"}:
        path = EXPORTS_DIR / f"{base}.json"
        path.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        return path

    if fmt in {"md", "markdown"}:
        path = EXPORTS_DIR / f"{base}.md"
        path.write_text(_report_to_markdown(report), encoding="utf-8")
        return path

    if fmt in {"pdf", "xlsx", "excel"}:
        return api_services._export_report(report, fmt, fallback_name=fallback_name)

    raise ValueError(f"Unsupported report format: {export_format}")


def _file_response(path: Path) -> FileResponse:
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {path.name}")
    return FileResponse(path=str(path), filename=path.name)


@app.get("/api/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/api/cluster/nodes")
def api_cluster_nodes() -> Dict[str, Any]:
    try:
        return {"nodes": api_services.list_nodes()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/cluster/namespaces")
def api_cluster_namespaces() -> Dict[str, Any]:
    try:
        return {"namespaces": api_services.list_namespaces()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/cluster/nodes/{node_name}/namespaces")
def api_cluster_namespaces_for_node(node_name: str) -> Dict[str, Any]:
    try:
        return {"namespaces": api_services.list_namespaces_for_node(node_name)}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/scenarios")
def api_scenarios() -> Dict[str, Any]:
    try:
        return {"scenarios": api_services.list_scenarios()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/scenarios/details")
def api_scenario_details() -> Dict[str, Any]:
    try:
        return {"scenarios": api_services.list_scenario_details()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/investigations")
def api_start_investigation(request: InvestigationRequest) -> Dict[str, Any]:
    try:
        return api_services.start_investigation_job(request)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/investigations/{job_id}")
def api_get_investigation(job_id: str) -> Dict[str, Any]:
    try:
        return api_services.get_job(job_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/investigations/{job_id}/report/{export_format}/download")
def api_download_job_report(job_id: str, export_format: str):
    try:
        report = api_services.get_job(job_id).get("report")
        if not report:
            raise FileNotFoundError("No report is available for this job yet.")
        return _file_response(_export_payload(report, export_format, fallback_name=job_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/reports/latest/{export_format}/download")
def api_download_latest_report(export_format: str):
    try:
        history = history_store.list_reports(limit=1)
        if not history:
            raise FileNotFoundError("No reports found.")
        report = history_store.get_report(history[0]["record_id"])
        return _file_response(_export_payload(report, export_format, fallback_name=history[0]["record_id"]))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/export/report/{export_format}")
def export_report_direct(export_format: str, report_payload: Any = Body(...)):
    try:
        return _file_response(_export_payload(report_payload, export_format, fallback_name="report"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/db/reports/save")
def api_save_report(report_payload: Any = Body(...)) -> Dict[str, Any]:
    try:
        report_payload = _normalize_report(report_payload)
        record_id = _record_id_for_report(report_payload)
        return history_store.save_report(report_payload, record_id=record_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history")
def api_db_history(limit: int = 20) -> Dict[str, Any]:
    try:
        return {"records": history_store.list_reports(limit=limit)}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history/{record_id}")
def api_db_history_report(record_id: str) -> Dict[str, Any]:
    try:
        return history_store.get_report(record_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history/{record_id}/report/{export_format}/download")
def api_db_history_export(record_id: str, export_format: str):
    try:
        report = history_store.get_report(record_id)
        return _file_response(_export_payload(report, export_format, fallback_name=record_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/feedback")
def api_feedback(payload: FeedbackRequest) -> Dict[str, Any]:
    try:
        return api_services.save_feedback(payload.model_dump())
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/debug/history")
def api_debug_history() -> Dict[str, Any]:
    try:
        return history_store.debug_state()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# Mount the frontend last so it does not swallow /api routes.


# =========================================================
# OpsLens v5 export endpoints
# Exact format export: pdf / xlsx / csv / json / markdown
# =========================================================

from fastapi import Body
from fastapi.responses import FileResponse, Response, JSONResponse
from pathlib import Path as _OpsPath
from datetime import datetime as _OpsDateTime
import json as _ops_json
import csv as _ops_csv
import re as _ops_re
import tempfile as _ops_tempfile

_OPS_EXPORT_DIR = PROJECT_ROOT / "data" / "final_incident_reports" / "exports"
_OPS_EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _ops_as_dict(payload):
    if payload is None:
        return {}

    if isinstance(payload, dict):
        if isinstance(payload.get("report"), dict):
            return payload["report"]
        if isinstance(payload.get("payload"), dict):
            return payload["payload"]
        return payload

    if isinstance(payload, str):
        try:
            parsed = _ops_json.loads(payload)
            return _ops_as_dict(parsed)
        except Exception:
            return {"title": "OpsLens Report", "incident_summary": payload}

    return {"title": "OpsLens Report", "incident_summary": str(payload)}


def _ops_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        parts = []
        for k, v in value.items():
            txt = _ops_text(v)
            if txt:
                parts.append(f"{k}: {txt}")
        return "\n".join(parts)

    if isinstance(value, list):
        return "\n".join(_ops_text(v) for v in value if _ops_text(v))

    return str(value)


def _ops_safe_name(value):
    value = str(value or "opslens-report").strip()
    value = _ops_re.sub(r"[^a-zA-Z0-9_.-]+", "_", value)
    return value.strip("_")[:80] or "opslens-report"


def _ops_report_title(report):
    return (
        report.get("title")
        or report.get("incident_title")
        or "OpsLens Incident Report"
    )


def _ops_markdown(report):
    report = _ops_as_dict(report)
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    lines = []
    lines.append(f"# {_ops_report_title(report)}")
    lines.append("")
    lines.append(f"**Severity:** {report.get('severity', 'unknown')}")
    lines.append(f"**Confidence:** {report.get('confidence', 'unknown')}")
    lines.append("")

    lines.append("## Affected Resources")
    lines.append(f"- Namespace: {affected.get('namespace', '')}")
    lines.append(f"- Service: {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- Deployment: {affected.get('deployment', '')}")
    lines.append(f"- Node: {affected.get('node', '')}")
    lines.append("")

    lines.append("## Incident Summary")
    lines.append(_ops_text(report.get("incident_summary") or report.get("summary")))
    lines.append("")

    lines.append("## Root Cause")
    lines.append(_ops_text(report.get("root_cause_story") or report.get("root_cause_hypothesis")))
    lines.append("")

    lines.append("## Evidence")
    evidence = report.get("evidence") or report.get("agent_reasoning") or report.get("supporting_signals") or []
    if isinstance(evidence, list):
        for item in evidence:
            lines.append(f"- {_ops_text(item)}")
    else:
        lines.append(_ops_text(evidence))
    lines.append("")

    lines.append("## Recommended Fix")
    if isinstance(fix, dict):
        lines.append(_ops_text(fix.get("strategy") or fix.get("explanation") or fix))
        actions = fix.get("actions") or []
        if actions:
            lines.append("")
            lines.append("### Actions")
            for action in actions:
                lines.append(f"- {_ops_text(action)}")
    else:
        lines.append(_ops_text(fix))
    lines.append("")

    lines.append("## Verification")
    if isinstance(verification, dict):
        lines.append(_ops_text(verification.get("intent") or verification))
        commands = verification.get("commands") or []
        if commands:
            lines.append("")
            lines.append("### Verification Commands")
            for command in commands:
                lines.append(f"- `{command}`")
    else:
        lines.append(_ops_text(verification))

    lines.append("")
    return "\n".join(lines)


def _ops_flat_rows(report):
    report = _ops_as_dict(report)
    affected = report.get("affected_resources") or {}

    return [
        ["title", _ops_report_title(report)],
        ["severity", report.get("severity", "")],
        ["confidence", report.get("confidence", "")],
        ["namespace", affected.get("namespace", "")],
        ["service", affected.get("service") or affected.get("service_name") or ""],
        ["deployment", affected.get("deployment", "")],
        ["node", affected.get("node", "")],
        ["incident_summary", _ops_text(report.get("incident_summary") or report.get("summary"))],
        ["root_cause", _ops_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["recommended_fix", _ops_text(report.get("recommended_fix"))],
        ["verification", _ops_text(report.get("verification"))],
    ]


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    report = _ops_as_dict(report)
    fmt = str(export_format or "pdf").lower().strip()

    if fmt in {"excel", "xls"}:
        fmt = "xlsx"
    if fmt in {"md"}:
        fmt = "markdown"

    title = _ops_report_title(report)
    stamp = _OpsDateTime.now().strftime("%Y%m%d_%H%M%S")
    base = _ops_safe_name(f"{fallback_name}_{stamp}")

    if fmt == "json":
        path = _OPS_EXPORT_DIR / f"{base}.json"
        path.write_text(_ops_json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        return path, "application/json"

    if fmt == "csv":
        path = _OPS_EXPORT_DIR / f"{base}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = _ops_csv.writer(f)
            writer.writerow(["field", "value"])
            writer.writerows(_ops_flat_rows(report))
        return path, "text/csv"

    if fmt == "xlsx":
        from openpyxl import Workbook

        path = _OPS_EXPORT_DIR / f"{base}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Report"
        ws.append(["field", "value"])
        for row in _ops_flat_rows(report):
            ws.append(row)

        for col in ("A", "B"):
            ws.column_dimensions[col].width = 35 if col == "A" else 110

        wb.save(path)
        return path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if fmt == "markdown":
        path = _OPS_EXPORT_DIR / f"{base}.md"
        path.write_text(_ops_markdown(report), encoding="utf-8")
        return path, "text/markdown"

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        path = _OPS_EXPORT_DIR / f"{base}.pdf"
        doc = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(_ops_report_title(report), styles["Title"]))
        story.append(Spacer(1, 12))

        for line in _ops_markdown(report).splitlines():
            clean = line.strip()
            if not clean:
                story.append(Spacer(1, 6))
                continue

            clean = clean.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

            if clean.startswith("# "):
                story.append(Paragraph(clean[2:], styles["Title"]))
            elif clean.startswith("## "):
                story.append(Paragraph(clean[3:], styles["Heading2"]))
            elif clean.startswith("### "):
                story.append(Paragraph(clean[4:], styles["Heading3"]))
            elif clean.startswith("- "):
                story.append(Paragraph("• " + clean[2:], styles["BodyText"]))
            else:
                story.append(Paragraph(clean, styles["BodyText"]))

        doc.build(story)
        return path, "application/pdf"

    raise HTTPException(status_code=400, detail=f"Unsupported export format: {export_format}")


@app.post("/api/v2/export/{export_format}")
async def opslens_v2_export(export_format: str, report_payload=Body(...)):
    # Keep v2 route for compatibility, but use the clean export pipeline.
    report = _normalize_report(_ops_as_dict(report_payload))
    affected = report.get("affected_resources") or {}
    fallback = (
        affected.get("service")
        or affected.get("service_name")
        or report.get("record_id")
        or "opslens-report"
    )

    return _file_response(_export_payload(report, export_format, fallback_name=fallback))


@app.get("/api/v2/history/{record_id}/download/{export_format}")
def opslens_v2_history_download(record_id: str, export_format: str):
    # Keep v2 route for compatibility, but use the clean export pipeline.
    report = history_store.get_report(record_id)
    return _file_response(_export_payload(report, export_format, fallback_name=record_id))


if WEB_DIR.exists():
    app.mount("/", StaticFiles(directory=str(WEB_DIR), html=True), name="web")

# =========================================================
# OpsLens final robust export override v4
# =========================================================
# OpsLens final robust export override v4
# Single clean export pipeline for:
# /api/export/report/{format}
# /api/db/history/{record_id}/report/{format}/download
# /api/v2/export/{format}
# /api/v2/history/{record_id}/download/{format}
# =========================================================

def _opslens_export_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        # Prefer command text when dict is a command object.
        if "command" in value:
            title = value.get("title") or value.get("name") or ""
            cmd = value.get("command") or ""
            return f"{title}: {cmd}".strip(": ")

        parts = []
        for key, val in value.items():
            txt = _opslens_export_text(val)
            if txt:
                parts.append(f"{key}: {txt}")
        return "\n".join(parts)

    if isinstance(value, (list, tuple, set)):
        return "\n".join(_opslens_export_text(item) for item in value if _opslens_export_text(item))

    return str(value)


def _opslens_export_safe_name(value, fallback="opslens-report"):
    import re

    value = str(value or fallback).strip()
    value = re.sub(r"[^a-zA-Z0-9_.-]+", "_", value)
    value = value.strip("_")
    return value[:120] or fallback


def _opslens_export_format(export_format):
    fmt = str(export_format or "pdf").lower().strip()

    if fmt in {"excel", "xls"}:
        return "xlsx"

    if fmt in {"md"}:
        return "markdown"

    return fmt


def _opslens_export_media_type(export_format):
    fmt = _opslens_export_format(export_format)

    return {
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "json": "application/json",
        "markdown": "text/markdown",
    }.get(fmt, "application/octet-stream")


def _opslens_export_report(report):
    try:
        report = _normalize_report(report)
    except Exception:
        try:
            report = _ops_as_dict(report)
        except Exception:
            report = report if isinstance(report, dict) else {}

    if not isinstance(report, dict):
        report = {}

    if isinstance(report.get("report"), dict):
        report = report["report"]

    return report


def _opslens_export_title(report):
    return (
        report.get("title")
        or report.get("incident_title")
        or "OpsLens Incident Report"
    )


def _opslens_export_affected(report):
    affected = report.get("affected_resources")
    return affected if isinstance(affected, dict) else {}


def _opslens_export_fix(report):
    fix = report.get("recommended_fix")
    return fix if isinstance(fix, dict) else {"strategy": _opslens_export_text(fix), "actions": [], "commands": []}


def _opslens_export_verification(report):
    verification = report.get("verification")
    return verification if isinstance(verification, dict) else {"intent": _opslens_export_text(verification), "commands": []}


def _opslens_export_commands(report):
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    rows = []

    for item in fix.get("commands") or fix.get("safe_commands") or []:
        rows.append(["Remediation", _opslens_export_text(item)])

    for item in verification.get("commands") or verification.get("verification_commands") or []:
        rows.append(["Verification", _opslens_export_text(item)])

    return rows


def _opslens_export_base(report, fallback_name="report"):
    from datetime import datetime

    affected = _opslens_export_affected(report)
    service = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("deployment")
        or report.get("record_id")
        or fallback_name
        or "opslens-report"
    )

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _opslens_export_safe_name(f"{service}_{stamp}")


def _opslens_export_markdown(report):
    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    lines = [
        f"# {_opslens_export_title(report)}",
        "",
        "## Overview",
        f"- Severity: {report.get('severity', 'unknown')}",
        f"- Confidence: {report.get('confidence', 'unknown')}",
        f"- Namespace: {affected.get('namespace', '')}",
        f"- Node: {affected.get('node') or affected.get('node_name') or ''}",
        f"- Service: {affected.get('service') or affected.get('service_name') or ''}",
        f"- Deployment: {affected.get('deployment') or affected.get('deployment_name') or ''}",
        "",
        "## Incident Summary",
        _opslens_export_text(report.get("incident_summary") or report.get("summary")),
        "",
        "## Root Cause",
        _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis")),
        "",
        "## Evidence",
    ]

    evidence = report.get("agent_reasoning") or report.get("evidence_trail") or report.get("evidence") or []

    if evidence:
        for item in evidence:
            lines.append(f"- {_opslens_export_text(item)}")
    else:
        lines.append("- No evidence rows available.")

    additional = report.get("additional_findings") or []

    if additional:
        lines.extend(["", "## Additional Findings"])
        for item in additional:
            lines.append(f"- {_opslens_export_text(item)}")

    lines.extend([
        "",
        "## Recommended Fix",
        _opslens_export_text(fix.get("strategy")),
        "",
        "## Commands",
    ])

    commands = _opslens_export_commands(report)

    if commands:
        for kind, command in commands:
            lines.append(f"- {kind}: `{command}`")
    else:
        lines.append("- No commands available.")

    lines.extend([
        "",
        "## Verification",
        _opslens_export_text(verification.get("intent")),
        "",
    ])

    return "\n".join(lines).strip() + "\n"


def _opslens_write_json(report, path):
    import json

    path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    return path


def _opslens_write_csv(report, path):
    import csv

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    rows = [
        ["section", "key", "value"],
        ["overview", "title", _opslens_export_title(report)],
        ["overview", "severity", report.get("severity", "")],
        ["overview", "confidence", report.get("confidence", "")],
        ["affected_resources", "namespace", affected.get("namespace", "")],
        ["affected_resources", "node", affected.get("node") or affected.get("node_name") or ""],
        ["affected_resources", "service", affected.get("service") or affected.get("service_name") or ""],
        ["affected_resources", "deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["incident", "summary", _opslens_export_text(report.get("incident_summary") or report.get("summary"))],
        ["incident", "root_cause", _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["recommended_fix", "strategy", _opslens_export_text(fix.get("strategy"))],
        ["verification", "intent", _opslens_export_text(verification.get("intent"))],
    ]

    for index, item in enumerate(report.get("agent_reasoning") or [], start=1):
        rows.append(["evidence", f"row_{index}", _opslens_export_text(item)])

    for index, item in enumerate(report.get("additional_findings") or [], start=1):
        rows.append(["additional_findings", f"row_{index}", _opslens_export_text(item)])

    for index, row in enumerate(_opslens_export_commands(report), start=1):
        rows.append(["commands", f"{row[0]}_{index}", row[1]])

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return path


def _opslens_write_xlsx(report, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def style(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="111827")

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 38 if col == 1 else 95

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    rows = [
        ["Title", _opslens_export_title(report)],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Incident Summary", _opslens_export_text(report.get("incident_summary") or report.get("summary"))],
        ["Root Cause", _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["Fix Strategy", _opslens_export_text(fix.get("strategy"))],
        ["Verification", _opslens_export_text(verification.get("intent"))],
    ]

    for row in rows:
        ws.append(row)
    style(ws)

    ws = wb.create_sheet("Evidence")
    ws.append(["Agent", "Finding", "Meaning"])

    for item in report.get("agent_reasoning") or []:
        if isinstance(item, dict):
            ws.append([
                _opslens_export_text(item.get("agent")),
                _opslens_export_text(item.get("finding")),
                _opslens_export_text(item.get("meaning")),
            ])
        else:
            ws.append(["Evidence", _opslens_export_text(item), ""])
    style(ws)

    ws = wb.create_sheet("Additional Findings")
    ws.append(["Resource", "Finding", "Impact", "Priority"])

    for item in report.get("additional_findings") or []:
        if isinstance(item, dict):
            ws.append([
                _opslens_export_text(item.get("resource")),
                _opslens_export_text(item.get("finding")),
                _opslens_export_text(item.get("impact")),
                _opslens_export_text(item.get("priority")),
            ])
        else:
            ws.append(["", _opslens_export_text(item), "", ""])
    style(ws)

    ws = wb.create_sheet("Commands")
    ws.append(["Type", "Command"])

    for row in _opslens_export_commands(report):
        ws.append(row)
    style(ws)

    wb.save(path)
    return path


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensTitleFinal",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=21,
        spaceAfter=10,
    )

    heading_style = ParagraphStyle(
        "OpsLensHeadingFinal",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=5,
    )

    body_style = ParagraphStyle(
        "OpsLensBodyFinal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "OpsLensCellFinal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.8,
        leading=9.5,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsLensHeaderFinal",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.5,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensCodeFinal",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.2,
        leading=8.8,
        wordWrap="CJK",
    )

    story = []

    def clean(value):
        return html.escape(_opslens_export_text(value)).replace("\n", "<br/>")

    def p(value, style=body_style):
        story.append(Paragraph(clean(value), style))

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(clean(value), heading_style))

    def table(headers, rows, widths=None):
        if not rows:
            p("No data available.")
            return

        usable = A4[0] - (2.5 * cm)

        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(clean(head), header_style) for head in headers]]

        for row in rows:
            data.append([Paragraph(clean(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(clean(_opslens_export_title(report)), title_style))

    overview_rows = [
        ["Severity", report.get("severity", "unknown")],
        ["Confidence", report.get("confidence", "unknown")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
    ]

    table(["Field", "Value"], overview_rows, [4.0 * cm, 13.0 * cm])

    heading("Incident Summary")
    p(report.get("incident_summary") or report.get("summary"))

    heading("Root Cause")
    p(report.get("root_cause_story") or report.get("root_cause_hypothesis"))

    evidence_rows = []

    for item in report.get("agent_reasoning") or []:
        if isinstance(item, dict):
            evidence_rows.append([
                item.get("agent", ""),
                item.get("finding", ""),
                item.get("meaning", ""),
            ])
        else:
            evidence_rows.append(["Evidence", _opslens_export_text(item), ""])

    heading("Evidence")
    table(["Agent", "Finding", "Meaning"], evidence_rows, [3.4 * cm, 5.4 * cm, 8.2 * cm])

    additional_rows = []

    for item in report.get("additional_findings") or []:
        if isinstance(item, dict):
            additional_rows.append([
                item.get("resource", ""),
                item.get("finding", ""),
                item.get("impact", ""),
                item.get("priority", ""),
            ])
        else:
            additional_rows.append(["", _opslens_export_text(item), "", ""])

    heading("Additional Findings")
    table(["Resource", "Finding", "Impact", "Priority"], additional_rows, [3.8 * cm, 4.8 * cm, 6.0 * cm, 2.4 * cm])

    heading("Recommended Fix")
    p(fix.get("strategy", ""))

    heading("Verification")
    p(verification.get("intent", ""))

    command_rows = _opslens_export_commands(report)

    heading("Commands")

    if command_rows:
        for index, row in enumerate(command_rows, start=1):
            p(f"{row[0]} Command {index}")
            story.append(Paragraph(clean(row[1]), code_style))
            story.append(Spacer(1, 5))
    else:
        p("No commands available.")

    doc.build(story)
    return path


def _opslens_write_export_file(report, export_format, fallback_name="report"):
    fmt = _opslens_export_format(export_format)
    report = _opslens_export_report(report)

    if not report:
        raise ValueError("No report payload was provided for export.")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    base = _opslens_export_base(report, fallback_name=fallback_name)

    if fmt == "json":
        return _opslens_write_json(report, EXPORTS_DIR / f"{base}.json")

    if fmt == "csv":
        return _opslens_write_csv(report, EXPORTS_DIR / f"{base}.csv")

    if fmt == "markdown":
        path = EXPORTS_DIR / f"{base}.md"
        path.write_text(_opslens_export_markdown(report), encoding="utf-8")
        return path

    if fmt == "xlsx":
        return _opslens_write_xlsx(report, EXPORTS_DIR / f"{base}.xlsx")

    if fmt == "pdf":
        return _opslens_write_pdf(report, EXPORTS_DIR / f"{base}.pdf")

    raise ValueError(f"Unsupported report format: {export_format}")


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(report, export_format, fallback_name=fallback_name)
    return path, _opslens_export_media_type(export_format)
