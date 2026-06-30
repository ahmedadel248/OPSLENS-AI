from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

from fastapi import Body, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles

from src.api import history_store
from src.api import services as api_services
from src.api.schemas import FeedbackRequest, InvestigationRequest


PROJECT_ROOT = Path(__file__).resolve().parents[2]
WEB_DIR = PROJECT_ROOT / "web"
EXPORTS_DIR = PROJECT_ROOT / "data" / "final_incident_reports" / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="OpsLens AI API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    history_store.init_db()


def _safe_name(value: Any, fallback: str = "opslens_report") -> str:
    value = str(value or fallback).strip()
    allowed = [ch if (ch.isalnum() or ch in "-_.") else "_" for ch in value]
    name = "".join(allowed).strip("_")
    return name or fallback



def _dict_or_empty(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _list_or_empty(value: Any) -> list:
    return value if isinstance(value, list) else []


def _payload_to_report(payload: Any) -> Dict[str, Any]:
    """Accept old/new frontend payload shapes and always return a report dict."""
    if isinstance(payload, str):
        try:
            return _payload_to_report(json.loads(payload))
        except Exception:
            raise ValueError("Report payload is a string, not a report JSON object.")

    if isinstance(payload, list):
        for item in payload:
            try:
                parsed = _payload_to_report(item)
                if parsed:
                    return parsed
            except Exception:
                continue
        raise ValueError("Report payload list does not contain a report object.")

    if not isinstance(payload, dict):
        raise ValueError("Report payload must be a JSON object.")

    if isinstance(payload.get("report"), (dict, str, list)):
        return _payload_to_report(payload.get("report"))

    return dict(payload)


def _normalize_report(report: Any) -> Dict[str, Any]:
    report = _payload_to_report(report)

    affected = report.get("affected_resources")
    if not isinstance(affected, dict):
        report["affected_resources"] = {}

    fix = report.get("recommended_fix")
    if isinstance(fix, list):
        report["recommended_fix"] = {
            "strategy": "\n".join(_stringify(item) for item in fix),
            "actions": [],
            "commands": [],
        }
    elif isinstance(fix, str):
        report["recommended_fix"] = {"strategy": fix, "actions": [], "commands": []}
    elif not isinstance(fix, dict):
        report["recommended_fix"] = {"strategy": "", "actions": [], "commands": []}
    else:
        fix = dict(fix)
        fix.setdefault("strategy", "")
        fix.setdefault("actions", [])
        fix.setdefault("commands", fix.get("safe_commands", []))
        if not isinstance(fix.get("actions"), list):
            fix["actions"] = []
        if not isinstance(fix.get("commands"), list):
            fix["commands"] = []
        report["recommended_fix"] = fix

    verification = report.get("verification")
    if isinstance(verification, list):
        report["verification"] = {
            "intent": "\n".join(_stringify(item) for item in verification),
            "commands": [],
        }
    elif isinstance(verification, str):
        report["verification"] = {"intent": verification, "commands": []}
    elif not isinstance(verification, dict):
        report["verification"] = {"intent": "", "commands": []}
    else:
        verification = dict(verification)
        verification.setdefault("intent", "")
        if not isinstance(verification.get("commands"), list):
            verification["commands"] = []
        report["verification"] = verification

    for key in ("agent_reasoning", "additional_findings"):
        if not isinstance(report.get(key), list):
            report[key] = []

    return report

def _report_service(report: Dict[str, Any]) -> str:
    affected = _dict_or_empty(report.get("affected_resources"))
    return (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or "opslens_report"
    )


def _record_id_for_report(report: Dict[str, Any]) -> str:
    """
    Stable ID for history records.

    Do NOT hash the full report because fields like created_at, modified_at,
    json paths, and export paths can change between renders and create duplicates.
    Prefer job_id. Otherwise hash only stable incident identity fields.
    """
    for key in ("job_id", "__opslens_job_id", "investigation_id", "run_id", "record_id", "id"):
        value = report.get(key)
        if value:
            return history_store.safe(str(value))

    if report.get("json_report_path"):
        return history_store.safe(Path(str(report["json_report_path"])).stem)

    affected = report.get("affected_resources") or {}

    stable_source = {
        "title": report.get("title", ""),
        "namespace": affected.get("namespace", ""),
        "service": affected.get("service") or affected.get("service_name") or "",
        "deployment": affected.get("deployment") or affected.get("deployment_name") or "",
        "node": affected.get("node") or affected.get("node_name") or "",
        "summary": report.get("incident_summary", ""),
        "root": report.get("root_cause_story", ""),
        "scenario": report.get("source_scenario", ""),
    }

    raw = json.dumps(stable_source, sort_keys=True, ensure_ascii=False, default=str)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return history_store.safe(digest)


def _export_base_name(report: Dict[str, Any], fallback: str = "opslens_report") -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _safe_name(f"{_report_service(report) or fallback}_{stamp}")[:140]


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _report_to_markdown(report: Dict[str, Any]) -> str:
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    lines = [
        f"# {report.get('title') or 'OpsLens Incident Report'}",
        "",
        "## Overview",
        f"- Severity: {report.get('severity', 'unknown')}",
        f"- Confidence: {report.get('confidence', 'unknown')}",
        f"- Namespace: {affected.get('namespace', '')}",
        f"- Node: {affected.get('node', '')}",
        f"- Service: {affected.get('service') or affected.get('service_name') or ''}",
        "",
        "## Incident Summary",
        _stringify(report.get("incident_summary", "")),
        "",
        "## Root Cause",
        _stringify(report.get("root_cause_story", "")),
        "",
        "## Recommended Fix",
        _stringify(fix.get("strategy", fix)),
        "",
        "## Verification",
        _stringify(verification.get("intent", verification)),
    ]
    return "\n".join(lines).strip() + "\n"


def _write_csv_report(report: Dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    rows = [
        ["section", "key", "value"],
        ["overview", "title", report.get("title", "")],
        ["overview", "severity", report.get("severity", "")],
        ["overview", "confidence", report.get("confidence", "")],
        ["affected_resources", "namespace", affected.get("namespace", "")],
        ["affected_resources", "node", affected.get("node", "")],
        ["affected_resources", "service", affected.get("service") or affected.get("service_name") or ""],
        ["affected_resources", "deployment", affected.get("deployment", "")],
        ["incident", "summary", report.get("incident_summary", "")],
        ["incident", "root_cause", report.get("root_cause_story", "")],
        ["recommended_fix", "strategy", fix.get("strategy", "") if isinstance(fix, dict) else _stringify(fix)],
        ["verification", "intent", verification.get("intent", "") if isinstance(verification, dict) else _stringify(verification)],
    ]

    for index, item in enumerate(report.get("agent_reasoning") or [], start=1):
        rows.append(["agent_reasoning", f"row_{index}", _stringify(item)])

    if isinstance(fix, dict):
        for index, item in enumerate(fix.get("actions") or [], start=1):
            rows.append(["recommended_fix.actions", f"action_{index}", _stringify(item)])
        for index, command in enumerate(fix.get("commands") or fix.get("safe_commands") or [], start=1):
            rows.append(["recommended_fix.commands", f"command_{index}", _stringify(command)])

    if isinstance(verification, dict):
        for index, command in enumerate(verification.get("commands") or [], start=1):
            rows.append(["verification.commands", f"command_{index}", _stringify(command)])

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return output_path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    report = _normalize_report(report)
    if not report:
        raise ValueError("No report payload was provided for export.")
    fmt = export_format.lower().strip()
    base = _export_base_name(report, fallback=fallback_name)

    if fmt == "csv":
        return _write_csv_report(report, EXPORTS_DIR / f"{base}.csv")

    if fmt in {"json"}:
        path = EXPORTS_DIR / f"{base}.json"
        path.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        return path

    if fmt in {"md", "markdown"}:
        path = EXPORTS_DIR / f"{base}.md"
        path.write_text(_report_to_markdown(report), encoding="utf-8")
        return path

    if fmt in {"pdf", "xlsx", "excel"}:
        return api_services._export_report(report, fmt, fallback_name=fallback_name)

    raise ValueError(f"Unsupported report format: {export_format}")


def _file_response(path: Path) -> FileResponse:
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {path.name}")
    return FileResponse(path=str(path), filename=path.name)


@app.get("/api/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/api/cluster/nodes")
def api_cluster_nodes() -> Dict[str, Any]:
    try:
        return {"nodes": api_services.list_nodes()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/cluster/namespaces")
def api_cluster_namespaces() -> Dict[str, Any]:
    try:
        return {"namespaces": api_services.list_namespaces()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/cluster/nodes/{node_name}/namespaces")
def api_cluster_namespaces_for_node(node_name: str) -> Dict[str, Any]:
    try:
        return {"namespaces": api_services.list_namespaces_for_node(node_name)}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/scenarios")
def api_scenarios() -> Dict[str, Any]:
    try:
        return {"scenarios": api_services.list_scenarios()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/scenarios/details")
def api_scenario_details() -> Dict[str, Any]:
    try:
        return {"scenarios": api_services.list_scenario_details()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/investigations")
def api_start_investigation(request: InvestigationRequest) -> Dict[str, Any]:
    try:
        return api_services.start_investigation_job(request)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/investigations/{job_id}")
def api_get_investigation(job_id: str) -> Dict[str, Any]:
    try:
        return api_services.get_job(job_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/investigations/{job_id}/report/{export_format}/download")
def api_download_job_report(job_id: str, export_format: str):
    try:
        report = api_services.get_job(job_id).get("report")
        if not report:
            raise FileNotFoundError("No report is available for this job yet.")
        return _file_response(_export_payload(report, export_format, fallback_name=job_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/reports/latest/{export_format}/download")
def api_download_latest_report(export_format: str):
    try:
        history = history_store.list_reports(limit=1)
        if not history:
            raise FileNotFoundError("No reports found.")
        report = history_store.get_report(history[0]["record_id"])
        return _file_response(_export_payload(report, export_format, fallback_name=history[0]["record_id"]))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/export/report/{export_format}")
def export_report_direct(export_format: str, report_payload: Any = Body(...)):
    try:
        return _file_response(_export_payload(report_payload, export_format, fallback_name="report"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/db/reports/save")
def api_save_report(report_payload: Any = Body(...)) -> Dict[str, Any]:
    try:
        report_payload = _normalize_report(report_payload)
        record_id = _record_id_for_report(report_payload)
        return history_store.save_report(report_payload, record_id=record_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history")
def api_db_history(limit: int = 20) -> Dict[str, Any]:
    try:
        return {"records": history_store.list_reports(limit=limit)}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history/{record_id}")
def api_db_history_report(record_id: str) -> Dict[str, Any]:
    try:
        return history_store.get_report(record_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/db/history/{record_id}/report/{export_format}/download")
def api_db_history_export(record_id: str, export_format: str):
    try:
        report = history_store.get_report(record_id)
        return _file_response(_export_payload(report, export_format, fallback_name=record_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.post("/api/feedback")
def api_feedback(payload: FeedbackRequest) -> Dict[str, Any]:
    try:
        return api_services.save_feedback(payload.model_dump())
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/api/debug/history")
def api_debug_history() -> Dict[str, Any]:
    try:
        return history_store.debug_state()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


# Mount the frontend last so it does not swallow /api routes.


# =========================================================
# OpsLens v5 export endpoints
# Exact format export: pdf / xlsx / csv / json / markdown
# =========================================================

from fastapi import Body
from fastapi.responses import FileResponse, Response, JSONResponse
from pathlib import Path as _OpsPath
from datetime import datetime as _OpsDateTime
import json as _ops_json
import csv as _ops_csv
import re as _ops_re
import tempfile as _ops_tempfile

_OPS_EXPORT_DIR = PROJECT_ROOT / "data" / "final_incident_reports" / "exports"
_OPS_EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _ops_as_dict(payload):
    if payload is None:
        return {}

    if isinstance(payload, dict):
        if isinstance(payload.get("report"), dict):
            return payload["report"]
        if isinstance(payload.get("payload"), dict):
            return payload["payload"]
        return payload

    if isinstance(payload, str):
        try:
            parsed = _ops_json.loads(payload)
            return _ops_as_dict(parsed)
        except Exception:
            return {"title": "OpsLens Report", "incident_summary": payload}

    return {"title": "OpsLens Report", "incident_summary": str(payload)}


def _ops_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        parts = []
        for k, v in value.items():
            txt = _ops_text(v)
            if txt:
                parts.append(f"{k}: {txt}")
        return "\n".join(parts)

    if isinstance(value, list):
        return "\n".join(_ops_text(v) for v in value if _ops_text(v))

    return str(value)


def _ops_safe_name(value):
    value = str(value or "opslens-report").strip()
    value = _ops_re.sub(r"[^a-zA-Z0-9_.-]+", "_", value)
    return value.strip("_")[:80] or "opslens-report"


def _ops_report_title(report):
    return (
        report.get("title")
        or report.get("incident_title")
        or "OpsLens Incident Report"
    )


def _ops_markdown(report):
    report = _ops_as_dict(report)
    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    lines = []
    lines.append(f"# {_ops_report_title(report)}")
    lines.append("")
    lines.append(f"**Severity:** {report.get('severity', 'unknown')}")
    lines.append(f"**Confidence:** {report.get('confidence', 'unknown')}")
    lines.append("")

    lines.append("## Affected Resources")
    lines.append(f"- Namespace: {affected.get('namespace', '')}")
    lines.append(f"- Service: {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- Deployment: {affected.get('deployment', '')}")
    lines.append(f"- Node: {affected.get('node', '')}")
    lines.append("")

    lines.append("## Incident Summary")
    lines.append(_ops_text(report.get("incident_summary") or report.get("summary")))
    lines.append("")

    lines.append("## Root Cause")
    lines.append(_ops_text(report.get("root_cause_story") or report.get("root_cause_hypothesis")))
    lines.append("")

    lines.append("## Evidence")
    evidence = report.get("evidence") or report.get("agent_reasoning") or report.get("supporting_signals") or []
    if isinstance(evidence, list):
        for item in evidence:
            lines.append(f"- {_ops_text(item)}")
    else:
        lines.append(_ops_text(evidence))
    lines.append("")

    lines.append("## Recommended Fix")
    if isinstance(fix, dict):
        lines.append(_ops_text(fix.get("strategy") or fix.get("explanation") or fix))
        actions = fix.get("actions") or []
        if actions:
            lines.append("")
            lines.append("### Actions")
            for action in actions:
                lines.append(f"- {_ops_text(action)}")
    else:
        lines.append(_ops_text(fix))
    lines.append("")

    lines.append("## Verification")
    if isinstance(verification, dict):
        lines.append(_ops_text(verification.get("intent") or verification))
        commands = verification.get("commands") or []
        if commands:
            lines.append("")
            lines.append("### Verification Commands")
            for command in commands:
                lines.append(f"- `{command}`")
    else:
        lines.append(_ops_text(verification))

    lines.append("")
    return "\n".join(lines)


def _ops_flat_rows(report):
    report = _ops_as_dict(report)
    affected = report.get("affected_resources") or {}

    return [
        ["title", _ops_report_title(report)],
        ["severity", report.get("severity", "")],
        ["confidence", report.get("confidence", "")],
        ["namespace", affected.get("namespace", "")],
        ["service", affected.get("service") or affected.get("service_name") or ""],
        ["deployment", affected.get("deployment", "")],
        ["node", affected.get("node", "")],
        ["incident_summary", _ops_text(report.get("incident_summary") or report.get("summary"))],
        ["root_cause", _ops_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["recommended_fix", _ops_text(report.get("recommended_fix"))],
        ["verification", _ops_text(report.get("verification"))],
    ]


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    report = _ops_as_dict(report)
    fmt = str(export_format or "pdf").lower().strip()

    if fmt in {"excel", "xls"}:
        fmt = "xlsx"
    if fmt in {"md"}:
        fmt = "markdown"

    title = _ops_report_title(report)
    stamp = _OpsDateTime.now().strftime("%Y%m%d_%H%M%S")
    base = _ops_safe_name(f"{fallback_name}_{stamp}")

    if fmt == "json":
        path = _OPS_EXPORT_DIR / f"{base}.json"
        path.write_text(_ops_json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        return path, "application/json"

    if fmt == "csv":
        path = _OPS_EXPORT_DIR / f"{base}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = _ops_csv.writer(f)
            writer.writerow(["field", "value"])
            writer.writerows(_ops_flat_rows(report))
        return path, "text/csv"

    if fmt == "xlsx":
        from openpyxl import Workbook

        path = _OPS_EXPORT_DIR / f"{base}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Report"
        ws.append(["field", "value"])
        for row in _ops_flat_rows(report):
            ws.append(row)

        for col in ("A", "B"):
            ws.column_dimensions[col].width = 35 if col == "A" else 110

        wb.save(path)
        return path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if fmt == "markdown":
        path = _OPS_EXPORT_DIR / f"{base}.md"
        path.write_text(_ops_markdown(report), encoding="utf-8")
        return path, "text/markdown"

    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        path = _OPS_EXPORT_DIR / f"{base}.pdf"
        doc = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(_ops_report_title(report), styles["Title"]))
        story.append(Spacer(1, 12))

        for line in _ops_markdown(report).splitlines():
            clean = line.strip()
            if not clean:
                story.append(Spacer(1, 6))
                continue

            clean = clean.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

            if clean.startswith("# "):
                story.append(Paragraph(clean[2:], styles["Title"]))
            elif clean.startswith("## "):
                story.append(Paragraph(clean[3:], styles["Heading2"]))
            elif clean.startswith("### "):
                story.append(Paragraph(clean[4:], styles["Heading3"]))
            elif clean.startswith("- "):
                story.append(Paragraph("• " + clean[2:], styles["BodyText"]))
            else:
                story.append(Paragraph(clean, styles["BodyText"]))

        doc.build(story)
        return path, "application/pdf"

    raise HTTPException(status_code=400, detail=f"Unsupported export format: {export_format}")


@app.post("/api/v2/export/{export_format}")
async def opslens_v2_export(export_format: str, report_payload=Body(...)):
    # Keep v2 route for compatibility, but use the clean export pipeline.
    report = _normalize_report(_ops_as_dict(report_payload))
    affected = report.get("affected_resources") or {}
    fallback = (
        affected.get("service")
        or affected.get("service_name")
        or report.get("record_id")
        or "opslens-report"
    )

    return _file_response(_export_payload(report, export_format, fallback_name=fallback))


@app.get("/api/v2/history/{record_id}/download/{export_format}")
def opslens_v2_history_download(record_id: str, export_format: str):
    # Keep v2 route for compatibility, but use the clean export pipeline.
    report = history_store.get_report(record_id)
    return _file_response(_export_payload(report, export_format, fallback_name=record_id))


if WEB_DIR.exists():
    app.mount("/", StaticFiles(directory=str(WEB_DIR), html=True), name="web")

# =========================================================
# OpsLens final robust export override v4
# =========================================================
# OpsLens final robust export override v4
# Single clean export pipeline for:
# /api/export/report/{format}
# /api/db/history/{record_id}/report/{format}/download
# /api/v2/export/{format}
# /api/v2/history/{record_id}/download/{format}
# =========================================================

def _opslens_export_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        # Prefer command text when dict is a command object.
        if "command" in value:
            title = value.get("title") or value.get("name") or ""
            cmd = value.get("command") or ""
            return f"{title}: {cmd}".strip(": ")

        parts = []
        for key, val in value.items():
            txt = _opslens_export_text(val)
            if txt:
                parts.append(f"{key}: {txt}")
        return "\n".join(parts)

    if isinstance(value, (list, tuple, set)):
        return "\n".join(_opslens_export_text(item) for item in value if _opslens_export_text(item))

    return str(value)


def _opslens_export_safe_name(value, fallback="opslens-report"):
    import re

    value = str(value or fallback).strip()
    value = re.sub(r"[^a-zA-Z0-9_.-]+", "_", value)
    value = value.strip("_")
    return value[:120] or fallback


def _opslens_export_format(export_format):
    fmt = str(export_format or "pdf").lower().strip()

    if fmt in {"excel", "xls"}:
        return "xlsx"

    if fmt in {"md"}:
        return "markdown"

    return fmt


def _opslens_export_media_type(export_format):
    fmt = _opslens_export_format(export_format)

    return {
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "json": "application/json",
        "markdown": "text/markdown",
    }.get(fmt, "application/octet-stream")


def _opslens_export_report(report):
    try:
        report = _normalize_report(report)
    except Exception:
        try:
            report = _ops_as_dict(report)
        except Exception:
            report = report if isinstance(report, dict) else {}

    if not isinstance(report, dict):
        report = {}

    if isinstance(report.get("report"), dict):
        report = report["report"]

    return report


def _opslens_export_title(report):
    return (
        report.get("title")
        or report.get("incident_title")
        or "OpsLens Incident Report"
    )


def _opslens_export_affected(report):
    affected = report.get("affected_resources")
    return affected if isinstance(affected, dict) else {}


def _opslens_export_fix(report):
    fix = report.get("recommended_fix")
    return fix if isinstance(fix, dict) else {"strategy": _opslens_export_text(fix), "actions": [], "commands": []}


def _opslens_export_verification(report):
    verification = report.get("verification")
    return verification if isinstance(verification, dict) else {"intent": _opslens_export_text(verification), "commands": []}


def _opslens_export_commands(report):
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    rows = []

    for item in fix.get("commands") or fix.get("safe_commands") or []:
        rows.append(["Remediation", _opslens_export_text(item)])

    for item in verification.get("commands") or verification.get("verification_commands") or []:
        rows.append(["Verification", _opslens_export_text(item)])

    return rows


def _opslens_export_base(report, fallback_name="report"):
    from datetime import datetime

    affected = _opslens_export_affected(report)
    service = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("deployment")
        or report.get("record_id")
        or fallback_name
        or "opslens-report"
    )

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _opslens_export_safe_name(f"{service}_{stamp}")


def _opslens_export_markdown(report):
    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    lines = [
        f"# {_opslens_export_title(report)}",
        "",
        "## Overview",
        f"- Severity: {report.get('severity', 'unknown')}",
        f"- Confidence: {report.get('confidence', 'unknown')}",
        f"- Namespace: {affected.get('namespace', '')}",
        f"- Node: {affected.get('node') or affected.get('node_name') or ''}",
        f"- Service: {affected.get('service') or affected.get('service_name') or ''}",
        f"- Deployment: {affected.get('deployment') or affected.get('deployment_name') or ''}",
        "",
        "## Incident Summary",
        _opslens_export_text(report.get("incident_summary") or report.get("summary")),
        "",
        "## Root Cause",
        _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis")),
        "",
        "## Evidence",
    ]

    evidence = report.get("agent_reasoning") or report.get("evidence_trail") or report.get("evidence") or []

    if evidence:
        for item in evidence:
            lines.append(f"- {_opslens_export_text(item)}")
    else:
        lines.append("- No evidence rows available.")

    additional = report.get("additional_findings") or []

    if additional:
        lines.extend(["", "## Additional Findings"])
        for item in additional:
            lines.append(f"- {_opslens_export_text(item)}")

    lines.extend([
        "",
        "## Recommended Fix",
        _opslens_export_text(fix.get("strategy")),
        "",
        "## Commands",
    ])

    commands = _opslens_export_commands(report)

    if commands:
        for kind, command in commands:
            lines.append(f"- {kind}: `{command}`")
    else:
        lines.append("- No commands available.")

    lines.extend([
        "",
        "## Verification",
        _opslens_export_text(verification.get("intent")),
        "",
    ])

    return "\n".join(lines).strip() + "\n"


def _opslens_write_json(report, path):
    import json

    path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    return path


def _opslens_write_csv(report, path):
    import csv

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    rows = [
        ["section", "key", "value"],
        ["overview", "title", _opslens_export_title(report)],
        ["overview", "severity", report.get("severity", "")],
        ["overview", "confidence", report.get("confidence", "")],
        ["affected_resources", "namespace", affected.get("namespace", "")],
        ["affected_resources", "node", affected.get("node") or affected.get("node_name") or ""],
        ["affected_resources", "service", affected.get("service") or affected.get("service_name") or ""],
        ["affected_resources", "deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["incident", "summary", _opslens_export_text(report.get("incident_summary") or report.get("summary"))],
        ["incident", "root_cause", _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["recommended_fix", "strategy", _opslens_export_text(fix.get("strategy"))],
        ["verification", "intent", _opslens_export_text(verification.get("intent"))],
    ]

    for index, item in enumerate(report.get("agent_reasoning") or [], start=1):
        rows.append(["evidence", f"row_{index}", _opslens_export_text(item)])

    for index, item in enumerate(report.get("additional_findings") or [], start=1):
        rows.append(["additional_findings", f"row_{index}", _opslens_export_text(item)])

    for index, row in enumerate(_opslens_export_commands(report), start=1):
        rows.append(["commands", f"{row[0]}_{index}", row[1]])

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return path


def _opslens_write_xlsx(report, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def style(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="111827")

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 38 if col == 1 else 95

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    rows = [
        ["Title", _opslens_export_title(report)],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Incident Summary", _opslens_export_text(report.get("incident_summary") or report.get("summary"))],
        ["Root Cause", _opslens_export_text(report.get("root_cause_story") or report.get("root_cause_hypothesis"))],
        ["Fix Strategy", _opslens_export_text(fix.get("strategy"))],
        ["Verification", _opslens_export_text(verification.get("intent"))],
    ]

    for row in rows:
        ws.append(row)
    style(ws)

    ws = wb.create_sheet("Evidence")
    ws.append(["Agent", "Finding", "Meaning"])

    for item in report.get("agent_reasoning") or []:
        if isinstance(item, dict):
            ws.append([
                _opslens_export_text(item.get("agent")),
                _opslens_export_text(item.get("finding")),
                _opslens_export_text(item.get("meaning")),
            ])
        else:
            ws.append(["Evidence", _opslens_export_text(item), ""])
    style(ws)

    ws = wb.create_sheet("Additional Findings")
    ws.append(["Resource", "Finding", "Impact", "Priority"])

    for item in report.get("additional_findings") or []:
        if isinstance(item, dict):
            ws.append([
                _opslens_export_text(item.get("resource")),
                _opslens_export_text(item.get("finding")),
                _opslens_export_text(item.get("impact")),
                _opslens_export_text(item.get("priority")),
            ])
        else:
            ws.append(["", _opslens_export_text(item), "", ""])
    style(ws)

    ws = wb.create_sheet("Commands")
    ws.append(["Type", "Command"])

    for row in _opslens_export_commands(report):
        ws.append(row)
    style(ws)

    wb.save(path)
    return path


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    affected = _opslens_export_affected(report)
    fix = _opslens_export_fix(report)
    verification = _opslens_export_verification(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensTitleFinal",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=21,
        spaceAfter=10,
    )

    heading_style = ParagraphStyle(
        "OpsLensHeadingFinal",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=5,
    )

    body_style = ParagraphStyle(
        "OpsLensBodyFinal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "OpsLensCellFinal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.8,
        leading=9.5,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsLensHeaderFinal",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.5,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensCodeFinal",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.2,
        leading=8.8,
        wordWrap="CJK",
    )

    story = []

    def clean(value):
        return html.escape(_opslens_export_text(value)).replace("\n", "<br/>")

    def p(value, style=body_style):
        story.append(Paragraph(clean(value), style))

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(clean(value), heading_style))

    def table(headers, rows, widths=None):
        if not rows:
            p("No data available.")
            return

        usable = A4[0] - (2.5 * cm)

        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(clean(head), header_style) for head in headers]]

        for row in rows:
            data.append([Paragraph(clean(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(clean(_opslens_export_title(report)), title_style))

    overview_rows = [
        ["Severity", report.get("severity", "unknown")],
        ["Confidence", report.get("confidence", "unknown")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
    ]

    table(["Field", "Value"], overview_rows, [4.0 * cm, 13.0 * cm])

    heading("Incident Summary")
    p(report.get("incident_summary") or report.get("summary"))

    heading("Root Cause")
    p(report.get("root_cause_story") or report.get("root_cause_hypothesis"))

    evidence_rows = []

    for item in report.get("agent_reasoning") or []:
        if isinstance(item, dict):
            evidence_rows.append([
                item.get("agent", ""),
                item.get("finding", ""),
                item.get("meaning", ""),
            ])
        else:
            evidence_rows.append(["Evidence", _opslens_export_text(item), ""])

    heading("Evidence")
    table(["Agent", "Finding", "Meaning"], evidence_rows, [3.4 * cm, 5.4 * cm, 8.2 * cm])

    additional_rows = []

    for item in report.get("additional_findings") or []:
        if isinstance(item, dict):
            additional_rows.append([
                item.get("resource", ""),
                item.get("finding", ""),
                item.get("impact", ""),
                item.get("priority", ""),
            ])
        else:
            additional_rows.append(["", _opslens_export_text(item), "", ""])

    heading("Additional Findings")
    table(["Resource", "Finding", "Impact", "Priority"], additional_rows, [3.8 * cm, 4.8 * cm, 6.0 * cm, 2.4 * cm])

    heading("Recommended Fix")
    p(fix.get("strategy", ""))

    heading("Verification")
    p(verification.get("intent", ""))

    command_rows = _opslens_export_commands(report)

    heading("Commands")

    if command_rows:
        for index, row in enumerate(command_rows, start=1):
            p(f"{row[0]} Command {index}")
            story.append(Paragraph(clean(row[1]), code_style))
            story.append(Spacer(1, 5))
    else:
        p("No commands available.")

    doc.build(story)
    return path


def _opslens_write_export_file(report, export_format, fallback_name="report"):
    fmt = _opslens_export_format(export_format)
    report = _opslens_export_report(report)

    if not report:
        raise ValueError("No report payload was provided for export.")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    base = _opslens_export_base(report, fallback_name=fallback_name)

    if fmt == "json":
        return _opslens_write_json(report, EXPORTS_DIR / f"{base}.json")

    if fmt == "csv":
        return _opslens_write_csv(report, EXPORTS_DIR / f"{base}.csv")

    if fmt == "markdown":
        path = EXPORTS_DIR / f"{base}.md"
        path.write_text(_opslens_export_markdown(report), encoding="utf-8")
        return path

    if fmt == "xlsx":
        return _opslens_write_xlsx(report, EXPORTS_DIR / f"{base}.xlsx")

    if fmt == "pdf":
        return _opslens_write_pdf(report, EXPORTS_DIR / f"{base}.pdf")

    raise ValueError(f"Unsupported report format: {export_format}")


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(report, export_format, fallback_name=fallback_name)
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens professional SRE export override
# =========================================================
# OpsLens professional SRE export override
# Professional PDF/Markdown incident report format.
# =========================================================

def _pro_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, dict):
        for preferred in ["summary", "finding", "meaning", "message", "command", "strategy", "intent"]:
            if value.get(preferred):
                return str(value.get(preferred)).strip()

        return "; ".join(f"{k}: {_pro_text(v)}" for k, v in value.items() if _pro_text(v))

    if isinstance(value, (list, tuple, set)):
        return "\n".join(_pro_text(item) for item in value if _pro_text(item))

    return str(value).strip()


def _pro_report(report):
    try:
        report = _normalize_report(report)
    except Exception:
        report = report if isinstance(report, dict) else {}

    if isinstance(report.get("report"), dict):
        report = report["report"]

    return report if isinstance(report, dict) else {}


def _pro_affected(report):
    affected = report.get("affected_resources")
    return affected if isinstance(affected, dict) else {}


def _pro_title(report):
    affected = _pro_affected(report)
    deployment = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or report.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )

    severity = str(report.get("severity") or "UNKNOWN").upper()
    return f"[{severity}] Incident Report: {deployment}"


def _pro_verdict(report):
    return (
        _pro_text(report.get("root_cause_story"))
        or _pro_text(report.get("root_cause_hypothesis"))
        or _pro_text(report.get("summary"))
        or "OpsLens identified an active Kubernetes incident from correlated signals."
    )


def _pro_summary(report):
    return (
        _pro_text(report.get("incident_summary"))
        or _pro_text(report.get("summary"))
        or _pro_verdict(report)
    )


def _pro_findings(report):
    findings = []

    for key in ["agent_reasoning", "evidence_trail", "evidence", "supporting_signals"]:
        value = report.get(key)

        if isinstance(value, list):
            findings.extend(value)

    clean = []

    for item in findings:
        text = _pro_text(item)
        if not text:
            continue

        lowered = text.lower()
        if "agentexecutionerror" in lowered or "circular reference" in lowered or "podlstmunavailable" in lowered:
            continue

        clean.append(item)

    return clean[:12]


def _pro_smoking_gun(report):
    findings = _pro_findings(report)

    priority_words = [
        "readiness probe failed",
        "statuscode: 404",
        "oomkilled",
        "crashloopbackoff",
        "imagepullbackoff",
        "connection refused",
        "targetport",
        "selector",
        "no available replicas",
        "traceback",
        "fatal",
        "error",
    ]

    for item in findings:
        text = _pro_text(item)
        lowered = text.lower()

        if any(word in lowered for word in priority_words):
            return text

    if findings:
        return _pro_text(findings[0])

    return _pro_summary(report)


def _pro_status(report):
    affected = _pro_affected(report)
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")

    if pod and deployment:
        return f"Pod {pod}; deployment {deployment} is impacted."

    if pod:
        return f"Pod {pod} is impacted."

    if deployment:
        return f"Deployment {deployment} is impacted."

    return "Active incident detected."


def _pro_blast_radius(report):
    affected = _pro_affected(report)
    service = affected.get("service") or affected.get("service_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    namespace = affected.get("namespace")

    parts = []

    if service:
        parts.append(f"Service: {service}")

    if deployment:
        parts.append(f"Deployment: {deployment}")

    if namespace:
        parts.append(f"Namespace: {namespace}")

    additional = report.get("additional_findings") or []
    if additional:
        parts.append(f"Related findings: {len(additional)}")

    return ", ".join(parts) if parts else "Limited to the detected Kubernetes scope."


def _pro_fix_steps(report):
    fix = report.get("recommended_fix")

    steps = []

    if isinstance(fix, dict):
        for key in ["strategy", "actions", "steps", "commands"]:
            value = fix.get(key)
            if isinstance(value, list):
                steps.extend(_pro_text(item) for item in value)
            elif value:
                steps.append(_pro_text(value))

    elif isinstance(fix, list):
        steps.extend(_pro_text(item) for item in fix)

    elif fix:
        steps.append(_pro_text(fix))

    title = (_pro_title(report) + " " + _pro_verdict(report)).lower()

    if not steps:
        if "readiness" in title:
            steps = [
                "Inspect the pod readinessProbe path and port.",
                "Verify the application exposes the configured health endpoint.",
                "Update either the probe configuration or the application health route.",
                "Roll out the deployment and confirm the pod becomes Ready.",
            ]
        elif "selector" in title:
            steps = [
                "Compare Service selector labels with pod template labels.",
                "Update the Service selector or pod labels so endpoints are created.",
                "Re-apply the manifest and verify endpoints.",
            ]
        elif "targetport" in title:
            steps = [
                "Compare Service targetPort with the containerPort.",
                "Update targetPort or containerPort to align.",
                "Verify endpoints and client connectivity.",
            ]
        else:
            steps = [
                "Inspect the affected Kubernetes resource.",
                "Validate events, logs, configuration, and resource metrics.",
                "Apply the targeted fix and verify recovery.",
            ]

    return [step for step in steps if step][:8]


def _pro_commands(report):
    commands = []

    fix = report.get("recommended_fix")
    verification = report.get("verification")

    def add_from(value):
        if isinstance(value, dict):
            for key in ["commands", "safe_commands", "verification_commands"]:
                items = value.get(key)
                if isinstance(items, list):
                    for item in items:
                        txt = _pro_text(item)
                        if txt and txt.startswith("kubectl"):
                            commands.append(txt)
            for v in value.values():
                if isinstance(v, str) and v.strip().startswith("kubectl"):
                    commands.append(v.strip())

        elif isinstance(value, list):
            for item in value:
                txt = _pro_text(item)
                if txt and txt.startswith("kubectl"):
                    commands.append(txt)

        elif isinstance(value, str) and value.strip().startswith("kubectl"):
            commands.append(value.strip())

    add_from(fix)
    add_from(verification)

    affected = _pro_affected(report)
    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    defaults = [
        f"kubectl get pods -n {ns}",
        f"kubectl get events -n {ns} --sort-by=.lastTimestamp",
    ]

    if pod:
        defaults.append(f"kubectl describe pod {pod} -n {ns}")
        defaults.append(f"kubectl logs {pod} -n {ns} --tail=120")

    if deployment:
        defaults.append(f"kubectl describe deployment {deployment} -n {ns}")
        defaults.append(f"kubectl rollout status deployment/{deployment} -n {ns}")

    if service:
        defaults.append(f"kubectl get endpoints {service} -n {ns}")
        defaults.append(f"kubectl get svc {service} -n {ns} -o yaml")

    for cmd in defaults:
        if cmd not in commands:
            commands.append(cmd)

    return commands[:10]


def _pro_related_findings(report):
    rows = []

    for item in report.get("additional_findings") or []:
        text = _pro_text(item)
        lowered = text.lower()

        if not text:
            continue

        if "agentexecutionerror" in lowered or "circular reference" in lowered or "podlstmunavailable" in lowered:
            continue

        if "not found in collected evidence" in lowered and "deployment" in lowered:
            continue

        if isinstance(item, dict):
            rows.append([
                _pro_text(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope"),
                _pro_text(item.get("finding") or item.get("anomaly_type") or item.get("summary")),
                _pro_text(item.get("impact") or item.get("meaning") or "Related operational signal."),
                _pro_text(item.get("priority") or item.get("severity") or "Follow-up"),
            ])
        else:
            rows.append(["scope", text, "Related operational signal.", "Follow-up"])

    return rows[:8]


def _opslens_export_markdown(report):
    report = _pro_report(report)
    affected = _pro_affected(report)

    lines = []
    lines.append(f"# {_pro_title(report)}")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Incident Snapshot")
    lines.append(f"- **Namespace:** {affected.get('namespace', '')}")
    lines.append(f"- **Detected Time:** {report.get('generated_at') or report.get('created_at') or ''}")
    lines.append(f"- **Affected Deployment:** {affected.get('deployment') or affected.get('deployment_name') or ''}")
    lines.append(f"- **Affected Pod:** {affected.get('pod') or affected.get('pod_name') or ''}")
    lines.append(f"- **Affected Service:** {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- **Node:** {affected.get('node') or affected.get('node_name') or ''}")
    lines.append(f"- **Blast Radius:** {_pro_blast_radius(report)}")
    lines.append(f"- **Status:** {_pro_status(report)}")
    lines.append("")
    lines.append("## 2. The Verdict")
    lines.append(_pro_verdict(report))
    lines.append("")
    lines.append("## 3. Evidence: The Smoking Gun")
    lines.append("```")
    lines.append(_pro_smoking_gun(report))
    lines.append("```")
    lines.append("")
    lines.append("## 4. Evidence Trail")
    for item in _pro_findings(report):
        lines.append(f"- {_pro_text(item)}")
    lines.append("")
    lines.append("## 5. Fix Plan")
    for index, step in enumerate(_pro_fix_steps(report), start=1):
        lines.append(f"{index}. {step}")
    lines.append("")
    lines.append("## 6. Verification Commands")
    for cmd in _pro_commands(report):
        lines.append(f"```bash\n{cmd}\n```")
    lines.append("")
    related = _pro_related_findings(report)
    if related:
        lines.append("## 7. Related / Separate Findings")
        for resource, finding, impact, priority in related:
            lines.append(f"- **{resource}** — {finding}. Impact: {impact}. Priority: {priority}.")
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _pro_report(report)
    affected = _pro_affected(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensProTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=8,
    )

    h_style = ParagraphStyle(
        "OpsLensProHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        spaceBefore=7,
        spaceAfter=5,
    )

    body = ParagraphStyle(
        "OpsLensProBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.6,
        leading=11,
        spaceAfter=4,
    )

    code = ParagraphStyle(
        "OpsLensProCode",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.4,
        leading=9,
        backColor=colors.HexColor("#F3F4F6"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.4,
        borderPadding=5,
        wordWrap="CJK",
    )

    cell = ParagraphStyle(
        "OpsLensProCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.7,
        leading=9.2,
        wordWrap="CJK",
    )

    head = ParagraphStyle(
        "OpsLensProHead",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.7,
        leading=9.2,
        textColor=colors.white,
        wordWrap="CJK",
    )

    story = []

    def esc(v):
        return html.escape(_pro_text(v)).replace("\n", "<br/>")

    def p(v, style=body):
        if _pro_text(v):
            story.append(Paragraph(esc(v), style))

    def heading(v):
        story.append(Spacer(1, 5))
        story.append(Paragraph(esc(v), h_style))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.5 * cm
        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc(h), head) for h in headers]]
        for row in rows:
            data.append([Paragraph(esc(x), cell) for x in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))

    story.append(Paragraph(esc(_pro_title(report)), title_style))

    snapshot_rows = [
        ["Namespace", affected.get("namespace", "")],
        ["Detected Time", report.get("generated_at") or report.get("created_at") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Pod", affected.get("pod") or affected.get("pod_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Blast Radius", _pro_blast_radius(report)],
        ["Status", _pro_status(report)],
    ]
    heading("1. Incident Snapshot")
    table(["Field", "Value"], snapshot_rows, [4.0 * cm, 13.0 * cm])

    heading("2. The Verdict")
    p(_pro_verdict(report))

    heading("3. Evidence: The Smoking Gun")
    story.append(Paragraph(esc(_pro_smoking_gun(report)), code))
    story.append(Spacer(1, 6))

    evidence_rows = []
    for item in _pro_findings(report):
        if isinstance(item, dict):
            evidence_rows.append([
                item.get("agent") or item.get("source_agent") or item.get("agent_name") or "Investigation",
                item.get("anomaly_type") or item.get("finding") or item.get("event_type") or "Signal",
                item.get("summary") or item.get("meaning") or _pro_text(item),
            ])
        else:
            evidence_rows.append(["Investigation", "Evidence", _pro_text(item)])

    heading("4. Evidence Trail")
    table(["Agent", "Finding", "Meaning"], evidence_rows, [3.5 * cm, 4.2 * cm, 9.3 * cm])

    heading("5. Fix Plan")
    for index, step in enumerate(_pro_fix_steps(report), start=1):
        p(f"{index}. {step}")

    heading("6. Verification Commands")
    for cmd in _pro_commands(report):
        story.append(Paragraph(esc(cmd), code))
        story.append(Spacer(1, 4))

    related = _pro_related_findings(report)
    if related:
        heading("7. Related / Separate Findings")
        table(["Resource", "Finding", "Impact", "Priority"], related, [3.7 * cm, 4.7 * cm, 6.4 * cm, 2.2 * cm])

    doc.build(story)
    return path

# =========================================================
# OpsLens final export binding guard
# =========================================================
# OpsLens final export binding guard
# Guarantees all report download routes use the final professional export pipeline.
# =========================================================

def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens final professional report content cleanup v2
# =========================================================
# OpsLens final professional report content cleanup v2
# Fixes:
# - GenericErrorLog as weak Smoking Gun
# - raw action_type dictionaries in Fix Plan
# - duplicated commands
# - duplicated evidence rows
# - missing pod extraction from evidence/summary
# =========================================================

import re as _opslens_pro_re


def _pro_flat_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, dict):
        parts = []
        for key in [
            "summary", "meaning", "message", "reason", "finding",
            "anomaly_type", "event_type", "command", "strategy", "intent"
        ]:
            if value.get(key):
                parts.append(str(value.get(key)))

        for val in value.values():
            if isinstance(val, (dict, list)):
                txt = _pro_flat_text(val)
                if txt:
                    parts.append(txt)

        return "\n".join(parts)

    if isinstance(value, (list, tuple, set)):
        return "\n".join(_pro_flat_text(item) for item in value if _pro_flat_text(item))

    return str(value)


def _pro_clean_line(value):
    text = _pro_flat_text(value)
    text = text.replace("\r", " ").replace("\n", " ")
    text = _opslens_pro_re.sub(r"\s+", " ", text).strip()
    return text


def _pro_text(value):
    if value is None:
        return ""

    if isinstance(value, str):
        return _pro_clean_line(value)

    if isinstance(value, dict):
        for preferred in ["summary", "meaning", "message", "reason", "finding", "strategy", "intent", "command"]:
            if value.get(preferred):
                return _pro_clean_line(value.get(preferred))

        return _pro_clean_line(value)

    if isinstance(value, (list, tuple, set)):
        return "\n".join(_pro_text(item) for item in value if _pro_text(item))

    return _pro_clean_line(value)


def _pro_is_internal_text(text):
    lowered = text.lower()
    blocked = [
        "agentexecutionerror",
        "circular reference",
        "podlstmunavailable",
        "serialization error",
        "internal monitoring tool failure",
        "agent_health",
        "tool-health",
        "metrics context: status=not_available",
    ]
    return any(item in lowered for item in blocked)


def _pro_is_command(text):
    return _pro_clean_line(text).lower().startswith("kubectl ")


def _pro_dedupe(items):
    seen = set()
    output = []

    for item in items:
        key = _pro_clean_line(item).lower()
        if not key or key in seen:
            continue
        seen.add(key)
        output.append(item)

    return output


def _pro_extract_first(patterns, text):
    for pattern in patterns:
        match = _opslens_pro_re.search(pattern, text, flags=_opslens_pro_re.I)
        if match:
            return match.group(1).strip(" '\"`.,")
    return ""


def _pro_affected(report):
    affected = report.get("affected_resources")
    affected = dict(affected) if isinstance(affected, dict) else {}

    all_text = _pro_flat_text(report)

    if not affected.get("pod") and not affected.get("pod_name"):
        pod = _pro_extract_first(
            [
                r"pod[/\s']+([a-z0-9][a-z0-9.-]*-[a-z0-9]+)",
                r"Pod/([a-z0-9][a-z0-9.-]*-[a-z0-9]+)",
                r"pod\s+'([^']+)'",
                r"pod\s+([a-z0-9][a-z0-9.-]*-[a-z0-9]+)",
            ],
            all_text,
        )
        if pod:
            affected["pod"] = pod

    if not affected.get("deployment") and not affected.get("deployment_name"):
        deployment = _pro_extract_first(
            [
                r"deployment\s+'([^']+)'",
                r"deployment[/\s]+([a-z0-9][a-z0-9.-]+)",
                r"Deployment\s+([a-z0-9][a-z0-9.-]+)",
            ],
            all_text,
        )
        if deployment:
            affected["deployment"] = deployment

    return affected


def _pro_title(report):
    affected = _pro_affected(report)
    workload = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )

    severity = str(report.get("severity") or "UNKNOWN").upper()
    return f"[{severity}] Incident Report: {workload}"


def _pro_verdict(report):
    text = (
        _pro_text(report.get("root_cause_story"))
        or _pro_text(report.get("root_cause_hypothesis"))
        or _pro_text(report.get("incident_summary"))
        or _pro_text(report.get("summary"))
    )

    if text:
        return text

    return "OpsLens identified an active Kubernetes incident from correlated Kubernetes, log, configuration, and metrics signals."


def _pro_summary(report):
    return (
        _pro_text(report.get("incident_summary"))
        or _pro_text(report.get("summary"))
        or _pro_verdict(report)
    )


def _pro_findings(report):
    raw_items = []

    for key in ["agent_reasoning", "evidence_trail", "evidence", "supporting_signals", "primary_signal"]:
        value = report.get(key)

        if isinstance(value, list):
            raw_items.extend(value)
        elif isinstance(value, dict):
            raw_items.append(value)
        elif value:
            raw_items.append(value)

    rows = []
    seen = set()

    for item in raw_items:
        if isinstance(item, dict):
            agent = (
                item.get("agent")
                or item.get("agent_name")
                or item.get("source_agent")
                or "Investigation"
            )

            finding = (
                item.get("finding")
                or item.get("anomaly_type")
                or item.get("event_type")
                or item.get("signal")
                or "Evidence"
            )

            meaning = (
                item.get("meaning")
                or item.get("summary")
                or item.get("message")
                or item.get("reason")
                or _pro_text(item)
            )
        else:
            agent = "Investigation"
            finding = "Evidence"
            meaning = _pro_text(item)

        agent = _pro_clean_line(agent)
        finding = _pro_clean_line(finding)
        meaning = _pro_clean_line(meaning)

        combined = f"{agent} {finding} {meaning}"

        if not meaning or _pro_is_internal_text(combined):
            continue

        lowered = combined.lower()

        # Remove duplicate meta narration rows.
        if lowered.startswith("investigation evidence primary signal from"):
            continue
        if lowered.startswith("investigation evidence supporting signal from"):
            continue
        if "metrics context: status=not_available" in lowered:
            continue

        key = f"{agent}|{finding}|{meaning}".lower()
        if key in seen:
            continue

        seen.add(key)
        rows.append({
            "agent": agent,
            "finding": finding,
            "meaning": meaning,
            "summary": meaning,
            "anomaly_type": finding,
        })

    return rows[:8]


def _pro_smoking_gun(report):
    all_text = _pro_flat_text(report)
    verdict = _pro_verdict(report)

    # Best handcrafted smoking gun for the common readiness failure case.
    if "readiness" in all_text.lower() and ("404" in all_text or "not-found" in all_text.lower()):
        path = _pro_extract_first([r"(/healthz[-a-zA-Z0-9_/]*)"], all_text)
        if path:
            return f"Readiness probe failed with HTTP 404 because the configured health endpoint '{path}' does not exist or is not being served by the application."
        return "Readiness probe failed with HTTP 404, preventing the pod from becoming Ready and leaving the deployment with no available replicas."

    priority_words = [
        "readiness probe failed",
        "http probe failed",
        "statuscode: 404",
        "oomkilled",
        "crashloopbackoff",
        "imagepullbackoff",
        "errimagepull",
        "connection refused",
        "targetport",
        "selector mismatch",
        "no available replicas",
        "traceback",
        "fatal",
    ]

    for item in _pro_findings(report):
        text = _pro_text(item.get("meaning") or item)
        lowered = text.lower()

        if _pro_is_internal_text(text):
            continue

        if text.strip().lower() in {"genericerrorlog", "evidence", "signal"}:
            continue

        if any(word in lowered for word in priority_words):
            return text

    if verdict and verdict.lower() != "genericerrorlog":
        return verdict

    return _pro_summary(report)


def _pro_status(report):
    affected = _pro_affected(report)

    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")

    if pod and deployment:
        return f"Pod {pod} is not Ready; deployment {deployment} has reduced or zero availability."

    if pod:
        return f"Pod {pod} is impacted."

    if deployment:
        return f"Deployment {deployment} is impacted."

    return "Active incident detected."


def _pro_blast_radius(report):
    affected = _pro_affected(report)

    parts = []

    service = affected.get("service") or affected.get("service_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    namespace = affected.get("namespace")

    if service:
        parts.append(f"service/{service}")

    if deployment:
        parts.append(f"deployment/{deployment}")

    if namespace:
        parts.append(f"namespace/{namespace}")

    return ", ".join(parts) if parts else "Limited to the detected Kubernetes scope."


def _pro_action_to_step(action):
    if isinstance(action, dict):
        action_type = str(action.get("action_type") or "").lower()
        reason = _pro_text(action.get("reason"))
        target = _pro_text(action.get("target_kind") or action.get("target"))

        if "update_readiness_probe" in action_type:
            return "Update the deployment readinessProbe path/port to match a real application health endpoint."

        if "investigate_application_logs" in action_type:
            return "Inspect the affected pod logs to confirm which health endpoint the application actually exposes."

        if reason and target:
            return f"{reason}"

        if reason:
            return reason

        return ""

    return _pro_text(action)


def _pro_fix_steps(report):
    fix = report.get("recommended_fix")

    steps = []

    if isinstance(fix, dict):
        strategy = _pro_text(fix.get("strategy") or fix.get("explanation"))
        if strategy and not _pro_is_command(strategy):
            steps.append(strategy)

        for action in fix.get("actions") or fix.get("steps") or []:
            step = _pro_action_to_step(action)
            if step and not _pro_is_command(step):
                steps.append(step)

    elif isinstance(fix, list):
        for item in fix:
            step = _pro_action_to_step(item)
            if step and not _pro_is_command(step):
                steps.append(step)

    elif fix:
        step = _pro_text(fix)
        if step and not _pro_is_command(step):
            steps.append(step)

    # Remove raw dict-looking output.
    cleaned = []
    for step in steps:
        lower = step.lower()
        if "action_type:" in lower or "target_kind:" in lower or "risk:" in lower:
            continue
        if step.strip().startswith("{") or step.strip().startswith("["):
            continue
        cleaned.append(step)

    steps = _pro_dedupe(cleaned)

    title = (_pro_title(report) + " " + _pro_verdict(report) + " " + _pro_summary(report)).lower()

    if not steps:
        if "readiness" in title:
            steps = [
                "Inspect the current readinessProbe configuration on the affected deployment.",
                "Verify the application exposes the configured health endpoint and returns HTTP 200.",
                "Update either the readinessProbe path/port or the application health route.",
                "Roll out the deployment and confirm the pod transitions to Ready.",
            ]
        elif "selector" in title:
            steps = [
                "Compare the Service selector with the pod template labels.",
                "Update the Service selector or pod labels so endpoints are created.",
                "Re-apply the manifest and verify the Service has ready endpoints.",
            ]
        elif "targetport" in title:
            steps = [
                "Compare the Service targetPort with the containerPort exposed by the pod.",
                "Update the Service targetPort or the containerPort so they match.",
                "Verify endpoints and client connectivity after applying the change.",
            ]
        else:
            steps = [
                "Inspect the affected Kubernetes resource and its events.",
                "Validate logs, configuration, and metrics for the affected scope.",
                "Apply the targeted fix and verify service recovery.",
            ]

    return steps[:6]


def _pro_extract_command(text):
    text = _pro_clean_line(text)

    match = _opslens_pro_re.search(r"(kubectl\s+.+)$", text)
    if match:
        return match.group(1).strip()

    if text.startswith("kubectl "):
        return text

    return ""


def _pro_commands(report):
    commands = []

    def add(value):
        if value is None:
            return

        if isinstance(value, dict):
            if value.get("command"):
                cmd = _pro_extract_command(str(value.get("command")))
                if cmd:
                    commands.append(cmd)

            for key in ["commands", "safe_commands", "verification_commands"]:
                items = value.get(key)
                if isinstance(items, list):
                    for item in items:
                        add(item)

            for val in value.values():
                if isinstance(val, str):
                    cmd = _pro_extract_command(val)
                    if cmd:
                        commands.append(cmd)

        elif isinstance(value, list):
            for item in value:
                add(item)

        else:
            cmd = _pro_extract_command(str(value))
            if cmd:
                commands.append(cmd)

    add(report.get("recommended_fix"))
    add(report.get("verification"))

    affected = _pro_affected(report)
    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    defaults = [
        f"kubectl get pods -n {ns}",
        f"kubectl get events -n {ns} --sort-by=.lastTimestamp",
    ]

    if pod:
        defaults.extend([
            f"kubectl describe pod {pod} -n {ns}",
            f"kubectl logs {pod} -n {ns} --tail=120",
        ])

    if deployment:
        defaults.extend([
            f"kubectl describe deployment {deployment} -n {ns}",
            f"kubectl rollout status deployment/{deployment} -n {ns}",
        ])

    if service:
        defaults.extend([
            f"kubectl get endpoints {service} -n {ns}",
            f"kubectl get svc {service} -n {ns} -o yaml",
        ])

    commands.extend(defaults)
    return _pro_dedupe(commands)[:9]


def _pro_related_findings(report):
    rows = []

    for item in report.get("additional_findings") or []:
        text = _pro_text(item)
        lowered = text.lower()

        if not text or _pro_is_internal_text(text):
            continue

        if "not found in collected evidence" in lowered:
            continue

        if isinstance(item, dict):
            rows.append([
                _pro_text(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope"),
                _pro_text(item.get("finding") or item.get("anomaly_type") or item.get("summary")),
                _pro_text(item.get("impact") or item.get("meaning") or "Related operational signal."),
                _pro_text(item.get("priority") or item.get("severity") or "Follow-up"),
            ])
        else:
            rows.append(["scope", text, "Related operational signal.", "Follow-up"])

    return rows[:6]

# =========================================================
# OpsLens final simple operator report v3
# =========================================================
# OpsLens final simple operator report v3
# Final user-facing report structure:
# 1. Metadata / Scope
# 2. Problem
# 3. Root Cause
# 4. Recommended Fix
# 5. Verification
# 6. Summary
# =========================================================

import re as _simple_report_re


def _simple_clean(value):
    if value is None:
        return ""

    if isinstance(value, str):
        text = value.replace("\r", " ").replace("\n", " ")
        return _simple_report_re.sub(r"\s+", " ", text).strip()

    if isinstance(value, dict):
        for key in ["summary", "meaning", "message", "reason", "strategy", "intent", "finding", "command"]:
            if value.get(key):
                return _simple_clean(value.get(key))

        parts = []
        for item in value.values():
            txt = _simple_clean(item)
            if txt:
                parts.append(txt)

        return " ".join(parts).strip()

    if isinstance(value, (list, tuple, set)):
        return " ".join(_simple_clean(item) for item in value if _simple_clean(item)).strip()

    return str(value).strip()


def _simple_report(report):
    try:
        report = _normalize_report(report)
    except Exception:
        report = report if isinstance(report, dict) else {}

    if isinstance(report.get("report"), dict):
        report = report["report"]

    return report if isinstance(report, dict) else {}


def _simple_all_text(report):
    return _simple_clean(report)


def _simple_extract(patterns, text):
    for pattern in patterns:
        match = _simple_report_re.search(pattern, text, flags=_simple_report_re.I)
        if match:
            return match.group(1).strip(" '\"`.,")
    return ""


def _simple_affected(report):
    affected = report.get("affected_resources")
    affected = dict(affected) if isinstance(affected, dict) else {}

    all_text = _simple_all_text(report)

    if not affected.get("pod") and not affected.get("pod_name"):
        pod = _simple_extract(
            [
                r"Pod/([a-z0-9][a-z0-9.-]*-[a-z0-9]+)",
                r"pod\s+'([^']+)'",
                r"pod\s+([a-z0-9][a-z0-9.-]*-[a-z0-9]+)",
            ],
            all_text,
        )
        if pod:
            affected["pod"] = pod

    if not affected.get("deployment") and not affected.get("deployment_name"):
        deployment = _simple_extract(
            [
                r"deployment\s+'([^']+)'",
                r"deployment/([a-z0-9][a-z0-9.-]+)",
                r"Deployment\s+([a-z0-9][a-z0-9.-]+)",
            ],
            all_text,
        )
        if deployment:
            affected["deployment"] = deployment

    return affected


def _simple_title(report):
    affected = _simple_affected(report)

    workload = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )

    severity = str(report.get("severity") or "unknown").upper()

    return f"[{severity}] Incident Report: {workload}"


def _simple_problem(report):
    text = (
        _simple_clean(report.get("incident_summary"))
        or _simple_clean(report.get("summary"))
        or _simple_clean(report.get("title"))
        or _simple_clean(report.get("incident_title"))
    )

    all_text = _simple_all_text(report).lower()

    if "readiness" in all_text and "not ready" in all_text:
        affected = _simple_affected(report)
        pod = affected.get("pod") or affected.get("pod_name") or "the affected pod"
        deployment = affected.get("deployment") or affected.get("deployment_name") or "the deployment"
        return f"{pod} is failing readiness checks, so Kubernetes keeps it NotReady and {deployment} has no available ready replica."

    return text or "OpsLens detected an active Kubernetes workload issue."


def _simple_root_cause(report):
    text = (
        _simple_clean(report.get("root_cause_story"))
        or _simple_clean(report.get("root_cause_hypothesis"))
    )

    all_text = _simple_all_text(report)
    lowered = all_text.lower()

    if "readiness" in lowered and ("404" in lowered or "not-found" in lowered):
        path = _simple_extract([r"(/healthz[-a-zA-Z0-9_/]*)"], all_text)

        if path:
            return f"The readiness probe is targeting '{path}', but the application returns HTTP 404 for that endpoint, so the pod never becomes Ready."

        return "The readiness probe is returning HTTP 404, so Kubernetes marks the pod as NotReady and the deployment remains unavailable."

    return text or "The root cause was selected from the strongest correlated Kubernetes, log, and configuration signals."


def _simple_summary(report):
    affected = _simple_affected(report)

    namespace = affected.get("namespace") or "the selected namespace"
    node = affected.get("node") or affected.get("node_name") or "the selected node"
    deployment = affected.get("deployment") or affected.get("deployment_name") or "the affected workload"

    problem = _simple_problem(report)
    root = _simple_root_cause(report)

    return (
        f"OpsLens investigated {deployment} in namespace {namespace} on node {node}. "
        f"{problem} {root}"
    )


def _simple_dedupe(items):
    seen = set()
    output = []

    for item in items:
        item = _simple_clean(item)

        if not item:
            continue

        key = item.lower()

        if key in seen:
            continue

        seen.add(key)
        output.append(item)

    return output


def _simple_extract_kubectl(value):
    text = _simple_clean(value)
    match = _simple_report_re.search(r"(kubectl\s+.+)$", text)
    if match:
        return match.group(1).strip()
    if text.startswith("kubectl "):
        return text
    return ""


def _simple_collect_existing_commands(value):
    commands = []

    if value is None:
        return commands

    if isinstance(value, dict):
        if value.get("command"):
            cmd = _simple_extract_kubectl(value.get("command"))
            if cmd:
                commands.append(cmd)

        for key in ["commands", "safe_commands", "verification_commands"]:
            items = value.get(key)
            if isinstance(items, list):
                for item in items:
                    commands.extend(_simple_collect_existing_commands(item))

        for item in value.values():
            if isinstance(item, str):
                cmd = _simple_extract_kubectl(item)
                if cmd:
                    commands.append(cmd)

    elif isinstance(value, list):
        for item in value:
            commands.extend(_simple_collect_existing_commands(item))

    else:
        cmd = _simple_extract_kubectl(value)
        if cmd:
            commands.append(cmd)

    return commands


def _simple_recommended_fix_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    commands = []

    all_text = _simple_all_text(report).lower()

    if deployment and "readiness" in all_text:
        commands.extend([
            f"# Inspect the current readinessProbe configuration",
            f"kubectl describe deployment {deployment} -n {ns}",
            "",
            f"# Edit the readinessProbe path/port to match a real application health endpoint",
            f"kubectl edit deployment {deployment} -n {ns}",
            "",
            f"# Watch the rollout after applying the fix",
            f"kubectl rollout status deployment/{deployment} -n {ns}",
        ])

    elif deployment:
        commands.extend([
            f"kubectl describe deployment {deployment} -n {ns}",
            f"kubectl rollout status deployment/{deployment} -n {ns}",
        ])

    if service:
        commands.append(f"kubectl get endpoints {service} -n {ns}")

    existing = _simple_collect_existing_commands(report.get("recommended_fix"))

    for cmd in existing:
        if cmd not in commands:
            commands.append(cmd)

    return _simple_dedupe(commands)[:10]


def _simple_verification_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    commands = []

    if deployment:
        commands.append(f"kubectl rollout status deployment/{deployment} -n {ns}")

    commands.append(f"kubectl get pods -n {ns}")

    if pod:
        commands.append(f"kubectl describe pod {pod} -n {ns}")

    if service:
        commands.append(f"kubectl get endpoints {service} -n {ns}")

    commands.append(f"kubectl get events -n {ns} --sort-by=.lastTimestamp")

    existing = _simple_collect_existing_commands(report.get("verification"))

    for cmd in existing:
        if cmd not in commands:
            commands.append(cmd)

    return _simple_dedupe(commands)[:8]


def _opslens_export_markdown(report):
    report = _simple_report(report)
    affected = _simple_affected(report)

    lines = []
    lines.append(f"# {_simple_title(report)}")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Metadata / Scope")
    lines.append(f"- **Namespace:** {affected.get('namespace', '')}")
    lines.append(f"- **Node:** {affected.get('node') or affected.get('node_name') or ''}")
    lines.append(f"- **Service:** {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- **Deployment:** {affected.get('deployment') or affected.get('deployment_name') or ''}")
    lines.append(f"- **Pod:** {affected.get('pod') or affected.get('pod_name') or ''}")
    lines.append(f"- **Severity:** {report.get('severity', '')}")
    lines.append(f"- **Confidence:** {report.get('confidence', '')}")
    lines.append(f"- **Detected Time:** {report.get('generated_at') or report.get('created_at') or ''}")
    lines.append("")
    lines.append("## 2. Problem")
    lines.append(_simple_problem(report))
    lines.append("")
    lines.append("## 3. Root Cause")
    lines.append(_simple_root_cause(report))
    lines.append("")
    lines.append("## 4. Recommended Fix")
    lines.append("```bash")
    lines.extend(_simple_recommended_fix_commands(report))
    lines.append("```")
    lines.append("")
    lines.append("## 5. Verification")
    lines.append("```bash")
    lines.extend(_simple_verification_commands(report))
    lines.append("```")
    lines.append("")
    lines.append("## 6. Summary")
    lines.append(_simple_summary(report))
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _simple_report(report)
    affected = _simple_affected(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.35 * cm,
        leftMargin=1.35 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensSimpleTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=10,
    )

    heading_style = ParagraphStyle(
        "OpsLensSimpleHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=5,
    )

    body_style = ParagraphStyle(
        "OpsLensSimpleBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "OpsLensSimpleCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9.8,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsLensSimpleHeader",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9.8,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensSimpleCode",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.4,
        leading=9,
        backColor=colors.HexColor("#F3F4F6"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.4,
        borderPadding=5,
        wordWrap="CJK",
    )

    story = []

    def esc(value):
        return html.escape(_simple_clean(value)).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(esc(value), heading_style))

    def paragraph(value):
        value = _simple_clean(value)
        if value:
            story.append(Paragraph(esc(value), body_style))

    def code_block(lines):
        text = "\n".join(lines if isinstance(lines, list) else [str(lines)])
        story.append(Paragraph(esc(text), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.7 * cm

        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc(h), header_style) for h in headers]]

        for row in rows:
            data.append([Paragraph(esc(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(esc(_simple_title(report)), title_style))

    metadata_rows = [
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Pod", affected.get("pod") or affected.get("pod_name") or ""],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Detected Time", report.get("generated_at") or report.get("created_at") or ""],
    ]

    heading("1. Metadata / Scope")
    table(["Field", "Value"], metadata_rows, [4.2 * cm, 12.8 * cm])

    heading("2. Problem")
    paragraph(_simple_problem(report))

    heading("3. Root Cause")
    paragraph(_simple_root_cause(report))

    heading("4. Recommended Fix")
    code_block(_simple_recommended_fix_commands(report))

    heading("5. Verification")
    code_block(_simple_verification_commands(report))

    heading("6. Summary")
    paragraph(_simple_summary(report))

    doc.build(story)

    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens PDF command block newline fix v4
# =========================================================
# OpsLens PDF command block newline fix v4
# Keeps the simple 6-section report but preserves command newlines.
# =========================================================

def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _simple_report(report)
    affected = _simple_affected(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.35 * cm,
        leftMargin=1.35 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensSimpleTitleV4",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=10,
    )

    heading_style = ParagraphStyle(
        "OpsLensSimpleHeadingV4",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=5,
    )

    body_style = ParagraphStyle(
        "OpsLensSimpleBodyV4",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "OpsLensSimpleCellV4",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9.8,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsLensSimpleHeaderV4",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9.8,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensSimpleCodeV4",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.4,
        leading=9.2,
        backColor=colors.HexColor("#F3F4F6"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.4,
        borderPadding=6,
        wordWrap="CJK",
    )

    story = []

    def esc_inline(value):
        return html.escape(_simple_clean(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(esc_inline(value), heading_style))

    def paragraph(value):
        value = _simple_clean(value)
        if value:
            story.append(Paragraph(esc_inline(value), body_style))

    def code_block(lines):
        if isinstance(lines, list):
            text = "\n".join(str(line) for line in lines)
        else:
            text = str(lines or "")

        if not text.strip():
            text = "No command available."

        story.append(Paragraph(esc_code(text), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.7 * cm

        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc_inline(h), header_style) for h in headers]]

        for row in rows:
            data.append([Paragraph(esc_inline(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(esc_inline(_simple_title(report)), title_style))

    metadata_rows = [
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Pod", affected.get("pod") or affected.get("pod_name") or ""],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Detected Time", report.get("generated_at") or report.get("created_at") or ""],
    ]

    heading("1. Metadata / Scope")
    table(["Field", "Value"], metadata_rows, [4.2 * cm, 12.8 * cm])

    heading("2. Problem")
    paragraph(_simple_problem(report))

    heading("3. Root Cause")
    paragraph(_simple_root_cause(report))

    heading("4. Recommended Fix")
    code_block(_simple_recommended_fix_commands(report))

    heading("5. Verification")
    code_block(_simple_verification_commands(report))

    heading("6. Summary")
    paragraph(_simple_summary(report))

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens command extraction and related finding summary fix v5
# =========================================================
# OpsLens command extraction and related finding summary fix v5
# Fixes dirty combined commands like:
# kubectl get pods ...' and 'kubectl rollout status ...
# Also summarizes additional_findings without changing the 6-section report structure.
# =========================================================

import re as _opslens_cmd_re


def _simple_extract_all_kubectl(text):
    if text is None:
        return []

    raw = str(text)

    raw = raw.replace("\r", "\n")
    raw = raw.replace("`", "")
    raw = raw.replace('" and "', "\n")
    raw = raw.replace("' and '", "\n")
    raw = raw.replace("' and ", "\n")
    raw = raw.replace(" and 'kubectl", "\nkubectl")
    raw = raw.replace(' and "kubectl', "\nkubectl")

    raw = _opslens_cmd_re.sub(r"\s+and\s+(?=kubectl\s)", "\n", raw, flags=_opslens_cmd_re.I)
    raw = _opslens_cmd_re.sub(r"(?<!^)\s+(?=kubectl\s)", "\n", raw)

    commands = []

    for line in raw.splitlines():
        line = line.strip()

        if "kubectl " not in line:
            continue

        line = line[line.lower().find("kubectl "):]
        line = line.strip(" '\"`.,;")

        if not line.startswith("kubectl "):
            continue

        line = _opslens_cmd_re.sub(r"\s+", " ", line).strip()
        line = line.strip(" '\"`.,;")

        if line and line not in commands:
            commands.append(line)

    return commands


def _simple_collect_existing_commands(value):
    commands = []

    if value is None:
        return commands

    if isinstance(value, dict):
        if value.get("command"):
            commands.extend(_simple_extract_all_kubectl(value.get("command")))

        for key in ["commands", "safe_commands", "verification_commands"]:
            items = value.get(key)
            if isinstance(items, list):
                for item in items:
                    commands.extend(_simple_collect_existing_commands(item))

        for item in value.values():
            if isinstance(item, str):
                commands.extend(_simple_extract_all_kubectl(item))

    elif isinstance(value, list):
        for item in value:
            commands.extend(_simple_collect_existing_commands(item))

    else:
        commands.extend(_simple_extract_all_kubectl(value))

    return _simple_dedupe(commands)


def _simple_related_findings_text(report):
    findings = report.get("additional_findings") or []

    if not isinstance(findings, list) or not findings:
        return ""

    lines = []

    for item in findings:
        if isinstance(item, dict):
            resource = _simple_clean(item.get("resource") or "related resource")
            finding = _simple_clean(item.get("finding") or item.get("summary") or "")
            impact = _simple_clean(item.get("impact") or "")
            priority = _simple_clean(item.get("priority") or "")

            sentence = f"Related finding: {resource} — {finding}"

            if impact:
                sentence += f". Impact: {impact}"

            if priority:
                sentence += f". Priority: {priority}"

            lines.append(sentence + ".")

        else:
            text = _simple_clean(item)
            if text:
                lines.append(f"Related finding: {text}")

    return " ".join(_simple_dedupe(lines))


def _simple_summary(report):
    affected = _simple_affected(report)

    namespace = affected.get("namespace") or "the selected namespace"
    node = affected.get("node") or affected.get("node_name") or "the selected node"
    deployment = affected.get("deployment") or affected.get("deployment_name") or "the affected workload"

    problem = _simple_problem(report)
    root = _simple_root_cause(report)
    related = _simple_related_findings_text(report)

    summary = (
        f"OpsLens investigated {deployment} in namespace {namespace} on node {node}. "
        f"{problem} {root}"
    )

    if related:
        summary += " " + related

    return summary


def _simple_verification_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    commands = []

    if deployment:
        commands.append(f"kubectl rollout status deployment/{deployment} -n {ns}")

    commands.append(f"kubectl get pods -n {ns}")

    if pod:
        commands.append(f"kubectl describe pod {pod} -n {ns}")

    if service:
        commands.append(f"kubectl get endpoints {service} -n {ns}")

    commands.append(f"kubectl get events -n {ns} --sort-by=.lastTimestamp")

    for cmd in _simple_collect_existing_commands(report.get("verification")):
        if cmd not in commands:
            commands.append(cmd)

    return _simple_dedupe(commands)[:8]

# =========================================================
# OpsLens final professional report layout v7
# =========================================================
# OpsLens final professional report layout v7
# Clean sections:
# 1. Metadata / Scope
# 2. Primary Problem
# 3. Root Cause
# 4. Recommended Fix
# 5. Verification
# 6. Related Findings / Agent Status
# 7. Executive Summary
# =========================================================

def _ops_related_rows(report):
    rows = []

    findings = report.get("additional_findings") or []

    if isinstance(findings, list):
        for item in findings:
            if isinstance(item, dict):
                resource = _simple_clean(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope")
                finding = _simple_clean(item.get("finding") or item.get("summary") or item.get("anomaly_type") or "")
                impact = _simple_clean(item.get("impact") or item.get("meaning") or "Related operational signal.")
                priority = _simple_clean(item.get("priority") or item.get("severity") or "Follow-up")

                if finding:
                    rows.append([resource, finding, impact, priority])
            else:
                text = _simple_clean(item)
                if text:
                    rows.append(["scope", text, "Related operational signal.", "Follow-up"])

    contributions = report.get("agent_contributions") or {}

    if isinstance(contributions, dict):
        for name, value in contributions.items():
            if not isinstance(value, dict):
                continue

            status = _simple_clean(value.get("status") or "completed")
            finding = _simple_clean(value.get("finding") or "")

            if not finding and isinstance(value.get("findings"), list):
                parts = []
                for item in value.get("findings") or []:
                    if isinstance(item, dict):
                        parts.append(_simple_clean(item.get("summary") or item.get("anomaly_type") or item.get("finding") or ""))
                    else:
                        parts.append(_simple_clean(item))
                finding = "; ".join(part for part in parts if part)

            if not finding:
                finding = "No active signal detected."

            rows.append([str(name), finding, f"Agent status: {status}", "Info"])

    return rows[:10]


def _ops_executive_summary(report):
    affected = _simple_affected(report)

    namespace = affected.get("namespace") or "the selected namespace"
    node = affected.get("node") or affected.get("node_name") or "the selected node"
    deployment = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "the affected workload"
    )

    base = (
        f"OpsLens investigated {deployment} in namespace {namespace} on node {node}. "
        f"The primary problem is: {_simple_problem(report)} "
        f"The root cause is: {_simple_root_cause(report)}"
    )

    related_count = len(report.get("additional_findings") or [])
    if related_count:
        base += f" OpsLens also found {related_count} related follow-up signal(s), listed separately so they are not confused with the primary root cause."

    return base


def _opslens_export_markdown(report):
    report = _simple_report(report)
    affected = _simple_affected(report)
    related = _ops_related_rows(report)

    lines = []
    lines.append(f"# {_simple_title(report)}")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Metadata / Scope")
    lines.append(f"- **Namespace:** {affected.get('namespace', '')}")
    lines.append(f"- **Node:** {affected.get('node') or affected.get('node_name') or ''}")
    lines.append(f"- **Service:** {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- **Deployment:** {affected.get('deployment') or affected.get('deployment_name') or ''}")
    lines.append(f"- **Pod:** {affected.get('pod') or affected.get('pod_name') or ''}")
    lines.append(f"- **Severity:** {report.get('severity', '')}")
    lines.append(f"- **Confidence:** {report.get('confidence', '')}")
    lines.append(f"- **Detected Time:** {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or ''}")
    lines.append("")
    lines.append("## 2. Primary Problem")
    lines.append(_simple_problem(report))
    lines.append("")
    lines.append("## 3. Root Cause")
    lines.append(_simple_root_cause(report))
    lines.append("")
    lines.append("## 4. Recommended Fix")
    lines.append("```bash")
    lines.extend(_simple_recommended_fix_commands(report))
    lines.append("```")
    lines.append("")
    lines.append("## 5. Verification")
    lines.append("```bash")
    lines.extend(_simple_verification_commands(report))
    lines.append("```")
    lines.append("")

    if related:
        lines.append("## 6. Related Findings / Agent Status")
        for resource, finding, impact, priority in related:
            lines.append(f"- **{resource}** — {finding}. Impact: {impact}. Priority: {priority}.")
        lines.append("")
        lines.append("## 7. Executive Summary")
    else:
        lines.append("## 6. Executive Summary")

    lines.append(_ops_executive_summary(report))
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _simple_report(report)
    affected = _simple_affected(report)
    related = _ops_related_rows(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.35 * cm,
        leftMargin=1.35 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensProTitleV7",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=10,
    )

    heading_style = ParagraphStyle(
        "OpsLensProHeadingV7",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        spaceBefore=8,
        spaceAfter=5,
    )

    body_style = ParagraphStyle(
        "OpsLensProBodyV7",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "OpsLensProCellV7",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9.8,
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsLensProHeaderV7",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9.8,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensProCodeV7",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.4,
        leading=9.2,
        backColor=colors.HexColor("#F3F4F6"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.4,
        borderPadding=6,
        wordWrap="CJK",
    )

    story = []

    def esc_inline(value):
        return html.escape(_simple_clean(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(esc_inline(value), heading_style))

    def paragraph(value):
        value = _simple_clean(value)
        if value:
            story.append(Paragraph(esc_inline(value), body_style))

    def code_block(lines):
        if isinstance(lines, list):
            text = "\n".join(str(line) for line in lines)
        else:
            text = str(lines or "")

        if not text.strip():
            text = "No command available."

        story.append(Paragraph(esc_code(text), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.7 * cm
        if widths is None:
            widths = [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc_inline(h), header_style) for h in headers]]

        for row in rows:
            data.append([Paragraph(esc_inline(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(esc_inline(_simple_title(report)), title_style))

    metadata_rows = [
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Pod", affected.get("pod") or affected.get("pod_name") or ""],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Detected Time", report.get("generated_at") or report.get("created_at") or report.get("finished_at") or ""],
    ]

    heading("1. Metadata / Scope")
    table(["Field", "Value"], metadata_rows, [4.2 * cm, 12.8 * cm])

    heading("2. Primary Problem")
    paragraph(_simple_problem(report))

    heading("3. Root Cause")
    paragraph(_simple_root_cause(report))

    heading("4. Recommended Fix")
    code_block(_simple_recommended_fix_commands(report))

    heading("5. Verification")
    code_block(_simple_verification_commands(report))

    section_number = 6

    if related:
        heading("6. Related Findings / Agent Status")
        table(
            ["Resource / Agent", "Finding", "Impact / Status", "Priority"],
            related,
            [3.7 * cm, 5.0 * cm, 6.2 * cm, 2.1 * cm],
        )
        section_number = 7

    heading(f"{section_number}. Executive Summary")
    paragraph(_ops_executive_summary(report))

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(
        report,
        export_format,
        fallback_name=fallback_name,
    )
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens final report wording v8
# =========================================================
# OpsLens final report wording v8
# Aligns PDF/Markdown wording with the site report.
# Removes [CRITICAL] from title and explains severity as Impact Level.
# Filters noisy internal/deployment-not-found related row.
# =========================================================

def _ops_impact_label(report):
    severity = str(report.get("severity") or "unknown").lower()

    if severity == "critical":
        return "Critical - workload unavailable"
    if severity == "high":
        return "High - service degraded"
    if severity == "medium":
        return "Medium - operational risk"
    if severity == "low":
        return "Low - informational"

    return severity


def _simple_title(report):
    affected = _simple_affected(report)
    workload = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )
    return f"Incident Report: {workload}"


def _simple_problem(report):
    all_text = _simple_all_text(report).lower()
    affected = _simple_affected(report)

    pod = affected.get("pod") or affected.get("pod_name") or "the affected pod"
    deployment = affected.get("deployment") or affected.get("deployment_name") or "the deployment"

    if "readiness" in all_text and "not ready" in all_text:
        return f"{pod} is failing readiness checks. Kubernetes keeps the pod NotReady, so {deployment} has no available ready replica."

    return (
        _simple_clean(report.get("incident_summary"))
        or _simple_clean(report.get("summary"))
        or "OpsLens detected an active Kubernetes workload issue."
    )


def _simple_root_cause(report):
    all_text = _simple_all_text(report)
    lowered = all_text.lower()

    if "readiness" in lowered and ("404" in lowered or "not-found" in lowered):
        path = _simple_extract([r"(/healthz[-a-zA-Z0-9_/]*)"], all_text)

        if path:
            return f"The readiness probe is configured to call '{path}', but the application returns HTTP 404 for that endpoint. Because the health check fails, Kubernetes never marks the pod as Ready."

        return "The readiness probe returns HTTP 404. Because the health check fails, Kubernetes keeps the pod NotReady and the deployment remains unavailable."

    return (
        _simple_clean(report.get("root_cause_story"))
        or _simple_clean(report.get("root_cause_hypothesis"))
        or "OpsLens selected the most likely root cause from the correlated Kubernetes, log, configuration, and metrics signals."
    )


def _ops_related_rows(report):
    rows = []
    findings = report.get("additional_findings") or []

    if isinstance(findings, list):
        for item in findings:
            text = _simple_clean(item).lower()

            if "deploymentnot found in collected evidence" in text:
                continue
            if "not found in collected evidence status" in text:
                continue

            if isinstance(item, dict):
                resource = _simple_clean(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope")
                finding = _simple_clean(item.get("finding") or item.get("summary") or item.get("anomaly_type") or "")
                impact = _simple_clean(item.get("impact") or item.get("meaning") or "Related operational signal.")
                priority = _simple_clean(item.get("priority") or item.get("severity") or "Follow-up")

                if finding:
                    rows.append([resource, finding, impact, priority])
            else:
                raw = _simple_clean(item)
                if raw:
                    rows.append(["scope", raw, "Related operational signal.", "Follow-up"])

    return rows[:6]


def _ops_executive_summary(report):
    affected = _simple_affected(report)

    namespace = affected.get("namespace") or "the selected namespace"
    node = affected.get("node") or affected.get("node_name") or "the selected node"
    workload = (
        affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("service")
        or affected.get("service_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "the affected workload"
    )

    summary = (
        f"OpsLens investigated {workload} in namespace {namespace} on node {node}. "
        f"The primary problem is that {_simple_problem(report)} "
        f"The root cause is that {_simple_root_cause(report)}"
    )

    related_count = len(_ops_related_rows(report))
    if related_count:
        summary += f" OpsLens also found {related_count} separate follow-up signal(s), listed separately so they are not confused with the primary root cause."

    return summary


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _simple_report(report)
    affected = _simple_affected(report)
    related = _ops_related_rows(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.35 * cm,
        leftMargin=1.35 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("OpsLensTitleV8", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16, leading=20, spaceAfter=10)
    heading_style = ParagraphStyle("OpsLensHeadingV8", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, spaceBefore=8, spaceAfter=5)
    body_style = ParagraphStyle("OpsLensBodyV8", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12, spaceAfter=5)
    cell_style = ParagraphStyle("OpsLensCellV8", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=9.8, wordWrap="CJK")
    header_style = ParagraphStyle("OpsLensHeaderV8", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=9.8, textColor=colors.white, wordWrap="CJK")
    code_style = ParagraphStyle("OpsLensCodeV8", parent=styles["BodyText"], fontName="Courier", fontSize=7.4, leading=9.2, backColor=colors.HexColor("#F3F4F6"), borderColor=colors.HexColor("#CBD5E1"), borderWidth=0.4, borderPadding=6, wordWrap="CJK")

    story = []

    def esc_inline(value):
        return html.escape(_simple_clean(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(esc_inline(value), heading_style))

    def paragraph(value):
        value = _simple_clean(value)
        if value:
            story.append(Paragraph(esc_inline(value), body_style))

    def code_block(lines):
        text = "\n".join(str(line) for line in lines) if isinstance(lines, list) else str(lines or "")
        story.append(Paragraph(esc_code(text or "No command available."), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.7 * cm
        widths = widths or [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc_inline(h), header_style) for h in headers]]
        for row in rows:
            data.append([Paragraph(esc_inline(cell), cell_style) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(esc_inline(_simple_title(report)), title_style))

    metadata_rows = [
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node") or affected.get("node_name") or ""],
        ["Service", affected.get("service") or affected.get("service_name") or ""],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or ""],
        ["Pod", affected.get("pod") or affected.get("pod_name") or ""],
        ["Impact Level", _ops_impact_label(report)],
        ["Confidence", report.get("confidence", "")],
        ["Detected Time", report.get("generated_at") or report.get("created_at") or report.get("finished_at") or ""],
    ]

    heading("1. Metadata / Scope")
    table(["Field", "Value"], metadata_rows, [4.2 * cm, 12.8 * cm])

    heading("2. Primary Problem")
    paragraph(_simple_problem(report))

    heading("3. Root Cause")
    paragraph(_simple_root_cause(report))

    heading("4. Recommended Fix")
    code_block(_simple_recommended_fix_commands(report))

    heading("5. Verification")
    code_block(_simple_verification_commands(report))

    section_number = 6
    if related:
        heading("6. Related Findings / Agent Status")
        table(["Resource / Agent", "Finding", "Impact / Status", "Priority"], related, [3.7 * cm, 5.0 * cm, 6.2 * cm, 2.1 * cm])
        section_number = 7

    heading(f"{section_number}. Executive Summary")
    paragraph(_ops_executive_summary(report))

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)

# =========================================================
# OpsLens LLM-owned recommended fix guard v10
# =========================================================
# OpsLens LLM-owned recommended fix guard v10
# Principle:
# - LLM owns recommended_fix.
# - Backend validates/cleans it.
# - Backend fallback is used only if LLM fix is missing or contradicts root cause.
# =========================================================

import re as _ops_fix_re


def _ops_fix_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_ops_fix_text(v) for v in value.values() if _ops_fix_text(v))
        if isinstance(value, list):
            return " ".join(_ops_fix_text(v) for v in value if _ops_fix_text(v))
        return str(value).strip()


def _ops_fix_all_text(report):
    return _ops_fix_text(report).lower()


def _ops_fix_incident_type(report):
    text = _ops_fix_all_text(report)

    if "oomkilled" in text or "outofmemory" in text or "out of memory" in text:
        return "oom"

    if "crashloopbackoff" in text and ("memory" in text or "oom" in text):
        return "oom"

    if "podlstmresourceanomaly" in text or "predicted baseline" in text or "higher than expected" in text:
        return "pod_cpu_anomaly"

    if "readiness" in text and ("404" in text or "not-found" in text):
        return "readiness_404"

    if "readiness" in text:
        return "readiness"

    if "crashloopbackoff" in text:
        return "crashloop"

    return "generic"


def _ops_extract_kubectl_commands_from_text(value):
    if value is None:
        return []

    raw = str(value)
    raw = raw.replace("\r", "\n")
    raw = raw.replace("`", "")
    raw = raw.replace("' and '", "\n")
    raw = raw.replace('" and "', "\n")
    raw = _ops_fix_re.sub(r"\s+and\s+(?=kubectl\s)", "\n", raw, flags=_ops_fix_re.I)
    raw = _ops_fix_re.sub(r"(?<!^)\s+(?=kubectl\s)", "\n", raw)

    commands = []

    for line in raw.splitlines():
        line = line.strip()

        if not line:
            continue

        if line.startswith("#"):
            commands.append(line)
            continue

        if "kubectl " not in line:
            continue

        line = line[line.lower().find("kubectl "):]
        line = line.strip(" '\"`.,;")
        line = _ops_fix_re.sub(r"\s+", " ", line).strip()

        if line.startswith("kubectl "):
            commands.append(line)

    return commands


def _ops_collect_llm_commands(value):
    commands = []

    if value is None:
        return commands

    if isinstance(value, dict):
        for key in ["commands", "safe_commands", "verification_commands"]:
            if isinstance(value.get(key), list):
                for item in value.get(key):
                    commands.extend(_ops_collect_llm_commands(item))

        for key in ["command", "strategy", "explanation", "intent"]:
            if value.get(key):
                commands.extend(_ops_extract_kubectl_commands_from_text(value.get(key)))

        # Sometimes actions contain command/reason fields.
        if isinstance(value.get("actions"), list):
            for action in value.get("actions"):
                commands.extend(_ops_collect_llm_commands(action))

    elif isinstance(value, list):
        for item in value:
            commands.extend(_ops_collect_llm_commands(item))

    else:
        commands.extend(_ops_extract_kubectl_commands_from_text(value))

    clean = []
    seen = set()

    for cmd in commands:
        key = cmd.strip().lower()

        if not key:
            continue

        if key in seen:
            continue

        seen.add(key)
        clean.append(cmd.strip())

    return clean


def _ops_commands_match_incident(commands, incident_type):
    joined = "\n".join(commands).lower()

    if not commands:
        return False

    if incident_type == "oom":
        if "readinessprobe" in joined or "health endpoint" in joined or "/healthz" in joined:
            return False

        return (
            "oom" in joined
            or "--previous" in joined
            or "top pod" in joined
            or "describe pod" in joined
            or "edit deployment" in joined
            or "describe deployment" in joined
        )

    if incident_type == "pod_cpu_anomaly":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False

        return (
            "top pod" in joined
            or "logs" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
        )

    if incident_type in {"readiness_404", "readiness"}:
        return (
            "readiness" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
            or "rollout status" in joined
            or "get endpoints" in joined
        )

    return True


def _ops_fallback_fix_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    incident_type = _ops_fix_incident_type(report)
    commands = []

    if incident_type == "oom":
        if pod:
            commands.extend([
                "# Inspect pod restart history and OOM termination reason",
                f"kubectl describe pod {pod} -n {ns}",
                "",
                "# Review logs from the previous crashed container",
                f"kubectl logs {pod} -n {ns} --previous",
                "",
                "# Check current resource usage",
                f"kubectl top pod {pod} -n {ns}",
            ])

        if deployment:
            commands.extend([
                "",
                "# Inspect memory requests and limits",
                f"kubectl describe deployment {deployment} -n {ns}",
                "",
                "# Update memory limits or fix the memory-heavy workload",
                f"kubectl edit deployment {deployment} -n {ns}",
                "",
                "# Verify rollout after the change",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        return [cmd for cmd in commands if cmd is not None]

    if incident_type == "pod_cpu_anomaly":
        if pod:
            commands.extend([
                "# Confirm current CPU usage",
                f"kubectl top pod {pod} -n {ns}",
                "",
                "# Inspect pod events and runtime configuration",
                f"kubectl describe pod {pod} -n {ns}",
                "",
                "# Review logs for runaway loop, batch job, or traffic spike",
                f"kubectl logs {pod} -n {ns} --tail=120",
            ])

        if deployment:
            commands.extend([
                "",
                "# Inspect deployment resources and scaling settings",
                f"kubectl describe deployment {deployment} -n {ns}",
            ])

        return [cmd for cmd in commands if cmd is not None]

    if incident_type in {"readiness_404", "readiness"}:
        if deployment:
            commands.extend([
                "# Inspect the current readinessProbe configuration",
                f"kubectl describe deployment {deployment} -n {ns}",
                "",
                "# Edit the readinessProbe path/port to match the real application health endpoint",
                f"kubectl edit deployment {deployment} -n {ns}",
                "",
                "# Watch the rollout after applying the fix",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        if service:
            commands.append(f"kubectl get endpoints {service} -n {ns}")

        return [cmd for cmd in commands if cmd is not None]

    if deployment:
        return [
            f"kubectl describe deployment {deployment} -n {ns}",
            f"kubectl rollout status deployment/{deployment} -n {ns}",
        ]

    if pod:
        return [
            f"kubectl describe pod {pod} -n {ns}",
            f"kubectl logs {pod} -n {ns} --tail=120",
        ]

    return [f"kubectl get pods -n {ns}"]


def _simple_recommended_fix_commands(report):
    incident_type = _ops_fix_incident_type(report)

    llm_commands = _ops_collect_llm_commands(report.get("recommended_fix"))

    if _ops_commands_match_incident(llm_commands, incident_type):
        return llm_commands[:12]

    fallback = _ops_fallback_fix_commands(report)
    return fallback[:12]


def _simple_verification_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    incident_type = _ops_fix_incident_type(report)

    commands = []

    llm_commands = _ops_collect_llm_commands(report.get("verification"))

    # Verification commands are usually safe; keep relevant LLM commands first.
    for cmd in llm_commands:
        lowered = cmd.lower()

        if not lowered.startswith("kubectl "):
            continue

        if "delete " in lowered:
            continue

        if cmd not in commands:
            commands.append(cmd)

    if deployment:
        commands.append(f"kubectl rollout status deployment/{deployment} -n {ns}")

    commands.append(f"kubectl get pods -n {ns}")

    if pod:
        commands.append(f"kubectl describe pod {pod} -n {ns}")

        if incident_type in {"oom", "pod_cpu_anomaly"}:
            commands.append(f"kubectl top pod {pod} -n {ns}")
            commands.append(f"kubectl logs {pod} -n {ns} --tail=120")

    if service:
        commands.append(f"kubectl get endpoints {service} -n {ns}")

    commands.append(f"kubectl get events -n {ns} --sort-by=.lastTimestamp")

    clean = []
    seen = set()

    for cmd in commands:
        key = cmd.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        clean.append(cmd.strip())

    return clean[:10]

# =========================================================
# OpsLens enterprise recommended fix guard v12
# =========================================================
# OpsLens enterprise recommended fix guard v12
# LLM owns the fix. Backend validates it against the PRIMARY incident only.
# Unrelated related_findings must not change the primary remediation plan.
# =========================================================

import re as _ops_v12_re


def _ops_v12_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_ops_v12_text(v) for v in value.values() if _ops_v12_text(v))
        if isinstance(value, list):
            return " ".join(_ops_v12_text(v) for v in value if _ops_v12_text(v))
        return str(value).strip()


def _ops_primary_evidence_text(report):
    """
    Classify from PRIMARY evidence only.
    Do not classify from recommended_fix, verification, or unrelated findings.
    """
    if not isinstance(report, dict):
        return _ops_v12_text(report).lower()

    primary_parts = []

    for key in [
        "title",
        "incident_summary",
        "root_cause_story",
        "root_cause_hypothesis",
        "primary_incident_group",
        "root_cause_facts",
        "agent_reasoning",
        "evidence_trail",
        "affected_resources",
    ]:
        if key in report:
            primary_parts.append(_ops_v12_text(report.get(key)))

    return " ".join(primary_parts).lower()


def _ops_fix_incident_type(report):
    text = _ops_primary_evidence_text(report)

    if (
        ("selector" in text and ("endpoint" in text or "service" in text))
        or "empty endpoints" in text
        or "endpoint list remains empty" in text
        or "service has no endpoints" in text
    ):
        return "service_selector"

    if "targetport" in text or "target port" in text:
        return "service_targetport"

    if (
        "imagepullbackoff" in text
        or "errimagepull" in text
        or "invalid registry" in text
        or "failed to pull image" in text
    ):
        return "image_pull"

    if (
        "missing env" in text
        or "environment variable" in text
        or "missing environment" in text
        or "keyerror" in text
    ):
        return "missing_env"

    if "python traceback" in text or "traceback" in text:
        return "app_traceback"

    if "oomkilled" in text or "outofmemory" in text or "out of memory" in text:
        return "oom"

    if (
        "podlstmresourceanomaly" in text
        or "predicted baseline" in text
        or "higher than predicted" in text
        or "higher than expected" in text
        or "lstm metrics agent" in text
    ):
        return "pod_cpu_anomaly"

    if "readiness" in text and ("404" in text or "not-found" in text):
        return "readiness_404"

    if "readiness" in text:
        return "readiness"

    if "crashloopbackoff" in text:
        return "crashloop"

    return "generic"


def _ops_extract_kubectl_commands_from_text(value):
    if value is None:
        return []

    raw = str(value)
    raw = raw.replace("\r", "\n")
    raw = raw.replace("`", "")
    raw = raw.replace("' and '", "\n")
    raw = raw.replace('" and "', "\n")
    raw = _ops_v12_re.sub(r"\s+and\s+(?=kubectl\s)", "\n", raw, flags=_ops_v12_re.I)
    raw = _ops_v12_re.sub(r"(?<!^)\s+(?=kubectl\s)", "\n", raw)

    commands = []

    for line in raw.splitlines():
        line = line.strip()

        if not line:
            continue

        if line.startswith("#"):
            commands.append(line)
            continue

        if "kubectl " not in line:
            continue

        line = line[line.lower().find("kubectl "):]
        line = line.strip(" '\"`.,;")
        line = _ops_v12_re.sub(r"\s+", " ", line).strip()

        if line.startswith("kubectl "):
            commands.append(line)

    clean = []
    seen = set()

    for cmd in commands:
        key = cmd.lower().strip()
        if key and key not in seen:
            seen.add(key)
            clean.append(cmd)

    return clean


def _ops_collect_llm_commands(value):
    commands = []

    if value is None:
        return commands

    if isinstance(value, dict):
        for key in ["commands", "safe_commands", "verification_commands"]:
            if isinstance(value.get(key), list):
                for item in value.get(key):
                    commands.extend(_ops_collect_llm_commands(item))

        for key in ["command", "strategy", "explanation", "intent"]:
            if value.get(key):
                commands.extend(_ops_extract_kubectl_commands_from_text(value.get(key)))

        if isinstance(value.get("actions"), list):
            for action in value.get("actions"):
                commands.extend(_ops_collect_llm_commands(action))

    elif isinstance(value, list):
        for item in value:
            commands.extend(_ops_collect_llm_commands(item))

    else:
        commands.extend(_ops_extract_kubectl_commands_from_text(value))

    clean = []
    seen = set()

    for cmd in commands:
        key = cmd.lower().strip()
        if key and key not in seen:
            seen.add(key)
            clean.append(cmd.strip())

    return clean


def _ops_commands_match_incident(commands, incident_type):
    if not commands:
        return False

    joined = "\n".join(commands).lower()

    if "delete " in joined:
        return False

    if incident_type == "service_selector":
        return (
            "get svc" in joined
            or "describe svc" in joined
            or "get endpoints" in joined
            or "show-labels" in joined
            or "edit svc" in joined
        )

    if incident_type == "service_targetport":
        return (
            "get svc" in joined
            or "describe svc" in joined
            or "get endpoints" in joined
            or "edit svc" in joined
            or "targetport" in joined
        )

    if incident_type == "image_pull":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False
        return (
            "describe pod" in joined
            or "get events" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
        )

    if incident_type in {"missing_env", "app_traceback"}:
        if "readinessprobe" in joined or "/healthz" in joined:
            return False
        return (
            "logs" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
            or "configmap" in joined
            or "secret" in joined
            or "edit deployment" in joined
        )

    if incident_type == "oom":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False
        return (
            "--previous" in joined
            or "top pod" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
        )

    if incident_type == "pod_cpu_anomaly":
        if "oom" in joined or "--previous" in joined or "readinessprobe" in joined:
            return False
        return (
            "top pod" in joined
            or "logs" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
        )

    if incident_type in {"readiness_404", "readiness"}:
        return (
            "readiness" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
            or "rollout status" in joined
            or "get endpoints" in joined
        )

    return True


def _ops_fallback_fix_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    incident_type = _ops_fix_incident_type(report)
    commands = []

    if incident_type == "service_selector":
        if service:
            commands.extend([
                "# Inspect the Service selector and endpoints",
                f"kubectl get svc {service} -n {ns} -o yaml",
                f"kubectl describe svc {service} -n {ns}",
                f"kubectl get endpoints {service} -n {ns}",
                "",
            ])

        commands.extend([
            "# Compare Service selector labels with Pod labels",
            f"kubectl get pods -n {ns} --show-labels",
        ])

        if service:
            commands.extend([
                "",
                "# Fix the Service selector so it matches the intended backend Pods",
                f"kubectl edit svc {service} -n {ns}",
                "",
                "# Verify endpoints are populated after the selector fix",
                f"kubectl get endpoints {service} -n {ns}",
            ])

        return commands[:12]

    if incident_type == "service_targetport":
        if service:
            commands.extend([
                f"kubectl get svc {service} -n {ns} -o yaml",
                f"kubectl describe svc {service} -n {ns}",
                f"kubectl get endpoints {service} -n {ns}",
                f"kubectl edit svc {service} -n {ns}",
                f"kubectl get endpoints {service} -n {ns}",
            ])
        return commands[:10]

    if incident_type == "image_pull":
        if pod:
            commands.extend([
                "# Inspect image pull error details",
                f"kubectl describe pod {pod} -n {ns}",
                f"kubectl get events -n {ns} --sort-by=.lastTimestamp",
            ])

        if deployment:
            commands.extend([
                "",
                "# Correct image name, tag, registry, or imagePullSecret",
                f"kubectl describe deployment {deployment} -n {ns}",
                f"kubectl edit deployment {deployment} -n {ns}",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        return commands[:10]

    if incident_type in {"missing_env", "app_traceback"}:
        if pod:
            commands.extend([
                "# Inspect application logs and traceback",
                f"kubectl describe pod {pod} -n {ns}",
                f"kubectl logs {pod} -n {ns} --tail=160",
            ])

        if deployment:
            commands.extend([
                "",
                "# Inspect environment variables, ConfigMaps, Secrets, and deployment configuration",
                f"kubectl describe deployment {deployment} -n {ns}",
                f"kubectl get configmap -n {ns}",
                f"kubectl get secret -n {ns}",
                "",
                "# Add the missing env/config value or fix application configuration",
                f"kubectl edit deployment {deployment} -n {ns}",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        return commands[:12]

    if incident_type == "oom":
        if pod:
            commands.extend([
                "# Inspect pod restart history and OOM termination reason",
                f"kubectl describe pod {pod} -n {ns}",
                "",
                "# Review logs from the previous crashed container",
                f"kubectl logs {pod} -n {ns} --previous",
                "",
                "# Check current resource usage",
                f"kubectl top pod {pod} -n {ns}",
            ])

        if deployment:
            commands.extend([
                "",
                "# Inspect memory requests and limits",
                f"kubectl describe deployment {deployment} -n {ns}",
                "",
                "# Update memory limits or fix the memory-heavy workload",
                f"kubectl edit deployment {deployment} -n {ns}",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        return commands[:12]

    if incident_type == "pod_cpu_anomaly":
        if pod:
            commands.extend([
                "# Confirm current CPU usage",
                f"kubectl top pod {pod} -n {ns}",
                "",
                "# Inspect pod events and runtime configuration",
                f"kubectl describe pod {pod} -n {ns}",
                "",
                "# Review logs for runaway loop, batch job, or traffic spike",
                f"kubectl logs {pod} -n {ns} --tail=120",
            ])

        if deployment:
            commands.extend([
                "",
                "# Inspect deployment resources and scaling settings",
                f"kubectl describe deployment {deployment} -n {ns}",
            ])

        return commands[:10]

    if incident_type in {"readiness_404", "readiness"}:
        if deployment:
            commands.extend([
                "# Inspect the current readinessProbe configuration",
                f"kubectl describe deployment {deployment} -n {ns}",
                "",
                "# Edit the readinessProbe path/port to match the real application health endpoint",
                f"kubectl edit deployment {deployment} -n {ns}",
                "",
                "# Watch the rollout after applying the fix",
                f"kubectl rollout status deployment/{deployment} -n {ns}",
            ])

        if service:
            commands.append(f"kubectl get endpoints {service} -n {ns}")

        return commands[:10]

    if deployment:
        return [
            f"kubectl describe deployment {deployment} -n {ns}",
            f"kubectl rollout status deployment/{deployment} -n {ns}",
        ]

    if pod:
        return [
            f"kubectl describe pod {pod} -n {ns}",
            f"kubectl logs {pod} -n {ns} --tail=120",
        ]

    return [f"kubectl get pods -n {ns}"]


def _simple_recommended_fix_commands(report):
    incident_type = _ops_fix_incident_type(report)
    llm_commands = _ops_collect_llm_commands(report.get("recommended_fix"))

    if _ops_commands_match_incident(llm_commands, incident_type):
        return llm_commands[:12]

    return _ops_fallback_fix_commands(report)[:12]


def _simple_verification_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    incident_type = _ops_fix_incident_type(report)
    commands = []

    llm_commands = _ops_collect_llm_commands(report.get("verification"))

    for cmd in llm_commands:
        lowered = cmd.lower()
        if not lowered.startswith("kubectl "):
            continue
        if "delete " in lowered:
            continue
        if cmd not in commands:
            commands.append(cmd)

    if service and incident_type in {"service_selector", "service_targetport"}:
        commands.append(f"kubectl get endpoints {service} -n {ns}")
        commands.append(f"kubectl get svc {service} -n {ns} -o yaml")

    if deployment:
        commands.append(f"kubectl rollout status deployment/{deployment} -n {ns}")

    commands.append(f"kubectl get pods -n {ns}")

    if pod:
        commands.append(f"kubectl describe pod {pod} -n {ns}")

        if incident_type == "oom":
            commands.append(f"kubectl logs {pod} -n {ns} --previous")
            commands.append(f"kubectl top pod {pod} -n {ns}")

        if incident_type == "pod_cpu_anomaly":
            commands.append(f"kubectl top pod {pod} -n {ns}")
            commands.append(f"kubectl logs {pod} -n {ns} --tail=120")

        if incident_type in {"missing_env", "app_traceback"}:
            commands.append(f"kubectl logs {pod} -n {ns} --tail=160")

    commands.append(f"kubectl get events -n {ns} --sort-by=.lastTimestamp")

    clean = []
    seen = set()

    for cmd in commands:
        key = cmd.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        clean.append(cmd.strip())

    return clean[:10]

# =========================================================
# OpsLens Enterprise PDF Renderer v13
# =========================================================
# OpsLens Enterprise PDF Renderer v13
# Enterprise report layout:
# Header / Executive Summary / Primary Incident /
# Related Findings / Remediation & Verification / Diagnostic Summary
# =========================================================

def _ent_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_ent_text(v) for v in value.values() if _ent_text(v))
        if isinstance(value, list):
            return " ".join(_ent_text(v) for v in value if _ent_text(v))
        return str(value).strip()


def _ent_report(report):
    try:
        return _simple_report(report)
    except Exception:
        return report if isinstance(report, dict) else {}


def _ent_affected(report):
    try:
        return _simple_affected(report)
    except Exception:
        affected = report.get("affected_resources") if isinstance(report, dict) else {}
        return affected if isinstance(affected, dict) else {}


def _ent_impact_label(report):
    try:
        return _ops_impact_label(report)
    except Exception:
        severity = str(report.get("severity") or "unknown").lower()
        if severity == "critical":
            return "Critical - workload unavailable"
        if severity == "warning":
            return "Warning - degraded or anomalous"
        if severity == "info":
            return "Info - informational"
        return severity


def _ent_title(report):
    affected = _ent_affected(report)
    workload = (
        affected.get("service")
        or affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )
    return f"Incident Report: {workload}"


def _ent_problem(report):
    try:
        return _simple_problem(report)
    except Exception:
        return _ent_text(report.get("incident_summary")) or "Analysis Inconclusive."


def _ent_root_cause(report):
    try:
        return _simple_root_cause(report)
    except Exception:
        return _ent_text(report.get("root_cause_story")) or "Analysis Inconclusive."


def _ent_sentences(text):
    import re
    clean = _ent_text(text)
    if not clean:
        return []
    parts = re.split(r"(?<=[.!?])\s+", clean)
    return [p.strip() for p in parts if p.strip()]


def _ent_executive_summary(report):
    affected = _ent_affected(report)
    workload = (
        affected.get("service")
        or affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "the workload"
    )
    namespace = affected.get("namespace") or "the selected namespace"

    failure = _ent_problem(report)
    root = _ent_root_cause(report)
    impact = f"The operational impact is scoped to {workload} in namespace {namespace}, with impact level: {_ent_impact_label(report)}."

    return [
        f"Failure: {failure}",
        f"Root Cause: {root}",
        impact,
    ]


def _ent_related_rows(report):
    rows = []
    findings = report.get("additional_findings") or report.get("separate_findings") or []

    if isinstance(findings, dict):
        findings = [findings]

    if isinstance(findings, list):
        for item in findings:
            text = _ent_text(item).lower()

            if "agentexecutionerror" in text or "podlstmunavailable" in text:
                continue

            if "deploymentnot found in collected evidence" in text:
                continue

            if isinstance(item, dict):
                resource = _ent_text(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope")
                finding = _ent_text(item.get("finding") or item.get("summary") or item.get("anomaly_type") or "")
                impact = _ent_text(item.get("impact") or item.get("meaning") or "Related operational signal.")
                priority = _ent_text(item.get("priority") or item.get("severity") or "Follow-up")

                if finding:
                    rows.append([resource, finding, f"{impact} / {priority}"])
            else:
                raw = _ent_text(item)
                if raw:
                    rows.append(["scope", raw, "Related operational signal / Follow-up"])

    return rows[:8]


def _ent_cli_only(lines, limit):
    clean = []
    seen = set()

    for line in lines or []:
        item = str(line or "").strip()

        if not item.startswith("kubectl "):
            continue

        key = item.lower()

        if key in seen:
            continue

        seen.add(key)
        clean.append(item)

        if len(clean) >= limit:
            break

    return clean


def _ent_remediation_commands(report):
    try:
        lines = _simple_recommended_fix_commands(report)
    except Exception:
        lines = []
    return _ent_cli_only(lines, 3)


def _ent_verification_commands(report):
    try:
        lines = _simple_verification_commands(report)
    except Exception:
        lines = []
    return _ent_cli_only(lines, 2)


def _ent_diagnostic_rows(report):
    affected = _ent_affected(report)
    rows = []

    if affected.get("service") or affected.get("service_name"):
        rows.append(["Service", affected.get("service") or affected.get("service_name"), "In investigation scope"])

    if affected.get("deployment") or affected.get("deployment_name"):
        rows.append(["Deployment", affected.get("deployment") or affected.get("deployment_name"), "Affected workload"])

    if affected.get("pod") or affected.get("pod_name"):
        rows.append(["Pod", affected.get("pod") or affected.get("pod_name"), "Affected pod"])

    agent_status = report.get("agent_run_status") or []

    if isinstance(agent_status, list):
        for item in agent_status[:8]:
            if not isinstance(item, dict):
                continue

            agent = _ent_text(item.get("agent") or "agent")
            status = _ent_text(item.get("status") or "completed")
            finding = _ent_text(item.get("finding") or "No active signal detected.")
            event_count = item.get("event_count", 0)

            rows.append([f"Agent: {agent}", f"{status} ({event_count} signal(s))", finding])

    return rows[:12]


def _opslens_export_markdown(report):
    report = _ent_report(report)
    affected = _ent_affected(report)

    remediation = _ent_remediation_commands(report)
    verification = _ent_verification_commands(report)
    related = _ent_related_rows(report)
    diagnostic = _ent_diagnostic_rows(report)

    lines = []
    lines.append(f"# {_ent_title(report)}")
    lines.append("")
    lines.append(f"**Namespace:** {affected.get('namespace', '')}  ")
    lines.append(f"**Node:** {affected.get('node') or affected.get('node_name') or ''}  ")
    lines.append(f"**Impact:** {_ent_impact_label(report)}  ")
    lines.append(f"**Confidence:** {report.get('confidence', '')}  ")
    lines.append(f"**Time:** {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or ''}")
    lines.append("")
    lines.append("## Executive Summary")
    for sentence in _ent_executive_summary(report):
        lines.append(f"- {sentence}")
    lines.append("")
    lines.append("## Primary Incident")
    lines.append(f"**Root Cause:** {_ent_root_cause(report)}")
    lines.append("")
    lines.append(f"**Why it is the blocker:** {_ent_problem(report)}")
    lines.append("")
    lines.append("## Related Findings")
    if related:
        lines.append("| Resource | Finding | Impact/Priority |")
        lines.append("| --- | --- | --- |")
        for row in related:
            lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    else:
        lines.append("No separate unrelated findings were identified.")
    lines.append("")
    lines.append("## Recommended Actions & Validation")
    lines.append("### Recommended Next Actions")
    lines.append("```bash")
    lines.extend(remediation or ["# Analysis Inconclusive: no safe remediation command available."])
    lines.append("```")
    lines.append("")
    lines.append("### Validation Checks")
    lines.append("```bash")
    lines.extend(verification or ["# Analysis Inconclusive: no verification command available."])
    lines.append("```")
    lines.append("")
    lines.append("## Diagnostic Summary")
    lines.append("| Type | Resource / Agent | Status |")
    lines.append("| --- | --- | --- |")
    for row in diagnostic:
        lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    report = _ent_report(report)
    affected = _ent_affected(report)

    remediation = _ent_remediation_commands(report)
    verification = _ent_verification_commands(report)
    related = _ent_related_rows(report)
    diagnostic = _ent_diagnostic_rows(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensEnterpriseTitleV13",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=21,
        alignment=1,
        spaceAfter=7,
    )

    subtitle_style = ParagraphStyle(
        "OpsLensEnterpriseSubtitleV13",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        alignment=1,
        textColor=colors.HexColor("#475569"),
        spaceAfter=12,
    )

    h_style = ParagraphStyle(
        "OpsLensEnterpriseHeadingV13",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        spaceBefore=8,
        spaceAfter=5,
        textColor=colors.HexColor("#111827"),
    )

    body = ParagraphStyle(
        "OpsLensEnterpriseBodyV13",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.7,
        leading=11.3,
        spaceAfter=4,
    )

    small = ParagraphStyle(
        "OpsLensEnterpriseSmallV13",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9.4,
        wordWrap="CJK",
    )

    small_bold = ParagraphStyle(
        "OpsLensEnterpriseSmallBoldV13",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.4,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensEnterpriseCodeV13",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.6,
        leading=9.5,
        backColor=colors.HexColor("#F8FAFC"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.35,
        borderPadding=6,
        wordWrap="CJK",
    )

    story = []

    def esc(value):
        return html.escape(_ent_text(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 4))
        story.append(Paragraph(esc(value), h_style))

    def p(value):
        value = _ent_text(value)
        if value:
            story.append(Paragraph(esc(value), body))

    def code_block(lines):
        text = "\n".join(lines or ["# Analysis Inconclusive"])
        story.append(Paragraph(esc_code(text), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.5 * cm
        widths = widths or [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc(h), small_bold) for h in headers]]
        for row in rows:
            data.append([Paragraph(esc(cell), small) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 7))

    story.append(Paragraph(esc(_ent_title(report)), title_style))

    subtitle = (
        f"{affected.get('namespace', '')} | "
        f"Node: {affected.get('node') or affected.get('node_name') or '-'} | "
        f"Impact: {_ent_impact_label(report)} | "
        f"Confidence: {report.get('confidence', '-')} | "
        f"Time: {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or '-'}"
    )
    story.append(Paragraph(esc(subtitle), subtitle_style))

    heading("Executive Summary")
    summary_rows = [[item] for item in _ent_executive_summary(report)]
    table(["Summary"], summary_rows, [A4[0] - 2.5 * cm])

    heading("Primary Incident")
    table(
        ["Field", "Details"],
        [
            ["Root Cause", _ent_root_cause(report)],
            ["Why it is the blocker", _ent_problem(report)],
        ],
        [4.3 * cm, 12.7 * cm],
    )

    heading("Related Findings")
    if related:
        table(["Resource", "Finding", "Impact / Priority"], related, [4.4 * cm, 6.1 * cm, 6.5 * cm])
    else:
        p("No separate unrelated findings were identified.")

    heading("Remediation & Verification")
    table(
        ["Plan", "Commands"],
        [
            ["Recommended Next Actions", "\n".join(remediation or ["Analysis Inconclusive"])],
            ["Validation Checks", "\n".join(verification or ["Analysis Inconclusive"])],
        ],
        [4.3 * cm, 12.7 * cm],
    )

    heading("Diagnostic Summary")
    table(["Type", "Resource / Agent", "Status"], diagnostic or [["Scope", "No diagnostic details", "Analysis Inconclusive"]], [4.0 * cm, 5.5 * cm, 7.5 * cm])

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(report, export_format, fallback_name=fallback_name)
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens Enterprise PDF Renderer v14
# =========================================================
# OpsLens Enterprise PDF Renderer v14
# Final professional PDF layout:
# - branded header
# - metadata table instead of pipe subtitle
# - remediation/verification as real code blocks
# - diagnostic summary with real agent findings
# =========================================================

def _ent_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_ent_text(v) for v in value.values() if _ent_text(v))
        if isinstance(value, list):
            return " ".join(_ent_text(v) for v in value if _ent_text(v))
        return str(value).strip()


def _ent_report(report):
    try:
        return _simple_report(report)
    except Exception:
        return report if isinstance(report, dict) else {}


def _ent_affected(report):
    try:
        return _simple_affected(report)
    except Exception:
        affected = report.get("affected_resources") if isinstance(report, dict) else {}
        return affected if isinstance(affected, dict) else {}


def _ent_impact_label(report):
    try:
        return _ops_impact_label(report)
    except Exception:
        severity = str(report.get("severity") or "unknown").lower()

        if severity == "critical":
            return "Critical - workload unavailable"
        if severity == "warning":
            return "Warning - degraded or anomalous"
        if severity == "info":
            return "Info - informational"

        return severity or "unknown"


def _ent_title(report):
    affected = _ent_affected(report)
    workload = (
        affected.get("service")
        or affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )
    return f"Incident Report: {workload}"


def _ent_problem(report):
    try:
        return _simple_problem(report)
    except Exception:
        return _ent_text(report.get("incident_summary")) or "Analysis Inconclusive."


def _ent_root_cause(report):
    try:
        return _simple_root_cause(report)
    except Exception:
        return _ent_text(report.get("root_cause_story")) or "Analysis Inconclusive."


def _ent_executive_summary(report):
    affected = _ent_affected(report)

    workload = (
        affected.get("service")
        or affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "the workload"
    )

    namespace = affected.get("namespace") or "the selected namespace"

    return [
        f"Failure: {_ent_problem(report)}",
        f"Root Cause: {_ent_root_cause(report)}",
        f"Operational Impact: Impact is scoped to {workload} in namespace {namespace}. Impact level: {_ent_impact_label(report)}.",
    ]


def _ent_related_rows(report):
    rows = []
    findings = report.get("additional_findings") or report.get("separate_findings") or []

    if isinstance(findings, dict):
        findings = [findings]

    if isinstance(findings, list):
        for item in findings:
            text = _ent_text(item).lower()

            if "agentexecutionerror" in text or "podlstmunavailable" in text:
                continue

            if "deploymentnot found in collected evidence" in text:
                continue

            if isinstance(item, dict):
                resource = _ent_text(
                    item.get("resource")
                    or item.get("pod_name")
                    or item.get("deployment_name")
                    or "scope"
                )
                finding = _ent_text(
                    item.get("finding")
                    or item.get("summary")
                    or item.get("anomaly_type")
                    or ""
                )
                impact = _ent_text(
                    item.get("impact")
                    or item.get("meaning")
                    or "Related operational signal."
                )
                priority = _ent_text(item.get("priority") or item.get("severity") or "Follow-up")

                if finding:
                    rows.append([resource, finding, f"{impact} / {priority}"])
            else:
                raw = _ent_text(item)
                if raw:
                    rows.append(["scope", raw, "Related operational signal / Follow-up"])

    return rows[:8]


def _ent_cli_only(lines, limit):
    clean = []
    seen = set()

    for line in lines or []:
        item = str(line or "").strip()

        if not item.startswith("kubectl "):
            continue

        key = item.lower()
        if key in seen:
            continue

        seen.add(key)
        clean.append(item)

        if len(clean) >= limit:
            break

    return clean


def _ent_remediation_commands(report):
    try:
        lines = _simple_recommended_fix_commands(report)
    except Exception:
        lines = []
    return _ent_cli_only(lines, 3)


def _ent_verification_commands(report):
    try:
        lines = _simple_verification_commands(report)
    except Exception:
        lines = []
    return _ent_cli_only(lines, 2)


def _ent_agent_finding_from_contribution(value):
    if not isinstance(value, dict):
        return "No active signal detected."

    finding = _ent_text(value.get("finding"))

    findings = value.get("findings")

    if isinstance(findings, list) and findings:
        parts = []

        for item in findings[:2]:
            if isinstance(item, dict):
                anomaly = _ent_text(item.get("anomaly_type") or item.get("finding") or "Signal")
                summary = _ent_text(item.get("summary") or item.get("meaning") or "")
                if summary:
                    parts.append(f"{anomaly}: {summary}")
                else:
                    parts.append(anomaly)
            else:
                parts.append(_ent_text(item))

        return " | ".join(part for part in parts if part)

    event_count = value.get("event_count", 0)

    if event_count and (not finding or finding.lower() == "no active signal detected."):
        return f"{event_count} signal(s) detected."

    return finding or "No active signal detected."


def _ent_diagnostic_rows(report):
    affected = _ent_affected(report)
    rows = []

    if affected.get("service") or affected.get("service_name"):
        rows.append(["Service", affected.get("service") or affected.get("service_name"), "In investigation scope"])

    if affected.get("deployment") or affected.get("deployment_name"):
        rows.append(["Deployment", affected.get("deployment") or affected.get("deployment_name"), "Affected workload"])

    if affected.get("pod") or affected.get("pod_name"):
        rows.append(["Pod", affected.get("pod") or affected.get("pod_name"), "Affected pod"])

    contributions = report.get("agent_contributions") or {}

    if isinstance(contributions, dict) and contributions:
        for agent_name, value in list(contributions.items())[:8]:
            if isinstance(value, dict):
                status = _ent_text(value.get("status") or "completed")
                event_count = value.get("event_count", 0)
                finding = _ent_agent_finding_from_contribution(value)
            else:
                status = "completed"
                event_count = 0
                finding = "No active signal detected."

            rows.append([f"Agent: {agent_name}", f"{status} ({event_count} signal(s))", finding])

        return rows[:12]

    agent_status = report.get("agent_run_status") or []

    if isinstance(agent_status, list):
        for item in agent_status[:8]:
            if not isinstance(item, dict):
                continue

            agent = _ent_text(item.get("agent") or "agent")
            status = _ent_text(item.get("status") or "completed")
            finding = _ent_text(item.get("finding") or "No active signal detected.")
            event_count = item.get("event_count", 0)

            rows.append([f"Agent: {agent}", f"{status} ({event_count} signal(s))", finding])

    return rows[:12]


def _ent_logo_path():
    candidates = [
        "web/assets/logo.png",
        "web/assets/opslens-logo.png",
        "web/assets/opslens_ai_logo.png",
        "web/logo.png",
        "web/opslens-logo.png",
        "static/logo.png",
    ]

    for item in candidates:
        p = Path(item)
        if p.exists():
            return str(p)

    return None


def _opslens_export_markdown(report):
    report = _ent_report(report)
    affected = _ent_affected(report)

    remediation = _ent_remediation_commands(report)
    verification = _ent_verification_commands(report)
    related = _ent_related_rows(report)
    diagnostic = _ent_diagnostic_rows(report)

    lines = []
    lines.append(f"# {_ent_title(report)}")
    lines.append("")
    lines.append("## Header")
    lines.append(f"- **Namespace:** {affected.get('namespace', '')}")
    lines.append(f"- **Node:** {affected.get('node') or affected.get('node_name') or ''}")
    lines.append(f"- **Service:** {affected.get('service') or affected.get('service_name') or ''}")
    lines.append(f"- **Deployment:** {affected.get('deployment') or affected.get('deployment_name') or ''}")
    lines.append(f"- **Pod:** {affected.get('pod') or affected.get('pod_name') or ''}")
    lines.append(f"- **Impact:** {_ent_impact_label(report)}")
    lines.append(f"- **Confidence:** {report.get('confidence', '')}")
    lines.append(f"- **Time:** {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or ''}")
    lines.append("")
    lines.append("## Executive Summary")
    for sentence in _ent_executive_summary(report):
        lines.append(f"- {sentence}")
    lines.append("")
    lines.append("## Primary Incident")
    lines.append(f"**Root Cause:** {_ent_root_cause(report)}")
    lines.append("")
    lines.append(f"**Why it is the blocker:** {_ent_problem(report)}")
    lines.append("")
    lines.append("## Related Findings")
    if related:
        lines.append("| Resource | Finding | Impact/Priority |")
        lines.append("| --- | --- | --- |")
        for row in related:
            lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    else:
        lines.append("No separate unrelated findings were identified.")
    lines.append("")
    lines.append("## Recommended Actions & Validation")
    lines.append("### Recommended Next Actions")
    lines.append("```bash")
    lines.extend(remediation or ["# Analysis Inconclusive: no safe remediation command available."])
    lines.append("```")
    lines.append("")
    lines.append("### Validation Checks")
    lines.append("```bash")
    lines.extend(verification or ["# Analysis Inconclusive: no verification command available."])
    lines.append("```")
    lines.append("")
    lines.append("## Diagnostic Summary")
    lines.append("| Type | Resource / Agent | Status |")
    lines.append("| --- | --- | --- |")
    for row in diagnostic:
        lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, Image

    report = _ent_report(report)
    affected = _ent_affected(report)

    remediation = _ent_remediation_commands(report)
    verification = _ent_verification_commands(report)
    related = _ent_related_rows(report)
    diagnostic = _ent_diagnostic_rows(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=0.85 * cm,
        bottomMargin=0.95 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsLensEnterpriseTitleV14",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16.5,
        leading=20,
        alignment=0,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=2,
    )

    brand_style = ParagraphStyle(
        "OpsLensBrandV14",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=15,
        textColor=colors.HexColor("#DC2626"),
    )

    subtitle_style = ParagraphStyle(
        "OpsLensSubtitleV14",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#64748B"),
    )

    h_style = ParagraphStyle(
        "OpsLensHeadingV14",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11.2,
        leading=13.5,
        spaceBefore=7,
        spaceAfter=5,
        textColor=colors.HexColor("#111827"),
    )

    body = ParagraphStyle(
        "OpsLensBodyV14",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        spaceAfter=4,
    )

    small = ParagraphStyle(
        "OpsLensSmallV14",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9.2,
        wordWrap="CJK",
    )

    small_bold = ParagraphStyle(
        "OpsLensSmallBoldV14",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.4,
        leading=9.2,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsLensCodeV14",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.7,
        leading=9.8,
        backColor=colors.HexColor("#F8FAFC"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.35,
        borderPadding=6,
        wordWrap="CJK",
    )

    story = []

    def esc(value):
        return html.escape(_ent_text(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 3))
        story.append(Paragraph(esc(value), h_style))

    def p(value):
        value = _ent_text(value)
        if value:
            story.append(Paragraph(esc(value), body))

    def code_block(lines):
        text = "\n".join(lines or ["# Analysis Inconclusive"])
        story.append(Paragraph(esc_code(text), code_style))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.5 * cm
        widths = widths or [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc(h), small_bold) for h in headers]]
        for row in rows:
            data.append([Paragraph(esc(cell), small) for cell in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))

    logo = _ent_logo_path()

    if logo:
        try:
            logo_obj = Image(logo, width=1.05 * cm, height=1.05 * cm)
        except Exception:
            logo_obj = Paragraph("OpsLens AI", brand_style)
    else:
        logo_obj = Paragraph("OpsLens AI", brand_style)

    header_left = [
        logo_obj,
    ]

    header_right = [
        Paragraph(esc(_ent_title(report)), title_style),
        Paragraph(
            esc(
                f"Namespace: {affected.get('namespace', '-')}"
                f"    Node: {affected.get('node') or affected.get('node_name') or '-'}"
                f"    Impact: {_ent_impact_label(report)}"
                f"    Confidence: {report.get('confidence', '-')}"
            ),
            subtitle_style,
        ),
        Paragraph(
            esc(f"Generated: {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or '-'}"),
            subtitle_style,
        ),
    ]

    header_table = Table(
        [[header_left, header_right]],
        colWidths=[2.4 * cm, A4[0] - 4.9 * cm],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8))

    metadata_rows = [
        ["Service", affected.get("service") or affected.get("service_name") or "-"],
        ["Deployment", affected.get("deployment") or affected.get("deployment_name") or "-"],
        ["Pod", affected.get("pod") or affected.get("pod_name") or "-"],
        ["Impact Level", _ent_impact_label(report)],
    ]

    heading("Header / Scope")
    table(["Field", "Value"], metadata_rows, [4.2 * cm, 12.8 * cm])

    heading("Executive Summary")
    summary_rows = [[item] for item in _ent_executive_summary(report)]
    table(["Summary"], summary_rows, [A4[0] - 2.5 * cm])

    heading("Primary Incident")
    table(
        ["Field", "Details"],
        [
            ["Root Cause", _ent_root_cause(report)],
            ["Why it is the blocker", _ent_problem(report)],
        ],
        [4.3 * cm, 12.7 * cm],
    )

    heading("Related Findings")
    if related:
        table(["Resource", "Finding", "Impact / Priority"], related, [4.4 * cm, 6.1 * cm, 6.5 * cm])
    else:
        p("No separate unrelated findings were identified.")

    heading("Recommended Next Actions")
    code_block(remediation or ["# Analysis Inconclusive: no safe remediation command available."])

    heading("Validation Checks")
    code_block(verification or ["# Analysis Inconclusive: no verification command available."])

    heading("Diagnostic Summary")
    table(
        ["Type", "Resource / Agent", "Status"],
        diagnostic or [["Scope", "No diagnostic details", "Analysis Inconclusive"]],
        [4.0 * cm, 5.5 * cm, 7.5 * cm],
    )

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(report, export_format, fallback_name=fallback_name)
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens Final Generic Backend Classification and Commands v17
# =========================================================
# OpsLens Final Generic Backend Classification and Commands v17
# General solution, not scenario-specific:
# - classify from primary evidence only
# - prefer valid LLM commands
# - fallback only when LLM commands contradict the primary incident
# - used by PDF/Markdown renderers through _simple_* functions
# =========================================================

import re as _ops_v17_re


def _ops_v17_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_ops_v17_text(v) for v in value.values() if _ops_v17_text(v))
        if isinstance(value, list):
            return " ".join(_ops_v17_text(v) for v in value if _ops_v17_text(v))
        return str(value).strip()


def _ops_v17_primary_text(report):
    if not isinstance(report, dict):
        return _ops_v17_text(report).lower()

    parts = []

    for key in [
        "title",
        "incident_summary",
        "root_cause_story",
        "root_cause_hypothesis",
        "primary_incident_group",
        "primary_signal",
        "root_cause_facts",
        "affected_resources",
    ]:
        if key in report:
            parts.append(_ops_v17_text(report.get(key)))

    # Only use agent_reasoning as supporting fallback if the LLM text is too weak.
    base = " ".join(parts).lower()

    if len(base.strip()) < 80 and report.get("agent_reasoning"):
        base += " " + _ops_v17_text(report.get("agent_reasoning")).lower()

    return base


def _ops_fix_incident_type(report):
    text = _ops_v17_primary_text(report)

    # Prioritize explicit failure modes over readiness symptoms.
    if "oomkilled" in text or "outofmemory" in text or "out of memory" in text:
        return "oom"

    if (
        "imagepullbackoff" in text
        or "errimagepull" in text
        or "failed to pull image" in text
        or "invalid registry" in text
    ):
        return "image_pull"

    if (
        "missing env" in text
        or "missing environment" in text
        or "environment variable" in text
        or "keyerror" in text
    ):
        return "missing_env"

    if (
        "crashloopbackoff" in text
        or "fatal error" in text
        or "fatal:" in text
        or "traceback" in text
        or "failed during startup" in text
        or "fails during startup" in text
        or "failing to start" in text
    ):
        return "crashloop"

    if (
        "podlstmresourceanomaly" in text
        or "higher than predicted" in text
        or "higher than expected" in text
        or "predicted baseline" in text
    ):
        return "pod_cpu_anomaly"

    # Exact service/endpoint matching only; avoid generic 'selected' false positives.
    if (
        "empty endpoints" in text
        or "service has no endpoints" in text
        or "endpoint list remains empty" in text
        or ("service selector" in text and ("pod labels" in text or "endpoint" in text))
        or ("selector mismatch" in text and ("service" in text or "endpoint" in text))
    ):
        return "service_selector"

    if "targetport" in text or "target port" in text:
        return "service_targetport"

    if "readiness" in text and ("404" in text or "not-found" in text):
        return "readiness_404"

    if "readiness" in text:
        return "readiness"

    return "generic"


def _ops_v17_extract_commands(value):
    if value is None:
        return []

    if isinstance(value, dict):
        out = []
        for key in ["commands", "safe_commands", "verification_commands"]:
            items = value.get(key)
            if isinstance(items, list):
                for item in items:
                    out.extend(_ops_v17_extract_commands(item))

        for key in ["command", "strategy", "explanation", "intent"]:
            if value.get(key):
                out.extend(_ops_v17_extract_commands(value.get(key)))

        if isinstance(value.get("actions"), list):
            for action in value.get("actions"):
                out.extend(_ops_v17_extract_commands(action))

        return _ops_v17_dedupe(out)

    if isinstance(value, list):
        out = []
        for item in value:
            out.extend(_ops_v17_extract_commands(item))
        return _ops_v17_dedupe(out)

    raw = str(value)
    raw = raw.replace("\r", "\n")
    raw = raw.replace("`", "")
    raw = raw.replace("' and '", "\n")
    raw = raw.replace('" and "', "\n")
    raw = _ops_v17_re.sub(r"\s+and\s+(?=kubectl\s)", "\n", raw, flags=_ops_v17_re.I)
    raw = _ops_v17_re.sub(r"(?<!^)\s+(?=kubectl\s)", "\n", raw)

    commands = []

    for line in raw.splitlines():
        line = line.strip()

        if "kubectl " not in line:
            continue

        line = line[line.lower().find("kubectl "):]
        line = line.strip(" '\"`.,;")
        line = _ops_v17_re.sub(r"\s+", " ", line).strip()

        if line.startswith("kubectl "):
            commands.append(line)

    return _ops_v17_dedupe(commands)


def _ops_v17_dedupe(items):
    out = []
    seen = set()

    for item in items or []:
        item = str(item or "").strip()
        key = item.lower()

        if not item or key in seen:
            continue

        seen.add(key)
        out.append(item)

    return out


def _ops_v17_commands_match(commands, incident_type):
    if not commands:
        return False

    joined = "\n".join(commands).lower()

    if "delete " in joined:
        return False

    if incident_type == "service_selector":
        return (
            "get svc" in joined
            or "describe svc" in joined
            or "get endpoints" in joined
            or "show-labels" in joined
            or "edit svc" in joined
        )

    if incident_type == "service_targetport":
        return (
            "get svc" in joined
            or "describe svc" in joined
            or "targetport" in joined
            or "get endpoints" in joined
            or "edit svc" in joined
        )

    if incident_type == "crashloop":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False

        has_pod_or_logs = (
            "describe pod" in joined
            or "logs" in joined
            or "--previous" in joined
        )

        has_context = (
            "describe deployment" in joined
            or "edit deployment" in joined
            or "rollout status" in joined
        )

        return has_pod_or_logs and has_context

    if incident_type == "missing_env":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False

        return (
            "logs" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
            or "configmap" in joined
            or "secret" in joined
            or "edit deployment" in joined
        )

    if incident_type == "oom":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False

        return (
            "--previous" in joined
            or "top pod" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
        )

    if incident_type == "pod_cpu_anomaly":
        if "oom" in joined or "--previous" in joined or "readinessprobe" in joined:
            return False

        return (
            "top pod" in joined
            or "logs" in joined
            or "describe pod" in joined
            or "describe deployment" in joined
        )

    if incident_type == "image_pull":
        if "readinessprobe" in joined or "/healthz" in joined:
            return False

        return (
            "describe pod" in joined
            or "get events" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
        )

    if incident_type in {"readiness_404", "readiness"}:
        return (
            "readiness" in joined
            or "describe deployment" in joined
            or "edit deployment" in joined
            or "rollout status" in joined
            or "get endpoints" in joined
        )

    return True


def _ops_v17_fallback_fix(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    t = _ops_fix_incident_type(report)

    if t == "service_selector":
        return _ops_v17_dedupe([
            f"kubectl get svc {service} -n {ns} -o yaml" if service else "",
            f"kubectl get endpoints {service} -n {ns}" if service else "",
            f"kubectl get pods -n {ns} --show-labels",
        ])[:3]

    if t == "service_targetport":
        return _ops_v17_dedupe([
            f"kubectl get svc {service} -n {ns} -o yaml" if service else "",
            f"kubectl describe svc {service} -n {ns}" if service else "",
            f"kubectl get endpoints {service} -n {ns}" if service else "",
        ])[:3]

    if t == "crashloop":
        return _ops_v17_dedupe([
            f"kubectl describe pod {pod} -n {ns}" if pod else "",
            f"kubectl logs {pod} -n {ns} --previous" if pod else "",
            f"kubectl describe deployment {deployment} -n {ns}" if deployment else "",
        ])[:3]

    if t == "missing_env":
        return _ops_v17_dedupe([
            f"kubectl logs {pod} -n {ns} --tail=160" if pod else "",
            f"kubectl describe deployment {deployment} -n {ns}" if deployment else "",
            f"kubectl get configmap -n {ns}",
        ])[:3]

    if t == "oom":
        return _ops_v17_dedupe([
            f"kubectl describe pod {pod} -n {ns}" if pod else "",
            f"kubectl logs {pod} -n {ns} --previous" if pod else "",
            f"kubectl top pod {pod} -n {ns}" if pod else "",
        ])[:3]

    if t == "pod_cpu_anomaly":
        return _ops_v17_dedupe([
            f"kubectl top pod {pod} -n {ns}" if pod else "",
            f"kubectl describe pod {pod} -n {ns}" if pod else "",
            f"kubectl logs {pod} -n {ns} --tail=120" if pod else "",
        ])[:3]

    if t == "image_pull":
        return _ops_v17_dedupe([
            f"kubectl describe pod {pod} -n {ns}" if pod else "",
            f"kubectl get events -n {ns} --sort-by=.lastTimestamp",
            f"kubectl describe deployment {deployment} -n {ns}" if deployment else "",
        ])[:3]

    if t in {"readiness_404", "readiness"}:
        return _ops_v17_dedupe([
            f"kubectl describe deployment {deployment} -n {ns}" if deployment else "",
            f"kubectl edit deployment {deployment} -n {ns}" if deployment else "",
            f"kubectl rollout status deployment/{deployment} -n {ns}" if deployment else "",
        ])[:3]

    return _ops_v17_dedupe([
        f"kubectl describe deployment {deployment} -n {ns}" if deployment else "",
        f"kubectl describe pod {pod} -n {ns}" if pod else "",
        f"kubectl get pods -n {ns}",
    ])[:3]


def _simple_recommended_fix_commands(report):
    incident_type = _ops_fix_incident_type(report)
    llm_commands = _ops_v17_extract_commands(report.get("recommended_fix"))

    if _ops_v17_commands_match(llm_commands, incident_type):
        return llm_commands[:3]

    return _ops_v17_fallback_fix(report)[:3]


def _simple_verification_commands(report):
    affected = _simple_affected(report)

    ns = affected.get("namespace") or "default"
    pod = affected.get("pod") or affected.get("pod_name")
    deployment = affected.get("deployment") or affected.get("deployment_name")
    service = affected.get("service") or affected.get("service_name")

    t = _ops_fix_incident_type(report)

    llm_commands = _ops_v17_extract_commands(report.get("verification"))
    safe_llm = [
        cmd for cmd in llm_commands
        if cmd.lower().startswith("kubectl ") and "delete " not in cmd.lower()
    ]

    if safe_llm:
        return safe_llm[:2]

    if t == "service_selector":
        return _ops_v17_dedupe([
            f"kubectl get endpoints {service} -n {ns}" if service else "",
            f"kubectl get pods -n {ns} --show-labels",
        ])[:2]

    if t == "crashloop":
        return _ops_v17_dedupe([
            f"kubectl get pods -n {ns}",
            f"kubectl logs {pod} -n {ns} --tail=80" if pod else "",
        ])[:2]

    if t == "pod_cpu_anomaly":
        return _ops_v17_dedupe([
            f"kubectl top pod {pod} -n {ns}" if pod else "",
            f"kubectl logs {pod} -n {ns} --tail=80" if pod else "",
        ])[:2]

    if deployment:
        return _ops_v17_dedupe([
            f"kubectl rollout status deployment/{deployment} -n {ns}",
            f"kubectl get pods -n {ns}",
        ])[:2]

    return [f"kubectl get pods -n {ns}"]


def _simple_problem(report):
    affected = _simple_affected(report)

    pod = affected.get("pod") or affected.get("pod_name") or "the affected pod"
    deployment = affected.get("deployment") or affected.get("deployment_name") or "the deployment"
    service = affected.get("service") or affected.get("service_name") or "the service"

    t = _ops_fix_incident_type(report)

    if t == "crashloop":
        return (
            f"{pod} is stuck in CrashLoopBackOff because the application container fails during startup. "
            f"As a result, {deployment} cannot reach a stable Ready state."
        )

    if t == "service_selector":
        return (
            f"{service} has no ready endpoints because its selector does not match the intended backend pods."
        )

    if t == "oom":
        return (
            f"{pod} is repeatedly restarting because the container was OOMKilled. "
            f"As a result, {deployment} has no available ready replica."
        )

    if t == "pod_cpu_anomaly":
        return (
            _ops_v17_text(report.get("incident_summary"))
            or f"{pod} is consuming CPU significantly above its learned baseline."
        )

    if t in {"readiness_404", "readiness"}:
        return (
            f"{pod} is failing readiness checks. Kubernetes keeps the pod NotReady, "
            f"so {deployment} has no available ready replica."
        )

    return (
        _ops_v17_text(report.get("incident_summary"))
        or "Analysis Inconclusive."
    )

# =========================================================
# OpsLens User Friendly PDF Renderer v18
# =========================================================
# OpsLens User Friendly PDF Renderer v18
# Visual/wording only.
# Does NOT change agent logic, classification, or backend analysis.
# =========================================================

def _uf_text(value):
    try:
        return _simple_clean(value)
    except Exception:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.replace("\r", " ").replace("\n", " ").strip()
        if isinstance(value, dict):
            return " ".join(_uf_text(v) for v in value.values() if _uf_text(v))
        if isinstance(value, list):
            return " ".join(_uf_text(v) for v in value if _uf_text(v))
        return str(value).strip()


def _uf_report(report):
    try:
        return _simple_report(report)
    except Exception:
        return report if isinstance(report, dict) else {}


def _uf_affected(report):
    try:
        return _simple_affected(report)
    except Exception:
        affected = report.get("affected_resources") if isinstance(report, dict) else {}
        return affected if isinstance(affected, dict) else {}


def _uf_impact(report):
    severity = str(report.get("severity") or "unknown").lower()

    if severity == "critical":
        return "Workload unavailable"
    if severity == "warning":
        return "Degraded or anomalous"
    if severity == "info":
        return "Informational"

    try:
        label = _ops_impact_label(report)
        return label.replace("Critical - ", "").replace("Warning - ", "").replace("Info - ")
    except Exception:
        return severity or "unknown"


def _uf_title(report):
    affected = _uf_affected(report)
    workload = (
        affected.get("service")
        or affected.get("deployment")
        or affected.get("deployment_name")
        or affected.get("pod")
        or affected.get("pod_name")
        or "Kubernetes Workload"
    )
    return f"Incident Report: {workload}"


def _uf_problem(report):
    try:
        return _simple_problem(report)
    except Exception:
        return _uf_text(report.get("incident_summary")) or "Analysis Inconclusive."


def _uf_cause(report):
    try:
        return _simple_root_cause(report)
    except Exception:
        return _uf_text(report.get("root_cause_story")) or "Analysis Inconclusive."


def _uf_commands(report, mode):
    try:
        if mode == "fix":
            lines = _simple_recommended_fix_commands(report)
            limit = 3
        else:
            lines = _simple_verification_commands(report)
            limit = 2
    except Exception:
        lines = []
        limit = 3 if mode == "fix" else 2

    clean = []
    seen = set()

    for line in lines or []:
        item = str(line or "").strip()
        if not item.startswith("kubectl "):
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        clean.append(item)
        if len(clean) >= limit:
            break

    return clean


def _uf_related_rows(report):
    rows = []
    findings = report.get("additional_findings") or report.get("separate_findings") or []

    if isinstance(findings, dict):
        findings = [findings]

    if isinstance(findings, list):
        for item in findings:
            raw = _uf_text(item).lower()

            if "agentexecutionerror" in raw:
                continue
            if "deploymentnot found in collected evidence" in raw:
                continue

            if isinstance(item, dict):
                resource = _uf_text(item.get("resource") or item.get("pod_name") or item.get("deployment_name") or "scope")
                finding = _uf_text(item.get("finding") or item.get("summary") or item.get("anomaly_type") or "")
                impact = _uf_text(item.get("impact") or item.get("meaning") or "Requires follow-up.")
                priority = _uf_text(item.get("priority") or item.get("severity") or "Follow-up")

                if finding:
                    rows.append([resource, finding, f"{impact} / {priority}"])
            else:
                value = _uf_text(item)
                if value:
                    rows.append(["scope", value, "Requires follow-up / Follow-up"])

    return rows[:6]


def _uf_agent_finding(value):
    if not isinstance(value, dict):
        return "No active signal detected."

    findings = value.get("findings")
    if isinstance(findings, list) and findings:
        parts = []
        for item in findings[:2]:
            if isinstance(item, dict):
                anomaly = _uf_text(item.get("anomaly_type") or item.get("finding") or "Signal")
                summary = _uf_text(item.get("summary") or item.get("meaning") or "")
                parts.append(f"{anomaly}: {summary}" if summary else anomaly)
            else:
                parts.append(_uf_text(item))
        return " | ".join(p for p in parts if p)

    finding = _uf_text(value.get("finding"))
    event_count = value.get("event_count", 0)

    if event_count and (not finding or finding.lower() == "no active signal detected."):
        return f"{event_count} signal(s) detected."

    return finding or "No active signal detected."


def _uf_evidence_rows(report):
    affected = _uf_affected(report)
    rows = []

    if affected.get("deployment") or affected.get("deployment_name"):
        rows.append(["Deployment", affected.get("deployment") or affected.get("deployment_name"), "Affected workload"])

    if affected.get("pod") or affected.get("pod_name"):
        rows.append(["Pod", affected.get("pod") or affected.get("pod_name"), "Affected pod"])

    if affected.get("service") or affected.get("service_name"):
        rows.append(["Service", affected.get("service") or affected.get("service_name"), "In scope"])

    contributions = report.get("agent_contributions") or {}

    if isinstance(contributions, dict) and contributions:
        for agent, value in list(contributions.items())[:8]:
            if isinstance(value, dict):
                status = _uf_text(value.get("status") or "completed")
                count = value.get("event_count", 0)
                finding = _uf_agent_finding(value)
            else:
                status = "completed"
                count = 0
                finding = "No active signal detected."

            rows.append([f"Agent: {agent}", f"{status} ({count} signal(s))", finding])

        return rows[:12]

    agent_status = report.get("agent_run_status") or []
    if isinstance(agent_status, list):
        for item in agent_status[:8]:
            if not isinstance(item, dict):
                continue
            rows.append([
                f"Agent: {_uf_text(item.get('agent') or 'agent')}",
                f"{_uf_text(item.get('status') or 'completed')} ({item.get('event_count', 0)} signal(s))",
                _uf_text(item.get("finding") or "No active signal detected.")
            ])

    return rows[:12]


def _uf_logo_path():
    candidates = [
        "web/assets/logo.png",
        "web/assets/opslens-logo.png",
        "web/assets/opslens_ai_logo.png",
        "web/logo.png",
        "web/opslens-logo.png",
        "static/logo.png",
    ]

    for item in candidates:
        p = Path(item)
        if p.exists():
            return str(p)

    return None


def _opslens_export_markdown(report):
    report = _uf_report(report)
    affected = _uf_affected(report)

    fix = _uf_commands(report, "fix")
    verify = _uf_commands(report, "verify")
    related = _uf_related_rows(report)
    evidence = _uf_evidence_rows(report)

    lines = []
    lines.append(f"# {_uf_title(report)}")
    lines.append("")
    lines.append("## At a glance")
    lines.append(f"- **Namespace:** {affected.get('namespace', '-')}")
    lines.append(f"- **Node:** {affected.get('node') or affected.get('node_name') or '-'}")
    lines.append(f"- **Deployment:** {affected.get('deployment') or affected.get('deployment_name') or '-'}")
    lines.append(f"- **Pod:** {affected.get('pod') or affected.get('pod_name') or '-'}")
    lines.append(f"- **Impact:** {_uf_impact(report)}")
    lines.append(f"- **Confidence:** {report.get('confidence', '-')}")
    lines.append("")
    lines.append("## 1. What happened")
    lines.append(_uf_problem(report))
    lines.append("")
    lines.append("## 2. Why it happened")
    lines.append(_uf_cause(report))
    lines.append("")
    lines.append("## 3. Recommended next actions")
    lines.append("These are recommended SRE next actions based on the evidence, not a guaranteed final fix.")
    lines.append("")
    lines.append("```bash")
    lines.extend(fix or ["# Analysis Inconclusive: no safe recommended command available."])
    lines.append("```")
    lines.append("")
    lines.append("## 4. Validation checks")
    lines.append("```bash")
    lines.extend(verify or ["# Analysis Inconclusive: no validation command available."])
    lines.append("```")
    lines.append("")
    lines.append("## 5. Related findings")
    if related:
        lines.append("| Resource | Finding | Impact / Priority |")
        lines.append("| --- | --- | --- |")
        for row in related:
            lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    else:
        lines.append("No separate unrelated findings were identified.")
    lines.append("")
    lines.append("## 6. Technical evidence")
    lines.append("| Type | Resource / Agent | Status |")
    lines.append("| --- | --- | --- |")
    for row in evidence:
        lines.append(f"| {row[0]} | {row[1]} | {row[2]} |")
    lines.append("")

    return "\n".join(lines).strip() + "\n"


def _opslens_write_pdf(report, path):
    import html
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, Image

    report = _uf_report(report)
    affected = _uf_affected(report)

    fix = _uf_commands(report, "fix")
    verify = _uf_commands(report, "verify")
    related = _uf_related_rows(report)
    evidence = _uf_evidence_rows(report)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=1.25 * cm,
        leftMargin=1.25 * cm,
        topMargin=0.85 * cm,
        bottomMargin=0.9 * cm,
    )

    styles = getSampleStyleSheet()

    title = ParagraphStyle(
        "UFTitleV18",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16.5,
        leading=20,
        alignment=0,
        textColor=colors.HexColor("#111827"),
        spaceAfter=2,
    )

    brand = ParagraphStyle(
        "UFBrandV18",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=15,
        textColor=colors.HexColor("#DC2626"),
    )

    sub = ParagraphStyle(
        "UFSubV18",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#64748B"),
    )

    h = ParagraphStyle(
        "UFHeadingV18",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=14,
        spaceBefore=8,
        spaceAfter=5,
        textColor=colors.HexColor("#111827"),
    )

    body = ParagraphStyle(
        "UFBodyV18",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=11.4,
        spaceAfter=4,
    )

    note = ParagraphStyle(
        "UFNoteV18",
        parent=styles["BodyText"],
        fontName="Helvetica-Oblique",
        fontSize=7.7,
        leading=9.5,
        textColor=colors.HexColor("#475569"),
        spaceAfter=5,
    )

    cell = ParagraphStyle(
        "UFCellV18",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9.2,
        wordWrap="CJK",
    )

    cell_head = ParagraphStyle(
        "UFCellHeadV18",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.2,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code = ParagraphStyle(
        "UFCodeV18",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.8,
        leading=9.8,
        backColor=colors.HexColor("#F8FAFC"),
        borderColor=colors.HexColor("#CBD5E1"),
        borderWidth=0.35,
        borderPadding=6,
        wordWrap="CJK",
    )

    story = []

    def esc(value):
        return html.escape(_uf_text(value))

    def esc_code(value):
        return html.escape(str(value or "")).replace("\n", "<br/>")

    def heading(value):
        story.append(Spacer(1, 3))
        story.append(Paragraph(esc(value), h))

    def p(value, style=body):
        value = _uf_text(value)
        if value:
            story.append(Paragraph(esc(value), style))

    def code_block(lines):
        text = "\n".join(lines or ["# Analysis Inconclusive"])
        story.append(Paragraph(esc_code(text), code))
        story.append(Spacer(1, 5))

    def table(headers, rows, widths=None):
        if not rows:
            return

        usable = A4[0] - 2.5 * cm
        widths = widths or [usable / len(headers)] * len(headers)

        data = [[Paragraph(esc(x), cell_head) for x in headers]]
        for row in rows:
            data.append([Paragraph(esc(x), cell) for x in row])

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 6))

    logo = _uf_logo_path()
    if logo:
        try:
            logo_obj = Image(logo, width=1.0 * cm, height=1.0 * cm)
        except Exception:
            logo_obj = Paragraph("OpsLens AI", brand)
    else:
        logo_obj = Paragraph("OpsLens AI", brand)

    header_right = [
        Paragraph(esc(_uf_title(report)), title),
        Paragraph(
            esc(
                f"Namespace: {affected.get('namespace', '-')}"
                f"    Node: {affected.get('node') or affected.get('node_name') or '-'}"
                f"    Impact: {_uf_impact(report)}"
                f"    Confidence: {report.get('confidence', '-')}"
            ),
            sub,
        ),
        Paragraph(
            esc(f"Generated: {report.get('generated_at') or report.get('created_at') or report.get('finished_at') or '-'}"),
            sub,
        ),
    ]

    header_tbl = Table(
        [[[logo_obj], header_right]],
        colWidths=[2.2 * cm, A4[0] - 4.7 * cm],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 8))

    heading("At a glance")
    table(
        ["Field", "Value"],
        [
            ["Service", affected.get("service") or affected.get("service_name") or "-"],
            ["Deployment", affected.get("deployment") or affected.get("deployment_name") or "-"],
            ["Pod", affected.get("pod") or affected.get("pod_name") or "-"],
            ["Impact", _uf_impact(report)],
        ],
        [4.0 * cm, 13.0 * cm],
    )

    heading("1. What happened")
    p(_uf_problem(report))

    heading("2. Why it happened")
    p(_uf_cause(report))

    heading("3. Recommended next actions")
    p("These are recommended SRE next actions based on the evidence, not a guaranteed final fix.", note)
    code_block(fix or ["# Analysis Inconclusive: no safe recommended command available."])

    heading("4. Validation checks")
    code_block(verify or ["# Analysis Inconclusive: no validation command available."])

    heading("5. Related findings")
    if related:
        table(["Resource", "Finding", "Impact / Priority"], related, [4.4 * cm, 6.1 * cm, 6.5 * cm])
    else:
        p("No separate unrelated findings were identified.")

    heading("6. Technical evidence")
    table(
        ["Type", "Resource / Agent", "Status"],
        evidence or [["Scope", "No technical evidence available", "Analysis Inconclusive"]],
        [4.0 * cm, 5.5 * cm, 7.5 * cm],
    )

    doc.build(story)
    return path


def _export_payload(report: Any, export_format: str, fallback_name: str = "report") -> Path:
    return _opslens_write_export_file(report, export_format, fallback_name=fallback_name)


def _ops_make_export_file(report, export_format, fallback_name="opslens-report"):
    path = _opslens_write_export_file(report, export_format, fallback_name=fallback_name)
    return path, _opslens_export_media_type(export_format)

# =========================================================
# OpsLens Friendly Technical Evidence Labels v19
# =========================================================
# OpsLens Friendly Technical Evidence Labels v19
# Display-only change.
# Does NOT change agent logic, findings, grouping, or LLM output.
# =========================================================

def _uf_friendly_source_label(source):
    raw = str(source or "").strip()

    cleaned = (
        raw
        .replace("Agent:", "")
        .strip()
        .lower()
    )

    labels = {
        "resource_metrics": "Node Resource Check",
        "pod_lstm_metrics": "Pod Anomaly Check",
        "kubernetes_events": "Kubernetes Events Check",
        "logs": "Application Logs Check",
        "ansible_config": "Workload Configuration Check",
        "configuration": "Workload Configuration Check",
        "config": "Workload Configuration Check",
    }

    if cleaned in labels:
        return labels[cleaned]

    if raw.lower().startswith("agent:"):
        return raw.replace("Agent:", "Check:").strip()

    return raw


def _uf_friendly_status_text(value):
    text = _uf_text(value)

    if text.lower().strip() == "no active signal detected.":
        return "No issue detected by this check."

    return text


if "_uf_evidence_rows" in globals():
    _uf_original_evidence_rows = _uf_evidence_rows

    def _uf_evidence_rows(report):
        rows = _uf_original_evidence_rows(report)
        friendly = []

        for row in rows:
            if not isinstance(row, list) or len(row) < 3:
                friendly.append(row)
                continue

            first = _uf_friendly_source_label(row[0])
            second = _uf_text(row[1])
            third = _uf_friendly_status_text(row[2])

            friendly.append([first, second, third])

        return friendly


if "_ent_diagnostic_rows" in globals():
    _ent_original_diagnostic_rows = _ent_diagnostic_rows

    def _ent_diagnostic_rows(report):
        rows = _ent_original_diagnostic_rows(report)
        friendly = []

        for row in rows:
            if not isinstance(row, list) or len(row) < 3:
                friendly.append(row)
                continue

            first = _uf_friendly_source_label(row[0])
            second = _uf_text(row[1])
            third = _uf_friendly_status_text(row[2])

            friendly.append([first, second, third])

        return friendly

# =========================================================
# OpsLens final indexed scenario catalog v2
# =========================================================
# OpsLens final indexed scenario catalog v2
# Source of truth: scenarios/scenarios_index.json
# =========================================================

import json as _ops_index_json
from pathlib import Path as _OpsIndexPath


def _ops_indexed_scenario_catalog():
    index_path = _OpsIndexPath("scenarios") / "scenarios_index.json"

    if not index_path.exists():
        return {"namespaces": [], "scenarios": []}

    try:
        payload = _ops_index_json.loads(index_path.read_text(encoding="utf-8"))
    except Exception:
        return {"namespaces": [], "scenarios": []}

    scenarios = payload.get("scenarios") or []
    namespaces = payload.get("namespaces") or []

    clean = []

    for item in scenarios:
        if not isinstance(item, dict):
            continue

        manifest = item.get("manifest") or ""
        namespace = item.get("namespace") or ""

        clean.append({
            "id": item.get("id") or manifest,
            "name": item.get("id") or manifest,
            "title": item.get("title") or item.get("id") or manifest,
            "namespace": namespace,
            "namespaces": [namespace] if namespace else [],
            "manifest": manifest,
            "filename": manifest,
            "path": f"scenarios/{manifest}" if manifest else "",
            "resources": item.get("resources") or [],
            "primary_resources": item.get("primary_resources") or item.get("resources") or [],
            "related_resources": item.get("related_resources") or [],
            "focus_label": item.get("focus_label") or item.get("id") or "",
        })

    ns_names = [
        ns.get("name")
        for ns in namespaces
        if isinstance(ns, dict) and ns.get("name")
    ]

    if not ns_names:
        ns_names = sorted({item.get("namespace") for item in clean if item.get("namespace")})

    return {
        "namespaces": ns_names,
        "namespace_details": namespaces,
        "scenarios": clean,
    }


@app.get("/api/scenarios/catalog")
def ops_indexed_scenario_catalog_endpoint():
    return _ops_indexed_scenario_catalog()

# =========================================================
# OpsLens final catalog route override v3
# =========================================================
# OpsLens final catalog route override v3
# Single source of truth: scenarios/scenarios_index.json
# Removes older duplicate /api/scenarios/catalog routes, then registers one final route.
# =========================================================

import json as _ops_catalog_json_v3
from pathlib import Path as _OpsCatalogPathV3


def _ops_final_catalog_v3():
    index_path = _OpsCatalogPathV3("scenarios") / "scenarios_index.json"

    if not index_path.exists():
        return {
            "namespaces": [],
            "namespace_details": [],
            "scenarios": [],
        }

    try:
        payload = _ops_catalog_json_v3.loads(index_path.read_text(encoding="utf-8"))
    except Exception:
        return {
            "namespaces": [],
            "namespace_details": [],
            "scenarios": [],
        }

    namespace_details = payload.get("namespaces") or []
    scenarios = payload.get("scenarios") or []

    clean_scenarios = []

    for item in scenarios:
        if not isinstance(item, dict):
            continue

        namespace = item.get("namespace") or ""
        scenario_id = item.get("id") or item.get("manifest") or ""

        clean_scenarios.append({
            "id": scenario_id,
            "name": scenario_id,
            "title": item.get("title") or scenario_id,
            "namespace": namespace,
            "namespaces": [namespace] if namespace else [],
            "manifest": item.get("manifest") or "",
            "filename": item.get("manifest") or "",
            "path": f"scenarios/{item.get('manifest')}" if item.get("manifest") else "",
            "resources": item.get("resources") or [],
            "primary_resources": item.get("primary_resources") or item.get("resources") or [],
            "related_resources": item.get("related_resources") or [],
            "focus_label": item.get("focus_label") or scenario_id,
        })

    ns_names = [
        item.get("name")
        for item in namespace_details
        if isinstance(item, dict) and item.get("name")
    ]

    if not ns_names:
        ns_names = sorted({
            item.get("namespace")
            for item in clean_scenarios
            if item.get("namespace")
        })

    return {
        "namespaces": ns_names,
        "namespace_details": namespace_details,
        "scenarios": clean_scenarios,
    }


def _ops_final_catalog_endpoint_v3():
    return _ops_final_catalog_v3()


# Remove older duplicate catalog routes.
try:
    app.router.routes = [
        route for route in app.router.routes
        if getattr(route, "path", None) != "/api/scenarios/catalog"
    ]
except Exception:
    pass

app.add_api_route(
    "/api/scenarios/catalog",
    _ops_final_catalog_endpoint_v3,
    methods=["GET"],
)
