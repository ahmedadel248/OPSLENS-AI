
from __future__ import annotations

import json
import shutil
import subprocess
import threading
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from src.api.schemas import InvestigationRequest
from src.core.investigation_scope import InvestigationScope
from src.workflows.investigation_runner import run_investigation as run_opslens_investigation


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCENARIOS_DIR = PROJECT_ROOT / "scenarios"
LIVE_DIR = PROJECT_ROOT / "data" / "live"
REPORTS_DIR = PROJECT_ROOT / "data" / "final_incident_reports"

PROTECTED_NAMESPACES = {
    "default",
    "kube-system",
    "kube-public",
    "kube-node-lease",
}

JOBS: Dict[str, Dict[str, Any]] = {}

DEFAULT_STAGES = [
    ("scope", "Investigation Scope"),
    ("scenario", "Scenario Preparation"),
    ("collectors", "Collectors"),
    ("config_agent", "Config Agent"),
    ("logs_agent", "Logs Agent"),
    ("metrics_agent", "Metrics Agent + Model"),
    ("supervisor", "Supervisor Agent"),
    ("llm", "LLM Reasoning"),
    ("safety", "Command Safety Layer"),
    ("report", "Report Generated"),
]


def _new_stages() -> List[Dict[str, str]]:
    return [
        {"key": key, "label": label, "status": "pending"}
        for key, label in DEFAULT_STAGES
    ]


def _set_stage(job_id: str, key: str, status: str) -> None:
    job = JOBS.get(job_id)
    if not job:
        return

    for stage in job["stages"]:
        if stage["key"] == key:
            stage["status"] = status
            break


def _set_all_remaining(job_id: str, status: str) -> None:
    job = JOBS.get(job_id)
    if not job:
        return

    for stage in job["stages"]:
        if stage["status"] in {"pending", "running"}:
            stage["status"] = status


def _run_cmd(args: List[str], check: bool = True) -> subprocess.CompletedProcess:
    result = subprocess.run(
        args,
        cwd=PROJECT_ROOT,
        text=True,
        capture_output=True,
    )

    if check and result.returncode != 0:
        raise RuntimeError(
            "Command failed: "
            + " ".join(args)
            + "\nSTDOUT:\n"
            + result.stdout
            + "\nSTDERR:\n"
            + result.stderr
        )

    return result


def _kubectl_json(args: List[str]) -> Dict[str, Any]:
    result = _run_cmd(["kubectl", *args, "-o", "json"])
    return json.loads(result.stdout)


def list_nodes() -> List[str]:
    payload = _kubectl_json(["get", "nodes"])
    return [
        item.get("metadata", {}).get("name")
        for item in payload.get("items", [])
        if item.get("metadata", {}).get("name")
    ]


def list_namespaces() -> List[str]:
    payload = _kubectl_json(["get", "namespaces"])
    return [
        item.get("metadata", {}).get("name")
        for item in payload.get("items", [])
        if item.get("metadata", {}).get("name")
    ]


def list_scenarios() -> List[str]:
    SCENARIOS_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(
        list(SCENARIOS_DIR.glob("*.yml"))
        + list(SCENARIOS_DIR.glob("*.yaml"))
    )

    return [path.name for path in files]


def _render_scenario_for_namespace(scenario_name: str, namespace: str, job_id: str) -> Path:
    source_path = SCENARIOS_DIR / scenario_name

    if not source_path.exists():
        raise FileNotFoundError(f"Scenario not found: {scenario_name}")

    text = source_path.read_text(encoding="utf-8")

    known_namespaces = [
        "opslens-chaos",
        "opslens-demo",
        "opslens-lab",
        "opslens-test",
    ]

    for known in known_namespaces:
        text = text.replace(known, namespace)

    LIVE_DIR.mkdir(parents=True, exist_ok=True)

    rendered_path = LIVE_DIR / f"rendered_{job_id}_{scenario_name}"
    rendered_path.write_text(text, encoding="utf-8")

    return rendered_path


def _reset_namespace(namespace: str) -> None:
    if namespace in PROTECTED_NAMESPACES:
        raise ValueError(f"Refusing to reset protected namespace: {namespace}")

    _run_cmd(["kubectl", "delete", "namespace", namespace, "--ignore-not-found"], check=False)


def _ensure_namespace(namespace: str) -> None:
    _run_cmd(["kubectl", "create", "namespace", namespace], check=False)


def _apply_scenario(request: InvestigationRequest, job_id: str) -> None:
    if not request.scenario_name:
        return

    rendered_path = _render_scenario_for_namespace(
        scenario_name=request.scenario_name,
        namespace=request.namespace,
        job_id=job_id,
    )

    _run_cmd(["kubectl", "apply", "-f", str(rendered_path)])


def _progress_ticker(job_id: str, stop_event: threading.Event) -> None:
    """
    V1 progress animation support.

    The real pipeline is currently synchronous, so this ticker gives the UI
    step-by-step movement while the backend investigation runs. When the
    real result returns, the worker marks all stages done/failed.
    """

    ordered = [
        "collectors",
        "config_agent",
        "logs_agent",
        "metrics_agent",
        "supervisor",
        "llm",
        "safety",
    ]

    for key in ordered:
        if stop_event.is_set():
            return

        _set_stage(job_id, key, "running")
        time.sleep(1.5)

        if stop_event.is_set():
            return

        _set_stage(job_id, key, "done")


def start_investigation_job(request: InvestigationRequest) -> Dict[str, Any]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    job_id = f"{timestamp}_{uuid4().hex[:8]}"

    JOBS[job_id] = {
        "job_id": job_id,
        "status": "queued",
        "request": request.model_dump(),
        "stages": _new_stages(),
        "report": None,
        "error": None,
        "created_at": datetime.now().isoformat(),
        "finished_at": None,
    }

    thread = threading.Thread(
        target=_run_job,
        args=(job_id, request),
        daemon=True,
    )
    thread.start()

    return get_job(job_id)


def _run_job(job_id: str, request: InvestigationRequest) -> None:
    stop_ticker = threading.Event()

    try:
        JOBS[job_id]["status"] = "running"

        _set_stage(job_id, "scope", "running")
        scope = InvestigationScope(
            node_name=request.node_name,
            namespace=request.namespace,
        )
        scope.validate()
        _set_stage(job_id, "scope", "done")

        _set_stage(job_id, "scenario", "running")

        if request.reset_namespace:
            _reset_namespace(request.namespace)

        _ensure_namespace(request.namespace)

        if request.apply_scenario:
            _apply_scenario(request, job_id)
            if request.wait_seconds:
                time.sleep(request.wait_seconds)

        _set_stage(job_id, "scenario", "done")

        ticker = threading.Thread(
            target=_progress_ticker,
            args=(job_id, stop_ticker),
            daemon=True,
        )
        ticker.start()

        # Product mode: never seed demo metrics.
        try:
            request.demo_seed_metrics = False
        except Exception:
            pass

        final_report = run_opslens_investigation(
            scope=scope,
            demo_seed_metrics=False,
        )

        # Guard report before it is stored in the job/result cache.
        try:
            final_report = _opslens_guard_payload(
                final_report,
                namespace=request.namespace,
                node=request.node_name,
            )
        except Exception:
            pass

        final_report = sanitize_report(final_report, namespace=request.namespace, node=request.node_name)

        stop_ticker.set()
        _set_all_remaining(job_id, "done")

        finished_at = datetime.now().isoformat()

        final_report = dict(final_report)
        final_report["job_id"] = job_id
        final_report["__opslens_job_id"] = job_id
        final_report["created_at"] = JOBS[job_id].get("created_at") or final_report.get("created_at") or finished_at
        final_report["finished_at"] = finished_at
        final_report["source_scenario"] = request.scenario_name

        try:
            api_history_store.save_report(final_report, record_id=job_id)
        except Exception as save_exc:
            JOBS[job_id]["history_save_error"] = f"{type(save_exc).__name__}: {save_exc}"

        JOBS[job_id]["status"] = "completed"
        JOBS[job_id]["report"] = final_report
        JOBS[job_id]["finished_at"] = finished_at

    except Exception as exc:
        stop_ticker.set()
        _set_all_remaining(job_id, "failed")

        JOBS[job_id]["status"] = "failed"
        JOBS[job_id]["error"] = f"{type(exc).__name__}: {exc}"
        JOBS[job_id]["traceback"] = traceback.format_exc()
        JOBS[job_id]["finished_at"] = datetime.now().isoformat()


def get_job(job_id: str) -> Dict[str, Any]:
    if job_id not in JOBS:
        raise KeyError(f"Investigation job not found: {job_id}")

    return JOBS[job_id]


def get_latest_job() -> Optional[Dict[str, Any]]:
    if not JOBS:
        return None

    return list(JOBS.values())[-1]


def get_report_path(job_id: str, kind: str = "markdown") -> Path:
    job = get_job(job_id)
    report = job.get("report") or {}

    if kind == "json":
        path_value = report.get("json_report_path")
    else:
        path_value = report.get("markdown_report_path")

    if not path_value:
        raise FileNotFoundError(f"No {kind} report path found for job {job_id}")

    path = PROJECT_ROOT / path_value

    if not path.exists():
        raise FileNotFoundError(f"Report file does not exist: {path}")

    return path


def get_latest_report_path(kind: str = "markdown") -> Path:
    completed_jobs = [
        job for job in JOBS.values()
        if job.get("status") == "completed" and job.get("report")
    ]

    if completed_jobs:
        return get_report_path(completed_jobs[-1]["job_id"], kind=kind)

    pattern = "*.json" if kind == "json" else "*.md"
    files = sorted(REPORTS_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)

    if not files:
        raise FileNotFoundError("No reports found.")

    return files[0]


# =========================================================
# OpsLens UI V2 additions:
# - node-scoped namespace filtering
# - PDF / Excel / Markdown / JSON report exports
# =========================================================

EXPORTS_DIR = REPORTS_DIR / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def list_namespaces_for_node(node_name: str) -> List[str]:
    """
    Kubernetes namespaces are cluster-scoped, not node-scoped.

    This function returns namespaces that currently have at least one Pod
    scheduled on the selected node. This matches the UI behavior:
    user selects node -> UI shows relevant namespaces for that node.
    """

    if not node_name:
        return list_namespaces()

    result = _run_cmd(
        [
            "kubectl",
            "get",
            "pods",
            "-A",
            "--field-selector",
            f"spec.nodeName={node_name}",
            "-o",
            "json",
        ]
    )

    payload = json.loads(result.stdout)

    namespaces = sorted(
        {
            item.get("metadata", {}).get("namespace")
            for item in payload.get("items", [])
            if item.get("metadata", {}).get("namespace")
        }
    )

    return namespaces


def _safe_filename(value: str) -> str:
    allowed = []

    for char in str(value or "opslens_report"):
        if char.isalnum() or char in {"-", "_", "."}:
            allowed.append(char)
        elif char.isspace():
            allowed.append("_")

    result = "".join(allowed).strip("_")
    return result or "opslens_report"


def _report_base_name(report: Dict[str, Any], fallback: str = "opslens_report") -> str:
    affected = report.get("affected_resources") or {}

    service_name = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or fallback
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _safe_filename(f"{service_name}_{timestamp}")[:140]


    service_name = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or fallback
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _safe_filename(f"{service_name}_{timestamp}")[:140]


    service_name = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or fallback
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _safe_filename(f"{service_name}_{timestamp}")[:140]


def _latest_json_report_path() -> Path:
    files = sorted(REPORTS_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)

    if not files:
        raise FileNotFoundError("No JSON reports found.")

    return files[0]


def _load_report_from_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _report_from_job(job_id: str) -> Dict[str, Any]:
    job = get_job(job_id)
    report = job.get("report")

    if not report:
        raise FileNotFoundError(f"No report available for job {job_id}")

    return report


def _paragraph_text(value: Any) -> str:
    import html

    text = str(value or "")
    return html.escape(text).replace("\n", "<br/>")


def _write_pdf_report(report: Dict[str, Any], output_path: Path) -> Path:
    import html
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    page_size = landscape(A4)
    page_width, _ = page_size
    margin = 1.15 * cm
    usable_width = page_width - (2 * margin)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=page_size,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OpsTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
        alignment=TA_LEFT,
    )

    section_style = ParagraphStyle(
        "OpsSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#111827"),
        spaceBefore=10,
        spaceAfter=8,
    )

    body_style = ParagraphStyle(
        "OpsBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6,
    )

    small_style = ParagraphStyle(
        "OpsSmall",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.6,
        leading=9.4,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
    )

    header_style = ParagraphStyle(
        "OpsHeader",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.4,
        textColor=colors.white,
        wordWrap="CJK",
    )

    code_style = ParagraphStyle(
        "OpsCode",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=7.3,
        leading=9,
        textColor=colors.HexColor("#E5E7EB"),
        backColor=colors.HexColor("#111827"),
        borderPadding=6,
        wordWrap="CJK",
    )

    def clean(value: Any) -> str:
        return html.escape(str(value or "")).replace("\n", "<br/>")
    def para(value: Any, style=body_style):
        story.append(Paragraph(clean(value), style))

    def heading(value: str):
        story.append(Spacer(1, 4))
        story.append(Paragraph(clean(value), section_style))

    def make_table(headers, rows, widths=None):
        if not rows:
            para("No data available.")
            return

        if widths is None:
            widths = [usable_width / len(headers)] * len(headers)

        data = [[Paragraph(clean(cell), header_style) for cell in headers]]

        for row in rows:
            data.append([Paragraph(clean(cell), small_style) for cell in row])

        table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 8))

    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}

    story = []

    story.append(Paragraph(clean(report.get("title", "OpsLens Incident Report")), title_style))

    summary_rows = [
        ["Severity", report.get("severity", "unknown")],
        ["Confidence", report.get("confidence", "unknown")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node", "")],
        ["Service", affected.get("service", "") or affected.get("service_name", "")],
        ["Deployment", affected.get("deployment", "")],
    ]
    make_table(["Field", "Value"], summary_rows, [usable_width * 0.22, usable_width * 0.78])

    heading("Incident Summary")
    para(report.get("incident_summary", ""))

    heading("Primary Root Cause")
    para(report.get("root_cause_story", ""))

    heading("Incident Story")
    story_rows = [
        ["Problem", report.get("incident_summary", "")],
        ["Evidence", ((report.get("agent_reasoning") or [{}])[0]).get("finding", "") if report.get("agent_reasoning") else ""],
        ["Root Cause", report.get("root_cause_story", "")],
        ["Recommended Fix", fix.get("strategy", "")],
        ["Verification", verification.get("intent", "")],
    ]
    make_table(["Step", "Details"], story_rows, [usable_width * 0.18, usable_width * 0.82])

    heading("Evidence Trail")
    evidence_rows = [
        [row.get("agent", ""), row.get("finding", ""), row.get("meaning", "")]
        for row in report.get("agent_reasoning", []) or []
    ]
    make_table(
        ["Agent", "Finding", "Meaning"],
        evidence_rows,
        [usable_width * 0.20, usable_width * 0.38, usable_width * 0.42],
    )

    heading("Additional Findings")
    additional_rows = [
        [
            row.get("resource", ""),
            row.get("finding", ""),
            row.get("impact", ""),
            row.get("priority", ""),
        ]
        for row in report.get("additional_findings", []) or []
    ]
    make_table(
        ["Resource", "Finding", "Impact", "Priority"],
        additional_rows,
        [usable_width * 0.23, usable_width * 0.30, usable_width * 0.34, usable_width * 0.13],
    )

    story.append(PageBreak())

    heading("Recommended Fix")
    para(fix.get("strategy", ""))

    action_rows = [
        [
            action.get("action_type", ""),
            action.get("target_kind", ""),
            action.get("reason", ""),
            action.get("risk", ""),
        ]
        for action in fix.get("actions", []) or []
    ]
    make_table(
        ["Action", "Target", "Reason", "Risk"],
        action_rows,
        [usable_width * 0.24, usable_width * 0.18, usable_width * 0.45, usable_width * 0.13],
    )

    heading("Safe Commands")
    commands = fix.get("commands", []) or []

    if commands:
        for index, command in enumerate(commands, start=1):
            story.append(Paragraph(clean(f"Command {index}"), body_style))
            story.append(Paragraph(clean(command), code_style))
            story.append(Spacer(1, 6))
    else:
        para("No commands available.")

    heading("Verification Plan")
    para(verification.get("intent", ""))

    verification_commands = verification.get("commands", []) or []

    if verification_commands:
        for index, command in enumerate(verification_commands, start=1):
            story.append(Paragraph(clean(f"Verification {index}"), body_style))
            story.append(Paragraph(clean(command), code_style))
            story.append(Spacer(1, 6))

    doc.build(story)
    return output_path



def _write_excel_report(report: Dict[str, Any], output_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 28

    def write_rows(ws, headers, rows):
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws)

    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}
    verification = report.get("verification") or {}
    facts = report.get("root_cause_facts") or {}

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    summary_rows = [
        ["Title", report.get("title", "")],
        ["Severity", report.get("severity", "")],
        ["Confidence", report.get("confidence", "")],
        ["Namespace", affected.get("namespace", "")],
        ["Node", affected.get("node", "")],
        ["Service", affected.get("service", "")],
        ["Deployment", affected.get("deployment", "")],
        ["Incident Summary", report.get("incident_summary", "")],
        ["Primary Root Cause", report.get("root_cause_story", "")],
        ["Fix Strategy", fix.get("strategy", "")],
        ["Verification Intent", verification.get("intent", "")],
    ]
    for row in summary_rows:
        ws.append(row)
    style_sheet(ws)
    ws["A1"].font = title_font

    ws = wb.create_sheet("Evidence Trail")
    write_rows(
        ws,
        ["Agent", "Finding", "Meaning"],
        [
            [row.get("agent", ""), row.get("finding", ""), row.get("meaning", "")]
            for row in report.get("agent_reasoning", []) or []
        ],
    )

    ws = wb.create_sheet("Additional Findings")
    write_rows(
        ws,
        ["Resource", "Finding", "Impact", "Priority"],
        [
            [
                row.get("resource", ""),
                row.get("finding", ""),
                row.get("impact", ""),
                row.get("priority", ""),
            ]
            for row in report.get("additional_findings", []) or []
        ],
    )

    ws = wb.create_sheet("Actions")
    write_rows(
        ws,
        ["Action Type", "Target Kind", "Reason", "Risk"],
        [
            [
                row.get("action_type", ""),
                row.get("target_kind", ""),
                row.get("reason", ""),
                row.get("risk", ""),
            ]
            for row in fix.get("actions", []) or []
        ],
    )

    ws = wb.create_sheet("Safe Commands")
    write_rows(
        ws,
        ["Type", "Command"],
        [["Remediation", command] for command in fix.get("commands", []) or []]
        + [["Verification", command] for command in verification.get("commands", []) or []],
    )

    ws = wb.create_sheet("Root Cause Facts")
    write_rows(
        ws,
        ["Fact", "Value"],
        [[key, json.dumps(value, ensure_ascii=False)] for key, value in facts.items()],
    )

    wb.save(output_path)
    return output_path



def _copy_export_source(source_path: Path, output_path: Path) -> Path:
    import shutil

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source_path, output_path)
    return output_path


def _export_report(report: Dict[str, Any], export_format: str, fallback_name: str) -> Path:
    fmt = export_format.lower().strip()
    base = _report_base_name(report, fallback=fallback_name)

    if fmt in {"md", "markdown"}:
        md_path_value = report.get("markdown_report_path")
        if not md_path_value:
            raise FileNotFoundError("Markdown report path is missing.")

        source_path = PROJECT_ROOT / md_path_value
        output_path = EXPORTS_DIR / f"{base}.md"
        return _copy_export_source(source_path, output_path)

    if fmt == "json":
        json_path_value = report.get("json_report_path")
        if not json_path_value:
            raise FileNotFoundError("JSON report path is missing.")

        source_path = PROJECT_ROOT / json_path_value
        output_path = EXPORTS_DIR / f"{base}.json"
        return _copy_export_source(source_path, output_path)

    if fmt == "pdf":
        output_path = EXPORTS_DIR / f"{base}.pdf"
        return _write_pdf_report(report, output_path)

    if fmt in {"xlsx", "excel"}:
        output_path = EXPORTS_DIR / f"{base}.xlsx"
        return _write_excel_report(report, output_path)

    raise ValueError(f"Unsupported report format: {export_format}")


def get_report_export_path(job_id: str, export_format: str) -> Path:
    report = _report_from_job(job_id)
    return _export_report(report, export_format, fallback_name=job_id)


def get_latest_report_export_path(export_format: str) -> Path:
    completed_jobs = [
        job for job in JOBS.values()
        if job.get("status") == "completed" and job.get("report")
    ]

    if completed_jobs:
        latest = completed_jobs[-1]
        return get_report_export_path(latest["job_id"], export_format)

    json_path = _latest_json_report_path()
    report = _load_report_from_json(json_path)
    return _export_report(report, export_format, fallback_name=json_path.stem)


# =========================================================
# Feedback support
# =========================================================

FEEDBACK_DIR = PROJECT_ROOT / "data" / "feedback"
FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)


def save_feedback(payload: Dict[str, Any]) -> Dict[str, Any]:
    feedback_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}"

    record = {
        "feedback_id": feedback_id,
        "created_at": datetime.now().isoformat(),
        **payload,
    }

    output_path = FEEDBACK_DIR / f"{feedback_id}.json"
    output_path.write_text(json.dumps(record, indent=2, ensure_ascii=False), encoding="utf-8")

    return {
        "status": "saved",
        "feedback_id": feedback_id,
    }


# =========================================================
# Persistent investigation history
# =========================================================

def _history_record_from_report_path(path: Path) -> Dict[str, Any]:
    try:
        report = _load_report_from_json(path)
    except Exception as exc:
        return {
            "record_id": path.stem,
            "title": path.stem,
            "status": "unreadable",
            "error": str(exc),
            "created_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
            "json_report_path": str(path.relative_to(PROJECT_ROOT)),
        }

    affected = report.get("affected_resources") or {}
    fix = report.get("recommended_fix") or {}

    service_name = (
        affected.get("service")
        or affected.get("service_name")
        or affected.get("workload")
        or affected.get("deployment")
        or affected.get("pod")
        or "unknown-service"
    )

    created_at = report.get("created_at") or report.get("timestamp")
    if not created_at:
        created_at = datetime.fromtimestamp(path.stat().st_mtime).isoformat()

    return {
        "record_id": path.stem,
        "title": report.get("title") or "OpsLens Investigation",
        "status": "completed",
        "created_at": created_at,
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        "service": service_name,
        "namespace": affected.get("namespace", ""),
        "node": affected.get("node", ""),
        "severity": report.get("severity", "unknown"),
        "confidence": report.get("confidence", "unknown"),
        "summary": report.get("incident_summary", ""),
        "root_cause": report.get("root_cause_story", ""),
        "fix_strategy": fix.get("strategy", ""),
        "json_report_path": str(path.relative_to(PROJECT_ROOT)),
        "markdown_report_path": report.get("markdown_report_path", ""),
    }


def list_investigation_history(limit: int = 50) -> List[Dict[str, Any]]:
    if not REPORTS_DIR.exists():
        return []

    paths = sorted(
        REPORTS_DIR.glob("*.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    return [_history_record_from_report_path(path) for path in paths[:limit]]


def _history_report_path(record_id: str) -> Path:
    safe_id = _safe_filename(record_id)
    path = REPORTS_DIR / f"{safe_id}.json"

    if not path.exists():
        raise FileNotFoundError(f"Investigation history record not found: {record_id}")

    return path


def get_history_report(record_id: str) -> Dict[str, Any]:
    return _load_report_from_json(_history_report_path(record_id))


def get_history_export_path(record_id: str, export_format: str) -> Path:
    report = get_history_report(record_id)
    return _export_report(report, export_format, fallback_name=record_id)


# =========================================================
# SQLite-backed scenarios and investigation history
# =========================================================



def list_scenarios() -> List[str]:
    return api_persistence.list_scenarios()


def list_scenario_details() -> Dict[str, Any]:
    return api_persistence.list_scenario_details()


def list_investigation_history(limit: int = 50) -> List[Dict[str, Any]]:
    return api_persistence.list_reports(limit=limit)


def get_history_report(record_id: str) -> Dict[str, Any]:
    return api_persistence.get_report(record_id)


def get_history_export_path(record_id: str, export_format: str) -> Path:
    report = api_persistence.get_report(record_id)
    return _export_report(report, export_format, fallback_name=record_id)


def save_report_to_database(report: Dict[str, Any], record_id: Optional[str] = None) -> Dict[str, Any]:
    return api_persistence.save_report_object(report, record_id=record_id)


# =========================================================
# Robust SQLite history services override
# =========================================================



def save_report_to_database(report: Dict[str, Any], record_id: Optional[str] = None) -> Dict[str, Any]:
    return api_persistence.save_report_object(report, record_id=record_id)


def list_investigation_history(limit: int = 50) -> List[Dict[str, Any]]:
    return api_persistence.list_reports(limit=limit)


def get_history_report(record_id: str) -> Dict[str, Any]:
    return api_persistence.get_report(record_id)


def get_history_export_path(record_id: str, export_format: str) -> Path:
    report = api_persistence.get_report(record_id)
    return _export_report(report, export_format, fallback_name=record_id)


def debug_history_state() -> Dict[str, Any]:
    return api_persistence.debug_history_state()


# =========================================================
# Final clean SQLite-backed API service overrides
# Source of truth: src/api/history_store.py
# =========================================================

from src.api import history_store as api_history_store
from src.core.report_safety import sanitize_report


def list_scenarios():
    return list(api_history_store.scenario_details().keys())


def list_scenario_details():
    return api_history_store.scenario_details()


def save_report_to_database(report, record_id=None):
    return api_history_store.save_report(report, record_id=record_id)


def list_investigation_history(limit=50):
    return api_history_store.list_reports(limit=limit)


def get_history_report(record_id):
    return api_history_store.get_report(record_id)


def get_history_export_path(record_id, export_format):
    report = api_history_store.get_report(record_id)
    return _export_report(report, export_format, fallback_name=record_id)


def debug_history_state():
    return api_history_store.debug_state()



# =========================================================
# Product report evidence guardrails
# No evidence => no incident. No demo resource pressure.
# =========================================================

def _opslens_text(value):
    try:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        if isinstance(value, (list, tuple)):
            return " ".join(_opslens_text(v) for v in value)
        if isinstance(value, dict):
            return " ".join(f"{k} {_opslens_text(v)}" for k, v in value.items())
        return str(value)
    except Exception:
        return ""


def _opslens_is_empty_evidence(value):
    text = _opslens_text(value).strip().lower()

    if not text:
        return True

    empty_markers = [
        "no data available",
        "no evidence available",
        "not available",
        "none",
        "[]",
        "{}",
    ]

    return text in empty_markers or all(marker in text for marker in ["no", "data", "available"])


def _opslens_namespace_from_payload(payload, fallback=None):
    if fallback:
        return fallback

    if isinstance(payload, dict):
        for key in ("namespace", "selected_namespace"):
            if payload.get(key):
                return payload.get(key)

        request = payload.get("request") or payload.get("scope") or {}
        if isinstance(request, dict):
            return request.get("namespace") or request.get("selected_namespace")

        report = payload.get("report") or payload.get("result") or {}
        if isinstance(report, dict):
            affected = report.get("affected_resources") or {}
            if isinstance(affected, dict):
                return affected.get("namespace")

    return None


def _opslens_node_from_payload(payload, fallback=None):
    if fallback:
        return fallback

    if isinstance(payload, dict):
        for key in ("node", "node_name", "selected_node"):
            if payload.get(key):
                return payload.get(key)

        request = payload.get("request") or payload.get("scope") or {}
        if isinstance(request, dict):
            return request.get("node_name") or request.get("node")

        report = payload.get("report") or payload.get("result") or {}
        if isinstance(report, dict):
            affected = report.get("affected_resources") or {}
            if isinstance(affected, dict):
                return affected.get("node") or affected.get("node_name")

    return None


def _opslens_fix_namespace_in_text(value, namespace):
    if not isinstance(value, str):
        return value

    if namespace and namespace != "default":
        value = value.replace("-n default", f"-n {namespace}")
        value = value.replace("--namespace default", f"--namespace {namespace}")

    return value


def _opslens_fix_commands(obj, namespace):
    if isinstance(obj, dict):
        for key, value in list(obj.items()):
            if isinstance(value, str):
                obj[key] = _opslens_fix_namespace_in_text(value, namespace)
            else:
                _opslens_fix_commands(value, namespace)

    elif isinstance(obj, list):
        for i, value in enumerate(obj):
            if isinstance(value, str):
                obj[i] = _opslens_fix_namespace_in_text(value, namespace)
            else:
                _opslens_fix_commands(value, namespace)

    return obj


def _opslens_report_looks_fake_resource_pressure(report):
    text = _opslens_text(report).lower()

    fake_markers = [
        "cpu utilization at 95",
        "cpu utilization at 95%",
        "memory utilization at 90",
        "memory utilization at 90%",
        "critical resource exhaustion",
        "critical resource pressure",
        "affected-deployment",
        "affected-service",
    ]

    return any(marker in text for marker in fake_markers)


def _opslens_report_has_empty_evidence(report):
    """
    Treat structured OpsLens signals as real evidence.
    Runtime incidents often arrive through agent_reasoning / primary_signal /
    incident_groups instead of evidence_trail.
    """
    evidence = (
        report.get("evidence_trail")
        or report.get("evidence")
        or report.get("timeline")
        or report.get("investigation_evidence")
        or report.get("agent_reasoning")
        or report.get("important_evidence")
        or report.get("primary_signal")
        or report.get("primary_incident_group")
        or report.get("incident_groups")
        or report.get("supporting_signals")
        or report.get("root_cause_facts")
    )

    additional = (
        report.get("additional_findings")
        or report.get("additional_issues")
        or report.get("secondary_findings")
        or report.get("separate_findings")
        or report.get("unclassified_findings")
    )

    return _opslens_is_empty_evidence(evidence) and _opslens_is_empty_evidence(additional)


def _opslens_make_healthy_report(namespace=None, node=None):
    namespace = namespace or "selected namespace"
    node = node or "selected node"

    return {
        "title": "No Active Incident Detected",
        "status": "healthy",
        "severity": "none",
        "confidence": "high",
        "incident_summary": (
            f"No active incident was detected in namespace '{namespace}' on node '{node}'. "
            "OpsLens did not find Kubernetes failure evidence in the selected scope."
        ),
        "affected_resources": {
            "namespace": namespace,
            "node": node,
        },
        "evidence_trail": [
            f"No failing pods detected in namespace '{namespace}'.",
            f"No Service endpoint failure evidence detected in namespace '{namespace}'.",
            "No supported warning signal was strong enough to classify this as an incident.",
        ],
        "additional_findings": [
            "No additional findings detected."
        ],
        "root_cause_story": (
            "No root cause was identified because no active failure condition was found "
            "in the collected Kubernetes evidence."
        ),
        "recommended_fix": {
            "strategy": "No remediation required. Continue monitoring the selected scope.",
            "safe_commands": [
                {
                    "title": "Check pods",
                    "command": f"kubectl get pods -n {namespace} --show-labels"
                },
                {
                    "title": "Check services and endpoints",
                    "command": f"kubectl get svc,endpoints -n {namespace}"
                },
                {
                    "title": "Check warning events",
                    "command": f"kubectl get events -n {namespace} --sort-by=.lastTimestamp"
                }
            ],
            "verification_plan": "Confirm that pods are Running/Ready and Services have endpoints.",
            "verification_commands": [
                {
                    "title": "Verify namespace workload state",
                    "command": f"kubectl get all -n {namespace}"
                }
            ]
        },
        "verification": [
            f"Run kubectl get all -n {namespace} and confirm there are no failing resources."
        ],
    }


def _opslens_guard_report(report, namespace=None, node=None):
    if not isinstance(report, dict):
        return report

    namespace = namespace or _opslens_namespace_from_payload(report) or "selected namespace"
    node = node or _opslens_node_from_payload(report) or "selected node"

    _opslens_fix_commands(report, namespace)

    # In product mode, fake/demo resource-pressure summaries are invalid.
    # If a report contains placeholders like affected-service/affected-deployment
    # or the old CPU95/Mem90 demo pattern, convert it to a Healthy report.
    if _opslens_report_looks_fake_resource_pressure(report):
        return _opslens_make_healthy_report(namespace=namespace, node=node)

    if _opslens_report_has_empty_evidence(report):
        summary_text = _opslens_text(report.get("incident_summary", "")).lower()
        if any(word in summary_text for word in ["critical", "incident", "failure", "unavailable", "degradation"]):
            return _opslens_make_healthy_report(namespace=namespace, node=node)

    return report


def _opslens_guard_payload(payload, namespace=None, node=None):
    if not isinstance(payload, dict):
        return payload

    namespace = _opslens_namespace_from_payload(payload, namespace)
    node = _opslens_node_from_payload(payload, node)

    # Direct report
    if "incident_summary" in payload or "recommended_fix" in payload or "evidence_trail" in payload:
        return _opslens_guard_report(payload, namespace=namespace, node=node)

    # Wrapped report
    for key in ("report", "result", "final_report", "incident_report"):
        if isinstance(payload.get(key), dict):
            payload[key] = _opslens_guard_report(payload[key], namespace=namespace, node=node)

    _opslens_fix_commands(payload, namespace)
    return payload


try:
    _opslens_original_start_investigation = start_investigation
except NameError:
    _opslens_original_start_investigation = None

try:
    _opslens_original_get_investigation = get_investigation
except NameError:
    _opslens_original_get_investigation = None

try:
    _opslens_original_get_latest_investigation = get_latest_investigation
except NameError:
    _opslens_original_get_latest_investigation = None


if _opslens_original_start_investigation is not None:
    def start_investigation(request):
        # Hard-disable demo metrics at API service boundary.
        try:
            request.demo_seed_metrics = False
        except Exception:
            if isinstance(request, dict):
                request["demo_seed_metrics"] = False

        result = _opslens_original_start_investigation(request)
        return _opslens_guard_payload(result)


if _opslens_original_get_investigation is not None:
    def get_investigation(job_id):
        result = _opslens_original_get_investigation(job_id)
        return _opslens_guard_payload(result)


if _opslens_original_get_latest_investigation is not None:
    def get_latest_investigation():
        result = _opslens_original_get_latest_investigation()
        return _opslens_guard_payload(result)

# =========================================================
# OpsLens final job persistence v3
# =========================================================
# OpsLens final job persistence v3
# Keeps active investigation visible after browser refresh.
# =========================================================

JOBS_DIR = LIVE_DIR / "jobs"
JOBS_DIR.mkdir(parents=True, exist_ok=True)


def _opslens_job_path(job_id: str) -> Path:
    safe_id = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in str(job_id))
    return JOBS_DIR / f"{safe_id}.json"


def _opslens_persist_job(job_id: str) -> None:
    try:
        job = JOBS.get(job_id)
        if not job:
            return

        payload = dict(job)
        payload["persisted_at"] = datetime.now().isoformat()

        _opslens_job_path(job_id).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
    except Exception:
        pass


def _opslens_load_job_snapshot(job_id: str) -> Optional[Dict[str, Any]]:
    path = _opslens_job_path(job_id)

    if not path.exists():
        return None

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _opslens_latest_job_snapshot() -> Optional[Dict[str, Any]]:
    try:
        files = sorted(JOBS_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    except Exception:
        return None

    for path in files:
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue

    return None


_opslens_original_set_stage = _set_stage
_opslens_original_set_all_remaining = _set_all_remaining
_opslens_original_run_job = _run_job
_opslens_original_start_job = start_investigation_job
_opslens_original_get_job = get_job
_opslens_original_get_latest_job = get_latest_job


def _set_stage(job_id: str, key: str, status: str) -> None:
    _opslens_original_set_stage(job_id, key, status)
    _opslens_persist_job(job_id)


def _set_all_remaining(job_id: str, status: str) -> None:
    _opslens_original_set_all_remaining(job_id, status)
    _opslens_persist_job(job_id)


def _run_job(job_id: str, request: InvestigationRequest) -> None:
    _opslens_persist_job(job_id)

    try:
        _opslens_original_run_job(job_id, request)
    finally:
        _opslens_persist_job(job_id)


def start_investigation_job(request: InvestigationRequest) -> Dict[str, Any]:
    result = _opslens_original_start_job(request)

    try:
        job_id = result.get("job_id")
        if job_id:
            _opslens_persist_job(job_id)
    except Exception:
        pass

    return result


def get_job(job_id: str) -> Dict[str, Any]:
    if job_id in JOBS:
        return JOBS[job_id]

    snapshot = _opslens_load_job_snapshot(job_id)

    if snapshot:
        status = str(snapshot.get("status", "")).lower()

        if status in {"queued", "running"}:
            snapshot["status"] = "failed"
            snapshot["error"] = (
                "The browser was refreshed after the job snapshot was saved, "
                "but the backend worker is no longer active. Start the investigation again."
            )
            snapshot["finished_at"] = snapshot.get("finished_at") or datetime.now().isoformat()

        return snapshot

    return _opslens_original_get_job(job_id)


def get_latest_job() -> Optional[Dict[str, Any]]:
    if JOBS:
        return list(JOBS.values())[-1]

    return _opslens_latest_job_snapshot()
